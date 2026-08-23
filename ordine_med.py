"""ordine_med.py — "Ordine MED" extraction: turns a raw ORDINE sheet into
the same "Ordine da creare" table Delta's Excel macro produces, plus two
things it didn't have: a raw-yarn availability/shortage check (against
Filato Disponibile stock) and automatic Consegna dates assigned by walking
the order top-to-bottom and queuing each color onto its machine using the
same "2 colors/day, Friday off" Copertura logic the rest of the app uses.

Reverse-engineered from the Power Query embedded in the reference
ORDINE_MED-MACRO.xlsm the user sent (see Section1.m: queries 'Ordine',
'PT-GG', 'Filato') and verified field-by-field against its real sheets
(ORDINE, Ordine da creare, Filato Disponibile, تحضير خيط خام).

Not replicated (left as follow-ups, flagged rather than guessed at):
- BAGNO PREPOSTO auto-numbering (a separate, self-contained sequencing
  rule keyed off a starting Bagno letter+number that isn't available yet).
- The exact Prezzo lookup key the macro uses (Colore + a category "type"
  derived from an external Category Colori.xlsx this session doesn't have
  visibility into) -- Prezzo here reuses the same Articolo+Codice Listini
  lookup Biglietti already uses, which is close but not proven identical.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

from biglietti_exporter import _clean, _get, _key, _number, _read_sheet_rows  # noqa: F401 -- shared normalization helpers


# ---------------------------------------------------------------------------
# GRUPPO MACCHINA lookup -- Rocche-per-ABBIN-group total -> the 3300-series
# code this sheet uses (distinct from Biglietti's own M/C machine numbers;
# both derive from the same Rocche-count idea but are separate ERP fields).
# ---------------------------------------------------------------------------
GRUPPO_MACCHINA_TABLE: dict[int, int] = {
    6: 3301, 7: 3301, 24: 3310, 32: 3306, 56: 3302,
    72: 3307, 128: 3303, 192: 3308, 384: 3304, 672: 3309,
}


def _polmoni_multiplier(polmoni_text: str) -> int:
    """'2 POLMONI' -> 8 (2 polmoni * 4 coni each), matching the reference
    query's Custom column."""
    digits = re.sub(r"[^0-9]", "", _clean(polmoni_text))
    return int(digits) * 4 if digits else 0


@dataclass
class OrdineMedRow:
    riga: int
    code_org: str
    titolo: str
    descr_col: str
    articolo: str
    colore: str
    rocc: int
    abbin: Any
    consegna_input: Any
    pt_grg: str
    pt_med: str
    polmoni: str
    cliente_note: str
    nota_grg: str
    nota_col: str
    kg_note: str
    fabb: Any
    prezz_note: Any
    mc: int = 0
    gruppo_macchina: Any = None
    commento: str = ""
    cliente: str = ""
    data_riconsegna: Any = None
    consegna: Any = None
    prezzo: Any = ""
    livello: Any = ""
    check_articolo: str = ""
    prezzo_plus2: Any = ""


def load_ordine(path: Path) -> list[OrdineMedRow]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = "ORDINE" if "ORDINE" in wb.sheetnames else wb.sheetnames[0]
        rows = _read_sheet_rows(wb[sheet])
    finally:
        wb.close()

    out: list[OrdineMedRow] = []
    for i, row in enumerate(rows, start=1):
        articolo = _clean(_get(row, "ARTICOLO"))
        if not articolo:
            continue
        out.append(OrdineMedRow(
            riga=i,
            code_org=_clean(_get(row, "CODE.ORG")),
            titolo=_clean(_get(row, "TITOLO")),
            descr_col=_clean(_get(row, "DESCR COL")),
            articolo=articolo,
            colore=_clean(_get(row, "COLORE")),
            rocc=int(_number(_get(row, "ROCC")) or 0),
            abbin=_get(row, "ABBIN"),
            consegna_input=_get(row, "CONSEGNA"),
            pt_grg=_clean(_get(row, "PT GRG")),
            pt_med=_clean(_get(row, "PT MED")),
            polmoni=_clean(_get(row, "POLMONI")),
            cliente_note=_clean(_get(row, "cliente")),
            nota_grg=_clean(_get(row, "NOTA grg")),
            nota_col=_clean(_get(row, "NOTA col")),
            kg_note=_clean(_get(row, "KG")),
            fabb=_get(row, "FABB"),
            prezz_note=_get(row, "PREZZ"),
        ))
    if not out:
        raise ValueError('Nessuna riga trovata nel foglio "ORDINE" (colonna ARTICOLO vuota su tutte le righe).')
    return out


def compute_mc_and_gruppo(records: list[OrdineMedRow]) -> None:
    """M/C = Rocche summed across every row sharing the same non-null ABBIN
    (+ Polmoni*4), applied to every row in that group; GRUPPO MACCHINA is
    the fixed lookup off that total."""
    groups: dict[Any, list[OrdineMedRow]] = {}
    for r in records:
        roc = r.rocc + _polmoni_multiplier(r.polmoni)
        r._roc_with_polmoni = roc  # type: ignore[attr-defined]
        if r.abbin not in (None, ""):
            groups.setdefault(r.abbin, []).append(r)
        else:
            r.mc = roc
            r.gruppo_macchina = GRUPPO_MACCHINA_TABLE.get(roc)
    for abbin, members in groups.items():
        total = sum(m._roc_with_polmoni for m in members)  # type: ignore[attr-defined]
        gm = GRUPPO_MACCHINA_TABLE.get(total)
        for m in members:
            m.mc = total
            m.gruppo_macchina = gm


def compute_titolo(records: list[OrdineMedRow], marca_map: dict[str, str]) -> None:
    """Overrides the raw sheet's own (sometimes corrupted -- can literally
    contain a date) Titolo with a lookup off CODE.ORG, same C/G-swap
    fallback Biglietti uses."""
    for r in records:
        art = r.code_org.upper()
        titolo = marca_map.get(art, "")
        if not titolo and art[:1] in ("C", "G"):
            swapped = ("G" if art[:1] == "C" else "C") + art[1:]
            titolo = marca_map.get(swapped, "")
        if titolo:
            r.titolo = titolo


def compute_cliente(records: list[OrdineMedRow]) -> None:
    for r in records:
        code = r.code_org.upper()
        if code.startswith("C130"):
            r.cliente = "3009"
        elif code.startswith("C010") or code.startswith("C011"):
            r.cliente = "3004"
        else:
            r.cliente = ""


def compute_commento(records: list[OrdineMedRow]) -> None:
    for r in records:
        segments = [
            "PG",
            r.pt_grg or "X",
            "PM",
            r.pt_med or "X",
            (r.polmoni[:10] if r.polmoni else ""),
            (r.cliente_note[:10] if r.cliente_note else ""),
            (r.nota_col[:10] if r.nota_col else ""),
        ]
        text = "-".join(segments)
        text = re.sub(r"-{2,}", "-", text).rstrip("-")
        r.commento = text


def compute_data_riconsegna(records: list[OrdineMedRow]) -> None:
    for r in records:
        v = r.consegna if r.consegna is not None else r.consegna_input
        try:
            r.data_riconsegna = v - timedelta(days=1)
        except TypeError:
            r.data_riconsegna = None


# ---------------------------------------------------------------------------
# Consegna auto-scheduling: walk the order top-to-bottom, and for every
# machine (GRUPPO MACCHINA -> plain 3-12 number), queue new colors onto
# whatever the machine's *current* Copertura coverage already is.
# ---------------------------------------------------------------------------

def assign_consegna(records: list[OrdineMedRow], machine_totals: dict[int, int]) -> None:
    """machine_totals: {machine_number (3-12): colors_already_queued}, from
    situazione_logic.compute_machine_totals(situation_df, copertura_df) --
    the same baseline the Copertura window itself shows. Rows whose machine
    isn't recognised or has no baseline just start counting from 0 (today)."""
    from situazione_logic import machine_coverage_until, machine_number_from_label

    running: dict[int, int] = {}
    for r in records:
        m = machine_number_from_label(str(r.gruppo_macchina)) if r.gruppo_macchina else None
        if m is None:
            r.consegna = r.consegna_input
            continue
        baseline = machine_totals.get(m, 0)
        running[m] = running.get(m, 0) + 1
        date_str = machine_coverage_until(baseline + running[m])
        r.consegna = datetime.strptime(date_str, "%Y-%m-%d") if date_str != "-" else r.consegna_input


# ---------------------------------------------------------------------------
# Prezzo -- reuses the same Articolo+Codice Listini lookup Biglietti uses.
# ---------------------------------------------------------------------------

def compute_prezzo(records: list[OrdineMedRow], price_lookup: dict[tuple, tuple]) -> None:
    for r in records:
        key = (r.articolo, r.colore)
        if key in price_lookup:
            r.livello, r.prezzo = price_lookup[key]


# Machines (by their Rocche-based M/C total) that get a $2 surcharge.
PREZZO_SURCHARGE_MACHINES = {56, 32, 24}


def compute_prezzo_plus2(records: list[OrdineMedRow]) -> None:
    """PREZZO + 2$ column: +2 on top of Prezzo for the 3 specific machine
    sizes (56/32/24 Rocche), otherwise the same price carried over as-is."""
    for r in records:
        if isinstance(r.prezzo, (int, float)) and r.mc in PREZZO_SURCHARGE_MACHINES:
            r.prezzo_plus2 = round(r.prezzo + 2, 2)
        else:
            r.prezzo_plus2 = r.prezzo


def load_dfm_articolo_colore(path: Path) -> set[tuple[str, str]]:
    """{(ARTICOLODFM, COLOREDFM)} pairs seen historically in the DFM
    export -- read from the raw DFM sheet directly (not the simplified
    situazione_loaders.load_dfm, which drops the color column), so 'Check
    Articolo' can tell a genuinely new Articolo+Colore combination from one
    that's simply missing a color code."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = "DFM" if "DFM" in wb.sheetnames else wb.sheetnames[0]
        rows = _read_sheet_rows(wb[sheet])
    finally:
        wb.close()
    pairs: set[tuple[str, str]] = set()
    for row in rows:
        art = _clean(_get(row, "ARTICOLODFM", "Articolo")).upper()
        col = _clean(_get(row, "COLOREDFM", "Colore"))
        if art:
            pairs.add((art, col))
    return pairs


def compute_check_articolo(records: list[OrdineMedRow], dfm_pairs: set[tuple[str, str]]) -> None:
    """'NEW' when this Articolo+Colore combination has never been dyed
    before (not found in DFM). Left blank rather than defaulting to 'NEW'
    when no DFM data is available at all, to avoid flagging everything
    as new just because the source wasn't provided."""
    if not dfm_pairs:
        for r in records:
            r.check_articolo = ""
        return
    for r in records:
        key = (r.articolo.upper(), r.colore)
        r.check_articolo = "" if key in dfm_pairs else "NEW"


# ---------------------------------------------------------------------------
# Availability / shortage check (the PT-GG query) -- group the order by
# (Articolo, Titolo, PT GRG) summing Rocche, join against Filato
# Disponibile's Mag. Rocche, compute Manca / Disponibilita'.
# ---------------------------------------------------------------------------

def load_filato_disponibile(path: Path) -> dict[int, int]:
    """{PARTITA: Mag.Rocche} -- filters to MAGAZZINO in {900160, 900910},
    excludes committed stock (MAGAZZINO=900160 and ORDINE=0), sums COLLI
    per PARTITA. Verified against real data."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        rows = _read_sheet_rows(wb.active)
    finally:
        wb.close()
    totals: dict[int, int] = {}
    for row in rows:
        magazzino = _number(_get(row, "MAGAZZINO"))
        if magazzino not in (900160, 900910):
            continue
        ordine = _number(_get(row, "ORDINE"))
        if magazzino == 900160 and (ordine or 0) == 0:
            continue
        partita = _number(_get(row, "PARTITA"))
        if partita is None:
            continue
        colli = _number(_get(row, "COLLI"))
        if colli is None:
            continue
        totals[int(partita)] = totals.get(int(partita), 0) + int(colli)
    return totals


@dataclass
class FilatoAvailabilityRow:
    articolo: str
    titolo: str
    pt_grg: int
    rocche: int
    kg: Any
    mag_rocche: Any
    manca: Any
    disponibilita: str


def compute_filato_availability(
    records: list[OrdineMedRow],
    densita_map: dict[int, dict[str, Any]] | None,
    stock_map: dict[int, int],
) -> list[FilatoAvailabilityRow]:
    """Groups by (ARTICOLO with C->G swap, TITOLO, PT GRG), sums Rocche,
    computes Kg via Densita' Query's peso_net when available (same source
    Biglietti's KG uses), joins Mag.Rocche from Filato Disponibile."""
    densita_map = densita_map or {}
    groups: dict[tuple, dict[str, Any]] = {}
    for r in records:
        try:
            pt_grg = int(_number(r.pt_grg))
        except (TypeError, ValueError):
            continue
        art_g = ("G" + r.articolo[1:]) if r.articolo[:1] == "C" else r.articolo
        key = (art_g, r.titolo, pt_grg)
        g = groups.setdefault(key, {"rocche": 0})
        g["rocche"] += r.rocc + _polmoni_multiplier(r.polmoni)

    out: list[FilatoAvailabilityRow] = []
    for (art_g, titolo, pt_grg), g in groups.items():
        rocche = g["rocche"]
        peso_net = densita_map.get(pt_grg, {}).get("peso_net")
        kg = round(peso_net * rocche, 2) if peso_net is not None else None
        mag_rocche = stock_map.get(pt_grg)
        manca = (mag_rocche - rocche) if mag_rocche is not None else None
        disponibilita = "OK" if (mag_rocche is not None and mag_rocche >= rocche) else "NO"
        out.append(FilatoAvailabilityRow(art_g, titolo, pt_grg, rocche, kg, mag_rocche, manca, disponibilita))
    out.sort(key=lambda x: x.pt_grg)
    return out


# ---------------------------------------------------------------------------
# Export -- two sheets: "Ordine da creare" (the full computed table, plus a
# ready-to-import CLIENTE..GRUPPO MACCHINA slice) and "Filato X Tinturia"
# (the availability/shortage table).
# ---------------------------------------------------------------------------

# The full field set, used to build the system-import slice
# ("Dati sistema", CLIENTE..GRUPPO MACCHINA).
FULL_ROW_HEADERS = [
    "Riga", "CLIENTE", "ARTICOLO", "COLORE", "Q.TA", "CONSEGNA", "COMMENTO",
    "LAVORANTE", "LAV. SUCC", "DATA RICONSEGNA", "MAG. GREGGIO",
    "SIGLA DISPOSIZONE", "BAGNO PREPOSTO", "GRUPPO MACCHINA",
    "TITOLO", "DESCR COL", "LIVELLO", "ABBIN", "M/C", "PREZZO",
    "PT GRG", "PT MED", "POLMONI", "Cliente MED", "NOTA grg", "NOTA col",
]

# The system-import slice: CLIENTE .. GRUPPO MACCHINA (columns B-N in the
# reference sheet's own layout, "Riga" excluded).
SYSTEM_IMPORT_HEADERS = FULL_ROW_HEADERS[1:14]

# "Ordine da creare"'s own display columns -- CLIENTE, COMMENTO and the
# fixed ERP constants (LAVORANTE..GRUPPO MACCHINA) live only in "Dati
# sistema" now; this sheet adds Check Articolo (after M/C) and
# PREZZO + 2$ (after PREZZO) instead.
ORDINE_DA_CREARE_HEADERS = [
    "Riga", "ARTICOLO", "COLORE", "Q.TA", "CONSEGNA", "DATA RICONSEGNA",
    "TITOLO", "DESCR COL", "LIVELLO", "ABBIN", "M/C", "Check Articolo",
    "PREZZO", "PREZZO + 2$",
    "PT GRG", "PT MED", "POLMONI", "Cliente MED", "NOTA grg", "NOTA col",
]

FILATO_AVAILABILITY_HEADERS = [
    "Articolo", "Titolo", "PT GRG", "Rocche", "Kg", "Mag. Rocche", "Manca", "Disponibilita'",
]


def _full_row(r: OrdineMedRow) -> dict[str, Any]:
    return dict(zip(FULL_ROW_HEADERS, [
        r.riga, r.cliente, r.articolo, r.colore, r.rocc, r.consegna, r.commento,
        "900901", "900161", r.data_riconsegna, "900923", "D", "", r.gruppo_macchina,
        r.titolo, r.descr_col, r.livello, r.abbin, r.mc, r.prezzo,
        r.pt_grg, r.pt_med, r.polmoni, r.cliente_note, r.nota_grg, r.nota_col,
    ]))


def _ordine_da_creare_row(r: OrdineMedRow) -> list[Any]:
    return [
        r.riga, r.articolo, r.colore, r.rocc, r.consegna, r.data_riconsegna,
        r.titolo, r.descr_col, r.livello, r.abbin, r.mc, r.check_articolo,
        r.prezzo, r.prezzo_plus2,
        r.pt_grg, r.pt_med, r.polmoni, r.cliente_note, r.nota_grg, r.nota_col,
    ]


def export_ordine_med_workbook(
    path: Path,
    records: list[OrdineMedRow],
    availability: list[FilatoAvailabilityRow],
) -> None:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    from biglietti_exporter import _style_sheet

    wb = Workbook()

    ws = wb.active
    ws.title = "Ordine da creare"
    ws.append(ORDINE_DA_CREARE_HEADERS)
    for r in records:
        ws.append(_ordine_da_creare_row(r))
    _style_sheet(ws, date_columns=("CONSEGNA", "DATA RICONSEGNA"))
    _style_ordine_da_creare_extras(ws)

    ws2 = wb.create_sheet("Dati sistema (B-N)")
    ws2.append(SYSTEM_IMPORT_HEADERS)
    for r in records:
        full = _full_row(r)
        ws2.append([full[h] for h in SYSTEM_IMPORT_HEADERS])
    _style_sheet(ws2, date_columns=("CONSEGNA", "DATA RICONSEGNA"))

    ws3 = wb.create_sheet("Filato X Tinturia")
    ws3.append(FILATO_AVAILABILITY_HEADERS)
    for a in availability:
        ws3.append([a.articolo, a.titolo, a.pt_grg, a.rocche, a.kg, a.mag_rocche, a.manca, a.disponibilita])
    _style_sheet(ws3)
    # Highlight shortages (Disponibilita' = NO) in red.
    if ws3.max_row > 1:
        red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        rng = f"A2:H{ws3.max_row}"
        ws3.conditional_formatting.add(
            rng, FormulaRule(formula=['$H2="NO"'], fill=red_fill, stopIfTrue=False)
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _style_ordine_da_creare_extras(ws) -> None:
    """The bespoke formatting "Ordine da creare" needs beyond the shared
    _style_sheet borders/banding: Check Articolo highlighted red when
    'NEW', LIVELLO solid red throughout, ABBIN duplicates highlighted
    green, and the PREZZO + 2$ header given a distinct fill so it reads
    as a derived column rather than raw input."""
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill

    header = [c.value for c in ws[1]]
    last_row = ws.max_row
    if last_row <= 1:
        return

    def col_letter(name: str) -> str | None:
        return ws.cell(row=1, column=header.index(name) + 1).column_letter if name in header else None

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

    check_col = col_letter("Check Articolo")
    if check_col:
        rng = f"{check_col}2:{check_col}{last_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{check_col}2="NEW"'], fill=red_fill, stopIfTrue=False))

    livello_col = col_letter("LIVELLO")
    if livello_col:
        col_idx = header.index("LIVELLO") + 1
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).fill = red_fill

    abbin_col = col_letter("ABBIN")
    if abbin_col:
        rng = f"{abbin_col}2:{abbin_col}{last_row}"
        formula = f'AND({abbin_col}2<>"",COUNTIF(${abbin_col}$2:${abbin_col}${last_row},{abbin_col}2)>1)'
        ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=green_fill, stopIfTrue=False))

    plus2_col = col_letter("PREZZO + 2$")
    if plus2_col:
        header_cell = ws.cell(row=1, column=header.index("PREZZO + 2$") + 1)
        header_cell.fill = PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="solid")
        header_cell.font = Font(bold=True, color="FFFFFFFF")
