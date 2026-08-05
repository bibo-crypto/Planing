"""
situazione_settimana_tab.py
Weekly Customer x Machine dyed-weight summary, embedded as the "Situazione
Settimana" tab. Shares its DFM and Produzione file uploads with the
Situazione tab (situazione_tab.py) via dfm_lookup.py / prod_lookup.py --
upload either file on either page and both pick it up automatically.
"""
import os
import threading
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd

import situazione_loaders as data_loaders
import situazione_settimana_logic as logic
from situazione_tab import SourceRow
from dfm_lookup import build_dfm_lookup, load_dfm_cache, save_dfm_cache
from prod_lookup import load_prod_cache, save_prod_cache
from utils import logger

SOURCE_LABELS = {"dfm": "Machine data (DFM)", "data_prod": "Production data"}
COLUMNS = ["cliente", "week_of_year", "machine_name", "total_peso", "batch_count"]
HEADERS = {"cliente": "Cliente", "week_of_year": "Settimana", "machine_name": "Machine",
           "total_peso": "Sum of Peso", "batch_count": "Batches"}


class SettimanaTab(ttk.Frame):
    """Embeddable 'Situazione Settimana' tab."""

    def __init__(self, master):
        super().__init__(master)
        self.loaded_frames = {}          # "dfm" / "data_prod" -> DataFrame
        self._shared_dfm_path = ""
        self._shared_prod_path = ""
        self.batch_weights = pd.DataFrame()

        self._build_upload_panel()
        self._build_toolbar()
        self._build_treeview()

    # ------------------------------------------------------------------ UI
    def _build_upload_panel(self):
        panel = ttk.LabelFrame(self, text="1) Upload files (shared with the Situazione tab)")
        panel.pack(side="top", fill="x", padx=8, pady=6)

        self.source_rows = {}
        for key, label in SOURCE_LABELS.items():
            row = SourceRow(panel, key, label, self._handle_upload)
            row.pack(fill="x", padx=4, pady=2)
            self.source_rows[key] = row

    def _build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(side="top", fill="x", padx=8, pady=4)

        self.refresh_btn = ttk.Button(bar, text="2) Calculate", command=self._on_refresh)
        self.refresh_btn.pack(side="left", padx=4)

        self.export_btn = ttk.Button(bar, text="Export Excel", command=self._on_export)
        self.export_btn.pack(side="left", padx=4)

        ttk.Label(bar, text="Week:").pack(side="left", padx=(16, 4))
        self.week_var = tk.StringVar(value="All")
        self.week_combo = ttk.Combobox(bar, textvariable=self.week_var, state="readonly",
                                        width=10, values=["All"])
        self.week_combo.pack(side="left")
        self.week_combo.bind("<<ComboboxSelected>>", lambda _e: self._render())

        self.summary_lbl = ttk.Label(bar, text="")
        self.summary_lbl.pack(side="right", padx=8)

    def _build_treeview(self):
        frame = ttk.Frame(self)
        frame.pack(side="top", fill="both", expand=True, padx=8, pady=6)

        self.tree = ttk.Treeview(frame, columns=COLUMNS, show="tree headings", selectmode="browse")
        self.tree.heading("#0", text="")
        self.tree.column("#0", width=10, stretch=False)
        for c in COLUMNS:
            self.tree.heading(c, text=HEADERS[c])
            anchor = "w" if c in ("cliente", "machine_name") else "center"
            self.tree.column(c, width=140, anchor=anchor)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tree.tag_configure("customer", background="#dbeafe", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("week", background="#eef2f7", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("grand_total", background="#16324f", foreground="#ffffff",
                                 font=("Segoe UI", 9, "bold"))

    # ------------------------------------------------------------- uploads
    def _handle_upload(self, key, path):
        # Settimana needs the FULL Produzione columns (machine/peso/codice/date),
        # not Situazione's narrow "data_prod" subset -- use load_produzione for that slot.
        loader_fn = data_loaders.load_produzione if key == "data_prod" else data_loaders.load_dfm
        try:
            df, errors = loader_fn(path)
        except Exception as exc:  # noqa: BLE001
            errors = [f"An error occurred while reading the file: {exc}"]
            df = None

        if errors or df is None or df.empty:
            msg = "; ".join(errors) if errors else "The file is empty after filtering"
            self.source_rows[key].set_status(False, f"❌ {msg}")
            self.loaded_frames.pop(key, None)
            return

        self.loaded_frames[key] = df
        self.source_rows[key].set_status(True, f"✅ {len(df)} rows")
        logger.info("Situazione Settimana: %s uploaded — %d rows (%s)", key, len(df), os.path.basename(path))

        if key == "dfm":
            self._save_shared_dfm(path)
        if key == "data_prod":
            self._save_shared_prod(path)

    # ---------------------------------------------------- shared DFM / Prod
    def sync_shared_dfm(self):
        """Load the DFM selected in either page from the shared persistent cache."""
        cache = load_dfm_cache()
        source_path = Path(str(cache.get("source_path", "")))
        if not source_path.is_file() or self._shared_dfm_path == str(source_path):
            return
        try:
            df, errors = data_loaders.load_dfm(str(source_path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not restore shared DFM file: %s", exc)
            return
        if errors or df is None or df.empty:
            return
        self.loaded_frames["dfm"] = df
        self._shared_dfm_path = str(source_path)
        self.source_rows["dfm"].set_status(True, f"✅ {len(df)} rows")

    def sync_shared_async(self):
        """Restore shared DFM/Produzione files without blocking the UI thread."""
        self._sync_shared_source_async("dfm")
        self._sync_shared_source_async("data_prod")

    def _sync_shared_source_async(self, key):
        cache = load_dfm_cache() if key == "dfm" else load_prod_cache()
        source_path = Path(str(cache.get("source_path", "")))
        syncing_key = f"_syncing_{key}"
        if not source_path.is_file() or getattr(self, syncing_key, False):
            return
        if getattr(self, f"_shared_{'dfm' if key == 'dfm' else 'prod'}_path", "") == str(source_path):
            return
        setattr(self, syncing_key, True)
        loader = data_loaders.load_dfm if key == "dfm" else data_loaders.load_produzione

        def worker():
            try:
                df, errors = loader(str(source_path))
            except Exception as exc:  # noqa: BLE001
                df, errors = None, [str(exc)]

            def apply_result():
                setattr(self, syncing_key, False)
                if errors or df is None or df.empty:
                    return
                self.loaded_frames[key] = df
                setattr(self, f"_shared_{'dfm' if key == 'dfm' else 'prod'}_path", str(source_path))
                self.source_rows[key].set_status(True, f"✅ {len(df)} rows")

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def _save_shared_dfm(self, path):
        try:
            entries = build_dfm_lookup(Path(path))
            if entries:
                save_dfm_cache(entries, Path(path).name, Path(path))
                self._shared_dfm_path = str(Path(path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not update shared DFM reference: %s", exc)

    def sync_shared_prod(self):
        """Load the Produzione file selected in either page from the shared cache."""
        cache = load_prod_cache()
        source_path = Path(str(cache.get("source_path", "")))
        if not source_path.is_file() or self._shared_prod_path == str(source_path):
            return
        try:
            df, errors = data_loaders.load_produzione(str(source_path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not restore shared Produzione file: %s", exc)
            return
        if errors or df is None or df.empty:
            return
        self.loaded_frames["data_prod"] = df
        self._shared_prod_path = str(source_path)
        self.source_rows["data_prod"].set_status(True, f"✅ {len(df)} rows")

    def _save_shared_prod(self, path):
        try:
            save_prod_cache(Path(path))
            self._shared_prod_path = str(Path(path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not update shared Produzione reference: %s", exc)

    # -------------------------------------------------------------- refresh
    def _on_refresh(self):
        dfm_df = self.loaded_frames.get("dfm")
        prod_df = self.loaded_frames.get("data_prod")
        if prod_df is None or dfm_df is None:
            messagebox.showwarning(
                "Missing data",
                "Upload both the Produzione file and the DFM file before calculating "
                "(either here or on the Situazione tab — they're shared)."
            )
            return

        self.batch_weights = logic.compute_batch_weights(prod_df, dfm_df)
        weeks = logic.available_weeks(self.batch_weights)
        self.week_combo["values"] = ["All"] + [str(w) for w in weeks]
        if self.week_var.get() not in self.week_combo["values"]:
            self.week_var.set("All")
        logger.info("Situazione Settimana: calculated — %d batches across %d week(s)",
                    len(self.batch_weights), len(weeks))
        self._render()

    # -------------------------------------------------------------- display
    def _render(self):
        self.tree.delete(*self.tree.get_children())
        if self.batch_weights.empty:
            self.summary_lbl.config(text="")
            return

        week_filter = self.week_var.get()
        summary = logic.summarize(self.batch_weights, week_of_year=week_filter)
        if summary.empty:
            self.summary_lbl.config(text="No batches for this week")
            return

        grand_peso = 0.0
        grand_batches = 0
        for cliente, cust_group in summary.groupby("cliente"):
            cust_peso = cust_group["total_peso"].sum()
            cust_batches = int(cust_group["batch_count"].sum())
            grand_peso += cust_peso
            grand_batches += cust_batches
            cust_node = self.tree.insert(
                "", "end", text="",
                values=(cliente, "", "", f"{cust_peso:g}", cust_batches),
                tags=("customer",), open=True,
            )
            for week, week_group in cust_group.groupby("week_of_year"):
                week_peso = week_group["total_peso"].sum()
                week_batches = int(week_group["batch_count"].sum())
                week_node = self.tree.insert(
                    cust_node, "end", text="",
                    values=("", int(week), "", f"{week_peso:g}", week_batches),
                    tags=("week",), open=True,
                )
                for _, r in week_group.sort_values("machine_name").iterrows():
                    self.tree.insert(
                        week_node, "end", text="",
                        values=("", "", r["machine_name"], f"{r['total_peso']:g}", int(r["batch_count"])),
                    )

        self.tree.insert(
            "", "end", text="",
            values=("Grand Total", "", "", f"{grand_peso:g}", grand_batches),
            tags=("grand_total",),
        )
        week_label = "all weeks" if week_filter == "All" else f"week {week_filter}"
        self.summary_lbl.config(text=f"{grand_batches} batches, {grand_peso:g} kg total ({week_label})")

    # -------------------------------------------------------------- export
    def _on_export(self):
        if self.batch_weights.empty:
            messagebox.showinfo("No data", "Calculate before exporting.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")],
                                             initialfile="Situazione_Settimana.xlsx")
        if not path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        week_filter = self.week_var.get()
        summary = logic.summarize(self.batch_weights, week_of_year=week_filter)
        week_ranges = logic.week_date_ranges(self.batch_weights)

        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="16324F", end_color="16324F", fill_type="solid")
        customer_fills = [PatternFill(start_color=c, end_color=c, fill_type="solid")
                           for c in ("FFFFFF", "F4F6F8")]

        wb = Workbook()
        ws = wb.active
        ws.title = "Situazione Settimana"

        # Flat table: one row per Cliente/Settimana/Machine, every cell filled in --
        # this is what lets Settimana (and every other column) get its own
        # working Excel filter dropdown, unlike a subtotal/outline layout.
        headers = [HEADERS[c] for c in COLUMNS]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        summary = summary.sort_values(["cliente", "week_of_year", "machine_name"])
        clientes_seen = []
        for _, r in summary.iterrows():
            if r["cliente"] not in clientes_seen:
                clientes_seen.append(r["cliente"])
            shade = customer_fills[clientes_seen.index(r["cliente"]) % 2]
            ws.append([r["cliente"], int(r["week_of_year"]), r["machine_name"],
                       r["total_peso"], int(r["batch_count"])])
            for cell in ws[ws.max_row]:
                cell.fill = shade
                cell.font = Font(name="Arial")
                cell.alignment = center
                cell.border = border

        last_row = ws.max_row
        last_col_letter = get_column_letter(len(COLUMNS))
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        ws.freeze_panes = "A2"
        for i, c in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        # Week -> date range reference table, off to the side
        ref_col = len(COLUMNS) + 2  # leave one blank column as a gap
        ref_col_letter = get_column_letter(ref_col)
        ws.cell(row=1, column=ref_col, value="Settimana").font = Font(bold=True, name="Arial", color="FFFFFF")
        ws.cell(row=1, column=ref_col + 1, value="From").font = Font(bold=True, name="Arial", color="FFFFFF")
        ws.cell(row=1, column=ref_col + 2, value="To").font = Font(bold=True, name="Arial", color="FFFFFF")
        for c in range(ref_col, ref_col + 3):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        for i, (week, (start, end)) in enumerate(week_ranges.items(), start=2):
            ws.cell(row=i, column=ref_col, value=int(week))
            ws.cell(row=i, column=ref_col + 1, value=start)
            ws.cell(row=i, column=ref_col + 2, value=end)
            for c in range(ref_col, ref_col + 3):
                cell = ws.cell(row=i, column=c)
                cell.alignment = center
                cell.border = border
                cell.font = Font(name="Arial")
                if c > ref_col:
                    cell.number_format = "dd/mm/yyyy"
        ws.column_dimensions[ref_col_letter].width = 12
        ws.column_dimensions[get_column_letter(ref_col + 1)].width = 12
        ws.column_dimensions[get_column_letter(ref_col + 2)].width = 12

        wb.save(path)
        logger.info("Situazione Settimana: exported to %s", path)
        messagebox.showinfo("Completed", f"Export completed successfully:\n{path}")
