"""
bolla_exporter.py — Builds an Excel workbook from parsed Bolla data.

Two sheets:
    "Summary" — one row per distinct (Articolo, Descrizione, Colore, Partita)
                combination, with Rocche and KgNetto summed across every
                matching row — across all Bolla files being exported
                together, not just one document.
    "Totals"  — one row per Bolla document, with its "Totale" summary line.

Formatting mirrors excel_exporter.py (header styling, freeze panes,
autofilter, auto-sized columns) so both document types feel consistent.
"""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bolla_parser import BollaRow, BollaTotals, group_rows_by_lot
from utils import logger

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

GROUPED_COLUMNS: list[tuple[str, str, str]] = [
    ("Articolo",    "articolo",    "text"),
    ("Descrizione", "descrizione", "text"),
    ("Colore",      "colore",      "integer"),
    ("Partita",     "partita",     "integer"),
    ("Rocche",      "rocche",      "integer"),
    ("KgNetto",     "kg_netto",    "number"),
]

TOTALS_COLUMNS: list[tuple[str, str, str]] = [
    ("Bolla No",   "bolla_no",   "text"),
    ("Del",        "bolla_date", "text"),
    ("N.Colli",    "n_colli",    "number"),
    ("N.Pallet",   "n_pallet",   "number"),
    ("TaraTub.",   "tara_tub",   "number"),
    ("TaraScat.",  "tara_scat",  "number"),
    ("Rocche",     "rocche",     "integer"),
    ("Netto",      "netto",      "number"),
    ("Lordo",      "lordo",      "number"),
]

# Styling constants (matches excel_exporter.py)
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
HEADER_FONT = Font(bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER_SIDE = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE,
)

NUMBER_FORMAT = "#,##0.00"
# General keeps KgNetto numeric while showing only meaningful decimals:
# 2119, 866.2 — never a thousands separator, trailing .00, or a lone dot.
KG_NETTO_FORMAT = "General"
INTEGER_FORMAT = "0"
TEXT_FORMAT = "@"
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 8
DESCRIPTION_COL_WIDTH = 40


def _write_sheet(
    ws, rows: Sequence[object], columns: list[tuple[str, str, str]]
) -> None:
    """Write a header row + data rows for *rows* using *columns* into *ws*."""
    ws.sheet_view.rightToLeft = False

    for col_idx, (label, _attr, _dtype) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    for row_idx, obj in enumerate(rows, start=2):
        for col_idx, (_, attr, dtype) in enumerate(columns, start=1):
            raw_value = getattr(obj, attr, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER

            if dtype == "number":
                cell.value = float(raw_value) if raw_value is not None else None
                cell.number_format = (
                    KG_NETTO_FORMAT if attr == "kg_netto" else NUMBER_FORMAT
                )
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif dtype == "integer":
                # Stored as a real number (not text), so a value like the
                # string "007777" becomes 7777 automatically — leading
                # zeros drop out because they aren't meaningful once it's
                # a number, not because of any separate string-stripping.
                try:
                    cell.value = (
                        int(round(float(raw_value)))
                        if raw_value not in (None, "")
                        else None
                    )
                except (TypeError, ValueError):
                    cell.value = str(raw_value)  # not numeric — keep as-is
                cell.number_format = INTEGER_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:  # "text"
                cell.value = str(raw_value) if raw_value not in (None, "") else ""
                cell.number_format = TEXT_FORMAT
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, (label, attr, _) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        if attr == "descrizione":
            ws.column_dimensions[col_letter].width = DESCRIPTION_COL_WIDTH
            continue
        max_width = len(label)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_width = max(max_width, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(MIN_COL_WIDTH, min(max_width + 4, MAX_COL_WIDTH))

    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


class BollaExporter:
    """
    Writes parsed Bolla data to an Excel workbook with two sheets:
    "Summary" (grouped by Articolo/Descrizione/Colore/Partita, with Rocche
    and KgNetto summed) and "Totals" (one row per document).

    Usage::

        exporter = BollaExporter(rows, totals_list, output_path=Path("bolle.xlsx"))
        exporter.export()
    """

    def __init__(
        self,
        rows: Sequence[BollaRow],
        totals: Sequence[BollaTotals],
        output_path: Path,
    ) -> None:
        self.rows = rows
        self.totals = totals
        self.output_path = output_path

    def export(self) -> None:
        grouped = group_rows_by_lot(list(self.rows))

        logger.info(
            "Exporting %d Bolla row(s) grouped into %d summary row(s), "
            "%d totals row(s) -> %s",
            len(self.rows), len(grouped), len(self.totals), self.output_path.name,
        )

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        _write_sheet(ws_summary, grouped, GROUPED_COLUMNS)

        ws_totals = wb.create_sheet("Totals")
        _write_sheet(ws_totals, self.totals, TOTALS_COLUMNS)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        logger.info("Export completed: %s", self.output_path)
