"""
excel_exporter.py — Builds a professional Excel workbook from parsed OrderRow data.

Formatting applied
------------------
* Header row: bold, centred, light blue fill.
* Freeze top row.
* AutoFilter on all columns.
* Numeric columns: right-aligned, number format.
* Date columns: date format.
* Colour column: text format (preserves leading zeros).
* Description column: text wrap + taller row height.
* All columns auto-sized to content (with a reasonable max width).
* Sheet direction: left-to-right.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Sequence

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from pdf_parser import OrderRow
from ordini_elvy import OrdiniElvyRow, build_ordini_elvy_rows, match_raw_yarn
from utils import logger

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# Each tuple: (header_label, attribute_name, data_type)
# data_type: "text" | "number" | "integer" | "date"
COLUMNS: list[tuple[str, str, str]] = [
    ("POD No",              "pod_no",              "text"),
    ("Po Date",             "po_date",             "date"),
    ("Pos",                 "pos",                 "integer"),
    ("Article No",          "article_no",          "text"),
    ("Articolo Delta",      "articolo_delta",      "text"),
    ("COLOREDFM",           "coloredfm",           "text"),
    ("CLDESCR",             "cldescr",             "text"),
    ("Article Description", "article_description", "text"),
    ("Yarn",                "yarn",                "text"),
    ("Nm",                  "nm",                  "text"),
    ("Ne",                  "ne",                  "text"),
    ("Dye Type",            "dye_type",            "text"),
    ("Lot",                 "lot",                 "text"),
    ("Colour",              "colour",              "text"),
    ("Quantity /Cones",     "quantity_cones",      "number"),
    ("Abbina",              "abbina",              "text"),
    ("Quantity /Kg",        "quantity_kg",         "number"),
    ("Price in $",          "price_usd",           "number"),
    ("Value in USD",        "value_usd",           "number"),
    ("Ship Date",           "ship_date",           "date"),
]

# "Ordini ELVY" sheet — derived columns, see ordini_elvy.py.
# Header labels match EXCEL_PER_ORDINE_VENDITA_EGITTO.xlsx exactly, so this
# sheet's data can be copied straight into that template.
ORDINI_ELVY_COLUMNS: list[tuple[str, str, str]] = [
    ("CLIENTE",              "cliente",              "integer"),
    ("ARTICOLO",             "articolo_delta",       "text"),
    ("COLORE",               "coloredfm",            "integer"),
    ("Q.TA",                 "quantity_cones",       "integer"),
    ("CONSEGNA",             "data_consegna",        "date"),
    ("COMMENTO",             "commento",             "text"),
    ("LAVORANTE",            "lavorante",            "integer"),
    ("LAV. SUCC",            "lavorante_successivo", "integer"),
    ("DATA RICONSEGNA",      "data_riconsegna",      "date"),
    ("MAG. GREGGIO",         "stock_mag",            "integer"),
    ("SIGLA DISPOSIZIONE",   "sigla_dispo",          "text"),
    ("BAGNO PREPOSTO",       "bagno_proposto",       "text"),
    ("GRUPPO MACCHINA",      "macchina",             "integer"),
]

FILATO_TINTURIA_COLUMNS: list[tuple[str, str, str]] = [
    ("Articolo", "articolo", "text"),
    ("Titolo",   "titolo",   "text"),
    ("Partita",  "partita",  "text"),
    ("Rocce",    "rocce",    "number"),
    ("Peso",     "peso",     "number"),
    ("تحضير خام", "label",   "text"),
]

# Styling constants
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")   # light blue
HEADER_FONT = Font(bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER_SIDE = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
)

NUMBER_FORMAT = "#,##0.00"
INTEGER_FORMAT = "0"
DATE_FORMAT = "DD/MM/YYYY"
TEXT_FORMAT = "@"   # force Excel to treat cell as text

MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 8
DESCRIPTION_COL_WIDTH = 50
ROW_HEIGHT_WRAPPED = 45  # points for wrapped description rows


# ---------------------------------------------------------------------------
# Exporter
# ---------------------------------------------------------------------------

def write_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    columns: list[tuple[str, str, str]],
) -> None:
    """Write and style the header row. Standalone so any exporter can reuse it."""
    for col_idx, (label, _attr, _dtype) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False
        )
        cell.border = CELL_BORDER


def write_data(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    columns: list[tuple[str, str, str]],
    rows: Sequence[object],
) -> None:
    """Write all data rows with appropriate formatting. Standalone so any exporter can reuse it."""
    description_col_idx = next(
        (i + 1 for i, (_, attr, _) in enumerate(columns) if attr == "article_description"),
        None,
    )

    for row_idx, order_row in enumerate(rows, start=2):
        for col_idx, (_, attr, dtype) in enumerate(columns, start=1):
            raw_value = getattr(order_row, attr, "")
            cell = ws.cell(row=row_idx, column=col_idx)

            cell.font = BODY_FONT
            cell.border = CELL_BORDER

            if dtype == "text":
                cell.value = str(raw_value) if raw_value not in (None, "") else ""
                cell.number_format = TEXT_FORMAT
                if col_idx == description_col_idx:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center"
                    )

            elif dtype == "integer":
                cell.value = int(raw_value) if raw_value is not None else None
                cell.number_format = INTEGER_FORMAT
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            elif dtype == "number":
                cell.value = float(raw_value) if raw_value is not None else None
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(
                    horizontal="right", vertical="center"
                )

            elif dtype == "date":
                if isinstance(raw_value, datetime):
                    # A real date we generated ourselves (e.g. Ordini
                    # ELVY's CONSEGNA) — write as an actual Excel date.
                    cell.value = raw_value
                    cell.number_format = "DD/MM/YYYY"
                else:
                    # Store as text; Excel will display as-is (dates
                    # from PDFs vary too much to parse reliably)
                    cell.value = str(raw_value) if raw_value not in (None, "") else ""
                    cell.number_format = TEXT_FORMAT
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

        # Taller rows for wrapped description text
        if description_col_idx:
            desc_val = getattr(order_row, "article_description", "")
            if desc_val and "\n" in str(desc_val):
                ws.row_dimensions[row_idx].height = ROW_HEIGHT_WRAPPED


def apply_column_widths(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    columns: list[tuple[str, str, str]],
) -> None:
    """Auto-size each column based on its content. Standalone so any exporter can reuse it."""
    for col_idx, (label, attr, _) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)

        if attr == "article_description":
            ws.column_dimensions[col_letter].width = DESCRIPTION_COL_WIDTH
            continue

        # Measure the widest cell value in this column
        max_width = len(label)
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                if cell.value is not None:
                    max_width = max(max_width, len(str(cell.value)))

        width = max(MIN_COL_WIDTH, min(max_width + 4, MAX_COL_WIDTH))
        ws.column_dimensions[col_letter].width = width


def freeze_and_filter(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    columns: list[tuple[str, str, str]],
) -> None:
    """Freeze the header row and enable AutoFilter. Standalone so any exporter can reuse it."""
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


class _LegacyMethodShims:
    """Thin instance-method wrappers kept so ExcelExporter's existing calls
    (self._write_header(...) etc.) keep working unchanged."""

    def _write_header(self, ws, columns):
        write_header(ws, columns)

    def _write_data(self, ws, columns, rows):
        write_data(ws, columns, rows)

    def _apply_column_widths(self, ws, columns):
        apply_column_widths(ws, columns)

    def _freeze_and_filter(self, ws, columns):
        freeze_and_filter(ws, columns)


class ExcelExporter(_LegacyMethodShims):
    """
    Writes a list of :class:`~pdf_parser.OrderRow` objects to an Excel file.

    Usage::

        exporter = ExcelExporter(rows, output_path=Path("output.xlsx"))
        exporter.export()
    """

    def __init__(self, rows: Sequence[OrderRow], output_path: Path,
                 magazino_summary=None, codes_map: dict | None = None) -> None:
        self.rows = rows
        self.output_path = output_path
        self.magazino_summary = magazino_summary
        self.codes_map = codes_map

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def export(self) -> None:
        """Build and save the Excel workbook."""
        logger.info("Exporting %d rows → %s", len(self.rows), self.output_path.name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Orders"
        ws.sheet_view.rightToLeft = False

        self._write_header(ws, COLUMNS)
        self._write_data(ws, COLUMNS, self.rows)
        self._apply_column_widths(ws, COLUMNS)
        self._freeze_and_filter(ws, COLUMNS)

        ordini_rows = build_ordini_elvy_rows(list(self.rows))

        raw_yarn_matches = None
        if self.magazino_summary is not None and not self.magazino_summary.empty:
            raw_yarn_matches = match_raw_yarn(ordini_rows, self.magazino_summary, self.codes_map)
            logger.info("Raw yarn matched: %d batch(es) assigned", len(raw_yarn_matches))

        ws2 = wb.create_sheet("Ordini ELVY")
        ws2.sheet_view.rightToLeft = False
        self._write_header(ws2, ORDINI_ELVY_COLUMNS)
        self._write_data(ws2, ORDINI_ELVY_COLUMNS, ordini_rows)
        self._apply_column_widths(ws2, ORDINI_ELVY_COLUMNS)
        self._freeze_and_filter(ws2, ORDINI_ELVY_COLUMNS)

        if raw_yarn_matches is not None:
            ws3 = wb.create_sheet("Filato x Tinturia")
            ws3.sheet_view.rightToLeft = False
            self._write_header(ws3, FILATO_TINTURIA_COLUMNS)
            self._write_data(ws3, FILATO_TINTURIA_COLUMNS, raw_yarn_matches)
            self._apply_column_widths(ws3, FILATO_TINTURIA_COLUMNS)
            self._freeze_and_filter(ws3, FILATO_TINTURIA_COLUMNS)

            label_col_idx = next(
                (i + 1 for i, (_, attr, _) in enumerate(FILATO_TINTURIA_COLUMNS) if attr == "label"), None
            )
            if label_col_idx:
                bold_label_font = Font(bold=True, name="Calibri", size=11)
                for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row,
                                          min_col=label_col_idx, max_col=label_col_idx):
                    for cell in row:
                        cell.font = bold_label_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        logger.info("Export completed: %s", self.output_path)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------


