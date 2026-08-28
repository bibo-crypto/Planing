from pathlib import Path
import tempfile
import openpyxl

import path_manager
from ordine_med import (
    FilatoAvailabilityRow,
    OrdineMedRow,
    export_erp_order_workbook,
    export_filato_availability_workbook,
)


def main() -> None:
    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        source = root / "Produzione.xlsx"
        path_manager.save_source("data_prod", source)
        assert path_manager.load_source("produzione")["source_path"] == str(source)

        record = OrdineMedRow(
            riga=1, code_org="C130", titolo="Titolo", descr_col="Blu",
            articolo="G130", colore="01", rocc=3, abbin="", consegna_input=None,
            pt_grg="12", pt_med="", polmoni="", cliente_note="", nota_grg="",
            nota_col="", kg_note="", fabb=None, prezz_note=None,
        )
        erp_output = root / "ERP" / "Ordine_MED_ERP.xlsx"
        export_erp_order_workbook(erp_output, [record])
        erp_workbook = openpyxl.load_workbook(erp_output, read_only=True)
        assert erp_workbook.active.title == "Dati sistema (B-N)"
        assert erp_workbook.active.max_row == 2
        erp_workbook.close()

        output = root / "ERP" / "Filato X Tinturia.xlsx"
        export_filato_availability_workbook(
            output,
            [FilatoAvailabilityRow("G130", "Titolo", 12, 3, 1.2, 2, -1, "NO")],
        )
        assert output.is_file()
        workbook = openpyxl.load_workbook(output, read_only=True)
        assert workbook.active.title == "Filato X Tinturia"
        assert workbook.active.max_row == 2
        workbook.close()


if __name__ == "__main__":
    main()
    print("focused_checks=ok")
