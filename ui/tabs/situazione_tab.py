"""
situazione_tab.py
Dyeing Situation Tracker — embedded as the "Situazione" tab of the Delta
Dyeing PDF/Excel converter. Replaces the old Old Situazione / WOORKSHEET /
New Situazione Excel + Power Query chain: each source is uploaded and
validated separately, and per-Partita Old/New Comment history is kept in
a local SQLite database (see situazione_db.py) instead of copy-pasted
sheets.
"""
import os
import json
import re
import threading
import tkinter as tk
from pathlib import Path
from tkinter import ttk, filedialog, messagebox
from tkinter import font as tkfont
from datetime import datetime, timedelta
from typing import Callable
import pandas as pd

import situazione_db as db
import situazione_loaders as data_loaders
import situazione_logic as business_logic
from abbina_suggestions import build_suggestions
from yarn_shortage_tab import YarnShortageTab
from dfm_lookup import build_dfm_lookup, load_dfm_cache, save_dfm_cache
from prod_lookup import load_prod_cache, save_prod_cache
from utils import logger

STATUS_COLORS = {
    "Filato": "#e0e0e0",
    "pronto da spedire": "#c6efce",
    "Spedita": "#bdd7ee",
    "spedita": "#bdd7ee",
    "Check": "#ffc7ce",
    "C.Q": "#fff2cc",
}


def color_for_status(status):
    if not status:
        return "#ffffff"
    if status.startswith("Ritinta"):
        return "#d9c6f0"
    return STATUS_COLORS.get(status, "#ffffff")


SOURCE_ORDER = ["copertura", "data_prod", "dfm", "wincoint", "uscita", "qualita"]

SOURCE_BUTTON_NAMES = {
    "copertura": "Copertura",
    "data_prod": "Produzione",
    "dfm": "DFM",
    "wincoint": "Wincoint",
    "uscita": "Uscita",
    "qualita": "Qualita",
    "codes": "Articoli",
}

# (internal_key, header shown in the app / exported file, cell type)
# Order and headers match the real "New Situazione" sheet exactly.
COLUMN_SPEC = [
    ("cliente", "CLIENTE", "text"),
    ("articolo", "Articolo", "text"),
    ("titolo", "Titolo", "text"),
    ("codice", "Codice", "number"),
    ("colore", "Colore", "text"),
    ("ordine", "Ordine", "text"),
    ("riga", "Riga", "number"),
    ("data", "Data", "date"),
    ("consegna", "Consegna", "date"),
    ("partita", "Partita", "number"),
    ("rocche", "Rocche", "number"),
    ("mc", "M/C", "number"),
    ("comment", "Comment", "text"),
    ("raw_yarn_match", "Filato Disponibile", "text"),
    ("cq", "C.Q", "text"),
    ("tinto", "Tinto", "date"),
    ("bagno", "Bagno", "text"),
    ("old_comment", "Old Comm.", "text"),
    ("new_comment", "New Comm.", "text"),
    ("planedate", "PlaneDate", "text"),
    ("data_qualita", "Data Qualita", "date"),
    ("data_uscita", "Data Uscita", "date"),
    ("custom", "Custom", "text"),
    ("days_in_qc", "Days in Q.C", "number"),
]
COLUMN_LABELS = {key: header for key, header, _ in COLUMN_SPEC}
COLUMN_TYPES = {key: ctype for key, _, ctype in COLUMN_SPEC}


class SourceRow(ttk.Frame):
    """Compact upload card: file name button + current upload status."""

    def __init__(self, master, key, label, on_upload):
        super().__init__(master)
        self.key = key
        self.on_upload = on_upload

        self.status_var = tk.StringVar(value="Not uploaded")
        self.dot = tk.Label(self, text="●", fg="#9e9e9e", bg="#f4f6f8", font=("Segoe UI", 11))
        self.dot.grid(row=1, column=0, padx=(4, 2), sticky="w")

        button_name = SOURCE_BUTTON_NAMES.get(key, label)
        self.button = ttk.Button(
            self,
            text=f"Upload {button_name}",
            command=self._browse,
            width=15,
            font=("Segoe UI", 9, "bold"),
            height=30,
        )
        self.button.grid(row=0, column=0, columnspan=2, padx=4, pady=(3, 2), sticky="ew")

        self.status_lbl = ttk.Label(self, textvariable=self.status_var, anchor="w",
                                    foreground="#667085", width=14)
        self.status_lbl.grid(row=1, column=1, padx=(2, 4), sticky="w")

        self.columnconfigure(1, weight=1)

    def _browse(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return
        self.on_upload(self.key, path)

    def set_status(self, ok, message):
        self.dot.config(fg="#2e7d32" if ok else "#c62828")
        self.status_var.set(message)


class SituazioneTab(ttk.Frame):
    """Embeddable 'Situazione' tab — hosted inside gui.py's main Notebook."""

    def __init__(self, master, on_shared_cache_changed: Callable[[], None] | None = None):
        super().__init__(master)

        self._configure_styles()

        self._on_shared_cache_changed = on_shared_cache_changed
        # Set post-construction from gui.py once MagazinoFilatoTab exists
        # (it's built after this tab). Used only to auto-fill the "Filato
        # Disponibile" column -- read lazily, so it's fine if it's not set
        # yet the first time _load_table_from_db() runs.
        self.magazino_tab = None

        db.init_db()

        self.loaded_frames = {}   # key -> DataFrame (validated, ready to merge)
        self.sort_state = {}      # column -> ascending bool
        self.current_df = pd.DataFrame()
        self._filter_after_id = None
        self._shared_syncing = False
        self._startup_restore_in_progress = False
        self._startup_snapshot_current = False
        self._table_loaded_callbacks = []
        self._data_revision = 0
        self._tree_render_generation = 0
        self._tree_render_after_id = None
        self._shared_dfm_path = ""
        self._shared_prod_path = ""
        self._copertura_revision = 0
        self._child_windows = {}

        self._build_upload_panel()
        self._build_toolbar()
        self._build_treeview()
        self._refresh_source_labels_from_db()
        # Show the last SQLite snapshot immediately.  Excel files are restored
        # later in a worker, so the application opens against the cache first
        # and refreshes when the latest source files have finished loading.
        self._load_table_from_db()
        self.after(700, self._auto_restore_saved_files)
        self.after(1000, self.sync_shared_async)

    # ------------------------------------------------------------------ UI
    def _configure_styles(self):
        # Note: the app's main window (ui/gui.py) already sets the global
        # ttk theme to "clam" once at startup for the same reason (the
        # native Windows theme ignores Treeview heading background colors).
        # Calling it again here is redundant but harmless (idempotent) —
        # kept as a safety net in case this tab is ever instantiated on its
        # own, outside the main app.
        style = ttk.Style(self)
        if "clam" in style.theme_names():
            style.theme_use("clam")
        style.configure("Upload.TLabelframe", background="#f4f6f8", borderwidth=1, relief="solid")
        style.configure("Upload.TLabelframe.Label", background="#f4f6f8", foreground="#344054",
                        font=("Segoe UI", 10, "bold"))
        style.configure("Situazione.Treeview", background="#ffffff", fieldbackground="#ffffff",
                        foreground="#1d2939", rowheight=29, borderwidth=1, relief="solid",
                        bordercolor="#b8c6d6", lightcolor="#b8c6d6", darkcolor="#b8c6d6",
                        font=("Segoe UI", 9))
        style.configure("Situazione.Treeview.Heading", background="#16324f", foreground="#ffffff",
                        relief="raised", borderwidth=1, padding=(8, 8),
                        font=("Segoe UI", 9, "bold"))
        style.map("Situazione.Treeview.Heading",
                  foreground=[("pressed", "#ffffff"), ("active", "#ffffff"), ("!active", "#ffffff")],
                  background=[("pressed", "#0b2239"), ("active", "#244b70"), ("!active", "#16324f")])
        style.map("Situazione.Treeview", background=[("selected", "#2563eb")],
                  foreground=[("selected", "#ffffff")])

    def _build_upload_panel(self):
        upload_area = ttk.Frame(self)
        upload_area.pack(side="top", fill="x", anchor="w", padx=8, pady=(6, 2))

        panel = ttk.LabelFrame(upload_area, text="1) Required files", style="Upload.TLabelframe")
        panel.pack(side="left", anchor="nw")
        panel.configure(padding=(5, 3))

        self.source_rows = {}
        for key in SOURCE_ORDER:
            label, _ = data_loaders.LOADERS[key]
            row = SourceRow(panel, key, label, self._handle_upload)
            row.grid(row=0, column=SOURCE_ORDER.index(key), padx=3, pady=1, sticky="nw")
            self.source_rows[key] = row

        codes_panel = ttk.LabelFrame(upload_area, text="Optional file", style="Upload.TLabelframe")
        codes_panel.pack(side="left", anchor="nw", padx=(8, 0), fill="y")
        codes_panel.configure(padding=(5, 3))

        codes_row = SourceRow(codes_panel, "codes", "Yarn codes (Articoli) - optional",
                               self._handle_codes_upload)
        codes_row.grid(row=0, column=0, padx=3, pady=1, sticky="nw")
        self.codes_row = codes_row

    def _build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(side="top", fill="x", padx=8, pady=(3, 4))

        self.refresh_btn = ttk.Button(bar, text="2) Refresh", command=self._on_refresh)
        self.refresh_btn.pack(side="left", padx=4)

        self.upload_data_btn = ttk.Button(bar, text="Upload Data", command=self._on_upload_data)
        self.upload_data_btn.pack(side="left", padx=4)

        self.export_btn = ttk.Button(bar, text="Export to Excel", command=self._on_export)
        self.export_btn.pack(side="left", padx=4)

        self.abbina_btn = ttk.Button(bar, text="Da abbinare", command=self._open_abbina)
        self.abbina_btn.pack(side="left", padx=4)

        self.yarn_shortage_btn = ttk.Button(bar, text="Mancanza Filato", command=self._open_yarn_shortage)
        self.yarn_shortage_btn.pack(side="left", padx=4)

        self.copertura_btn = ttk.Button(bar, text="Copertura", command=self._open_copertura)
        self.copertura_btn.pack(side="left", padx=4)

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search_changed)
        ttk.Label(bar, text="Search:").pack(side="left", padx=(18, 4))
        ttk.Entry(bar, textvariable=self.search_var, width=30).pack(side="left")

        self.summary_lbl = ttk.Label(bar, text="")
        self.summary_lbl.pack(side="right", padx=8)

    def _open_copertura(self):
        """Show and export the dyeing coverage summary for machines 3–12."""
        if self._focus_child_window("copertura"):
            return
        copertura = self.loaded_frames.get("copertura")
        if not isinstance(copertura, pd.DataFrame) or copertura.empty:
            messagebox.showinfo("Copertura", "Upload the Copertura file first.", parent=self)
            return
        if "machine" not in copertura.columns or self.current_df.empty:
            messagebox.showinfo("Copertura", "Refresh Situazione data first.", parent=self)
            return

        situation = self.current_df.copy()

        def bagno_key(value):
            digits = re.sub(r"\D", "", str(value or "")).lstrip("0")
            return digits or str(value or "").strip().casefold()

        for frame in (situation, copertura):
            frame["bagno"] = frame["bagno"].fillna("").astype(str).str.strip()
            frame["bagno_key"] = frame["bagno"].map(bagno_key)
        merged = situation.merge(
            copertura[["bagno_key", "machine"]].drop_duplicates("bagno_key"),
            on="bagno_key", how="inner",
        )

        def machine_number(value):
            text = str(value).strip()
            digits = re.sub(r"\D", "", text)
            if digits and 3300 <= int(digits) <= 3399:
                return int(digits) - 3300
            match = re.search(r"(?<!\d)0*(1[0-2]|[3-9])(?:\.0+)?(?!\d)", text)
            return int(match.group(1)) if match else None

        merged["machine_number"] = merged["machine"].map(machine_number)
        merged = merged[merged["machine_number"].between(3, 12, inclusive="both")]
        if merged.empty:
            messagebox.showinfo("Copertura", "No Situazione colours match machines 3–12.", parent=self)
            return

        def coverage_until(count):
            if not count:
                return "-"
            target_days = (int(count) + 1) // 2
            day, done = datetime.now().date(), 0
            while done < target_days:
                if day.weekday() != 4:
                    done += 1
                if done < target_days:
                    day += timedelta(days=1)
            return day.strftime("%Y-%m-%d")

        def build_summary(view):
            columns = ["machine_number", "cliente", "total_colors", "pgx", "available", "covered_until"]
            if view.empty:
                return pd.DataFrame(columns=columns)
            work = view.copy()
            work["is_pgx"] = work["comment"].fillna("").astype(str).str.upper().str.startswith("PG-X")
            summary = work.groupby(["machine_number", "cliente"], dropna=False).agg(
                total_colors=("colore", "size"), pgx=("is_pgx", "sum"),
            ).reset_index()
            summary["available"] = summary["total_colors"] - summary["pgx"]
            totals = summary.groupby("machine_number")["total_colors"].sum().to_dict()
            summary["covered_until"] = summary["machine_number"].map(lambda x: coverage_until(totals.get(x, 0)))
            return summary

        window = tk.Toplevel(self)
        self._child_windows["copertura"] = window
        window.title("Copertura — macchine 3–12")
        window.geometry("1050x600")
        window.minsize(800, 450)
        # Keep the normal Windows title bar so the native minimize, maximize,
        # restore, and close buttons remain available beside X.
        window.resizable(True, True)
        ttk.Label(window, text="Copertura: 2 colori al giorno per macchina — venerdì escluso", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10, pady=(10, 4))
        search_row = ttk.Frame(window)
        search_row.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Label(search_row, text="Filtra per:").pack(side="left", padx=(0, 6))
        filter_kind = tk.StringVar(value="Tutti")
        filter_value = tk.StringVar(value="Tutti")
        kind_combo = ttk.Combobox(
            search_row, textvariable=filter_kind, state="readonly", width=14,
            values=("Tutti", "Macchina", "Cliente"),
        )
        kind_combo.pack(side="left", padx=(0, 6))
        value_combo = ttk.Combobox(search_row, textvariable=filter_value, state="readonly", width=24)
        value_combo.pack(side="left")
        status = ttk.Label(search_row, text="")
        status.pack(side="right")

        frame = ttk.Frame(window)
        frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        columns = ("machine", "cliente", "colors", "pgx", "available", "covered_until")
        labels = {"machine": "Macchina", "cliente": "Cliente", "colors": "Totale colori", "pgx": "Manca Filato (PG-X)", "available": "Disponibile", "covered_until": "Coperta fino al"}
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        widths = [85, 130, 110, 140, 105, 150]
        for column, width in zip(columns, widths):
            tree.heading(column, text=labels[column])
            tree.column(column, width=width, anchor="center")
        yscroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        tree.tag_configure("total", background="#e8f1fb", font=("Segoe UI", 9, "bold"))

        def current_view():
            selected_kind = filter_kind.get()
            selected_value = filter_value.get()
            if selected_kind == "Tutti" or selected_value == "Tutti":
                return merged
            if selected_kind == "Macchina":
                return merged.loc[merged["machine_number"].astype(str).eq(selected_value)]
            return merged.loc[merged["cliente"].astype(str).eq(selected_value)]

        def refresh_filter_values(*_args):
            selected_kind = filter_kind.get()
            if selected_kind == "Macchina":
                values = ["Tutti"] + [str(n) for n in sorted(merged["machine_number"].dropna().unique())]
            elif selected_kind == "Cliente":
                values = ["Tutti"] + sorted(merged["cliente"].dropna().astype(str).unique())
            else:
                values = ["Tutti"]
            value_combo.configure(values=values)
            filter_value.set("Tutti")
            render()

        def render():
            view = current_view()
            summary = build_summary(view)
            tree.delete(*tree.get_children())
            machines = range(3, 13)
            if filter_kind.get() == "Macchina" and filter_value.get() != "Tutti":
                machines = [int(filter_value.get())]
            for machine in machines:
                rows = summary[summary["machine_number"] == machine]
                total = int(rows["total_colors"].sum()) if not rows.empty else 0
                until = coverage_until(total)
                if rows.empty:
                    tree.insert("", "end", values=(machine, "-", 0, 0, 0, "-"))
                else:
                    for _, row in rows.sort_values("cliente").iterrows():
                        tree.insert("", "end", values=(machine, row["cliente"], int(row["total_colors"]), int(row["pgx"]), int(row["available"]), row["covered_until"]))
                    tree.insert("", "end", values=(machine, "TOTALE", total, int(rows["pgx"].sum()), total - int(rows["pgx"].sum()), until), tags=("total",))
            status.config(text=f"{len(view):,} colori")

        def export_summary():
            path = filedialog.asksaveasfilename(parent=window, title="Esporta Copertura", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="copertura.xlsx")
            if not path:
                return
            view = current_view()
            summary = build_summary(view)
            rows = []
            machines = range(3, 13)
            if filter_kind.get() == "Macchina" and filter_value.get() != "Tutti":
                machines = [int(filter_value.get())]
            for machine in machines:
                part = summary[summary["machine_number"] == machine]
                total = int(part["total_colors"].sum()) if not part.empty else 0
                if part.empty:
                        rows.append([machine, "-", 0, 0, 0, "-"])
                else:
                    for _, row in part.sort_values("cliente").iterrows():
                        rows.append([machine, row["cliente"], int(row["total_colors"]), int(row["pgx"]), int(row["available"]), row["covered_until"]])
                    pgx_total = int(part["pgx"].sum())
                    rows.append([machine, "TOTALE", total, pgx_total, total - pgx_total, coverage_until(total)])
            try:
                pd.DataFrame(rows, columns=[labels[c] for c in columns]).to_excel(path, index=False, sheet_name="Copertura")
                messagebox.showinfo("Copertura", f"Export completato:\n{path}", parent=window)
            except Exception as exc:  # noqa: BLE001
                messagebox.showerror("Copertura", str(exc), parent=window)

        kind_combo.bind("<<ComboboxSelected>>", refresh_filter_values)
        value_combo.bind("<<ComboboxSelected>>", lambda _event: render())
        buttons = ttk.Frame(window)
        buttons.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(buttons, text="Estrai Excel", command=export_summary).pack(side="right", padx=3)
        refresh_filter_values()
        render()

    def _focus_child_window(self, key):
        """Focus an already-open child window instead of opening a duplicate."""
        window = self._child_windows.get(key)
        if window is None:
            return False
        try:
            if not window.winfo_exists():
                self._child_windows.pop(key, None)
                return False
            if window.state() == "iconic":
                window.deiconify()
            window.lift()
            window.focus_force()
            return True
        except tk.TclError:
            self._child_windows.pop(key, None)
            return False

    def _build_treeview(self):
        cols = [key for key, _, _ in COLUMN_SPEC]
        self.columns = cols

        frame = ttk.Frame(self, borderwidth=1, relief="solid")
        frame.pack(side="top", fill="both", expand=True, padx=8, pady=(2, 8))

        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse",
                                 style="Situazione.Treeview")
        for c in cols:
            ctype = COLUMN_TYPES.get(c, "text")
            anchor = "center" if ctype in ("number", "date") else "w"
            self.tree.heading(c, text=COLUMN_LABELS.get(c, c), anchor="center",
                              command=lambda c=c: self._sort_by(c))
            self.tree.column(c, width=100, minwidth=60, anchor=anchor, stretch=False)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        for status, color in STATUS_COLORS.items():
            self.tree.tag_configure(status, background=color)
        self.tree.tag_configure("Ritinta", background="#d9c6f0")
        self.tree.tag_configure("stripe", background="#f3f6fa")
        self.tree.tag_configure("normal", background="#ffffff")

    def _autosize_columns(self, df):
        """Fit columns to visible content while keeping the table usable."""
        body_font = tkfont.Font(family="Segoe UI", size=9)
        heading_font = tkfont.Font(family="Segoe UI", size=9, weight="bold")
        for col in self.columns:
            values = []
            if not df.empty and col in df.columns:
                values = [str(value) for value in df[col].fillna("").tolist()]
            # Measuring only the longest string is much faster than measuring
            # every cell, especially for large Excel exports.
            longest_value = max(values, key=len, default="")
            longest = max(heading_font.measure(COLUMN_LABELS.get(col, col)), body_font.measure(longest_value))
            width = max(65, min(longest + 22, 300))
            self.tree.column(col, width=width, minwidth=min(width, 65), stretch=False)

    # ------------------------------------------------------------- shared DFM / uploads
    def _on_upload_data(self):
        """Reload every previously uploaded source from its saved file path."""
        uploads = db.get_all_uploads()
        reloaded = []
        missing = []

        for key in SOURCE_ORDER:
            info = uploads.get(key, {})
            path = info.get("file_path", "")
            if not path:
                missing.append(SOURCE_BUTTON_NAMES[key] + " (not uploaded yet)")
            elif not os.path.isfile(path):
                missing.append(f"{SOURCE_BUTTON_NAMES[key]} ({path})")
            else:
                self._handle_upload(key, path)
                reloaded.append(SOURCE_BUTTON_NAMES[key])

        codes_info = uploads.get("codes", {})
        codes_path = codes_info.get("file_path", "")
        if not codes_path:
            missing.append("Articoli (not uploaded yet)")
        elif not os.path.isfile(codes_path):
            missing.append(f"Articoli ({codes_path})")
        else:
            self._handle_codes_upload("codes", codes_path)
            reloaded.append("Articoli")

        summary = []
        if reloaded:
            summary.append("Reloaded: " + ", ".join(reloaded))
        if missing:
            summary.append("Not available:\n- " + "\n- ".join(missing))
        messagebox.showinfo("Upload Data", "\n\n".join(summary) or "No saved file paths found.")

    def _auto_restore_saved_files(self):
        """Restore saved upload paths and refresh once at application start.

        Loading is done off the Tk thread because the source workbooks can be
        large.  The automatic refresh deliberately preserves comment history:
        opening the program must rebuild the table, not advance New Comm. to
        Old Comm.
        """
        if getattr(self, "_auto_restore_started", False):
            return
        self._auto_restore_started = True
        uploads = db.get_all_uploads()

        # The SQLite snapshot is already the fast local cache for the
        # Situazione grid. If none of the saved source workbooks changed since
        # their last upload, do not parse all six Excel files on every startup.
        # The user can still use Upload Data when a fresh rebuild is needed.
        if self._saved_snapshot_is_current(uploads):
            self._startup_snapshot_current = True
            # The SQLite snapshot is enough for the main grid, but the
            # Copertura dashboard also needs the physical machine column.
            self._restore_saved_copertura(uploads)
            logger.info("Situazione: startup snapshot is current; skipped Excel restore")
            return

        paths = {
            key: str(uploads.get(key, {}).get("file_path", ""))
            for key in SOURCE_ORDER + ["codes"]
        }
        if not any(paths.values()):
            return

        self._startup_restore_in_progress = True

        def worker():
            loaded = {}
            errors = {}
            for key in SOURCE_ORDER:
                path = paths.get(key, "")
                if not path:
                    errors[key] = "not saved"
                    continue
                if not os.path.isfile(path):
                    errors[key] = f"file not found: {path}"
                    continue
                try:
                    df, load_errors = data_loaders.LOADERS[key][1](path)
                    if load_errors or df is None or df.empty:
                        errors[key] = "; ".join(load_errors) if load_errors else "file is empty"
                    else:
                        loaded[key] = df
                except Exception as exc:  # noqa: BLE001
                    errors[key] = str(exc)

            codes_df = None
            codes_error = None
            codes_path = paths.get("codes", "")
            if codes_path:
                if os.path.isfile(codes_path):
                    try:
                        codes_df, load_errors = data_loaders.load_codes(codes_path)
                        if load_errors or codes_df is None or codes_df.empty:
                            codes_error = "; ".join(load_errors) if load_errors else "file is empty"
                    except Exception as exc:  # noqa: BLE001
                        codes_error = str(exc)
                else:
                    codes_error = f"file not found: {codes_path}"

            def apply_result():
                self._startup_restore_in_progress = False
                for key, df in loaded.items():
                    self.loaded_frames[key] = df
                    self.source_rows[key].set_status(True, f"✅ {len(df)} rows")
                    db.save_upload(key, os.path.basename(paths[key]), len(df), "ok",
                                   f"✅ {len(df)} rows - {os.path.basename(paths[key])}",
                                   file_path=paths[key])
                for key, error in errors.items():
                    if key in self.source_rows:
                        self.source_rows[key].set_status(False, f"❌ {error}")
                if codes_df is not None and not codes_error:
                    db.save_codes(codes_df)
                    self.codes_row.set_status(True, f"✅ Saved ({len(codes_df)} codes)")
                    db.save_upload("codes", os.path.basename(codes_path), len(codes_df), "ok",
                                   f"✅ Saved ({len(codes_df)} codes) - {os.path.basename(codes_path)}",
                                   file_path=codes_path)
                elif codes_error:
                    self.codes_row.set_status(False, f"❌ {codes_error}")

                if not errors:
                    self._on_refresh(preserve_comment_history=True)
                else:
                    logger.warning("Situazione: automatic restore skipped refresh; missing/invalid files: %s",
                                   ", ".join(errors))

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def _restore_saved_copertura(self, uploads):
        """Restore Copertura even when the main Situazione snapshot is current."""
        info = uploads.get("copertura", {}) if isinstance(uploads, dict) else {}
        path = str(info.get("file_path", ""))
        if not path or not os.path.isfile(path) or "copertura" in self.loaded_frames:
            return

        def worker():
            try:
                df, errors = data_loaders.load_schedulato(path)
            except Exception as exc:  # noqa: BLE001
                df, errors = None, [str(exc)]

            def apply_result():
                if errors or df is None or df.empty:
                    logger.warning("Situazione: saved Copertura restore failed: %s", errors)
                    return
                self.loaded_frames["copertura"] = df
                self._copertura_revision += 1
                self.source_rows["copertura"].set_status(True, f"✅ {len(df)} rows")

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    @staticmethod
    def _saved_snapshot_is_current(uploads):
        """Return True when the SQLite table can be used without Excel I/O."""
        if not db.get_all_states():
            return False
        for key in SOURCE_ORDER:
            info = uploads.get(key, {})
            path = info.get("file_path", "")
            uploaded_at = info.get("uploaded_at", "")
            if info.get("status") != "ok" or not path or not os.path.isfile(path) or not uploaded_at:
                return False
            try:
                uploaded_timestamp = datetime.fromisoformat(str(uploaded_at)).timestamp()
                if os.path.getmtime(path) > uploaded_timestamp + 1:
                    return False
            except (OSError, TypeError, ValueError):
                return False
        return True

    def sync_shared_dfm(self):
        """Load the DFM selected in either page from the shared persistent cache."""
        cache = load_dfm_cache()
        source_path = Path(str(cache.get("source_path", "")))
        if not source_path.is_file():
            return
        if getattr(self, "_shared_dfm_path", "") == str(source_path):
            return

        try:
            df, errors = data_loaders.load_dfm(str(source_path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not restore shared DFM file: %s", exc)
            return
        if errors or df is None or df.empty:
            return

        self.loaded_frames["dfm"] = df
        self._shared_dfm_path = str(source_path)
        msg = f"✅ {len(df)} rows - {source_path.name}"
        self.source_rows["dfm"].set_status(True, f"✅ {len(df)} rows")
        db.save_upload("dfm", source_path.name, len(df), "ok", msg, file_path=str(source_path))

    def sync_shared_async(self):
        """Restore shared Excel files without blocking the Tk event loop."""
        # Startup restore owns the source loading pass. Running this second
        # pass at the same time would read DFM/Produzione twice.
        if self._startup_restore_in_progress:
            return
        if self._startup_snapshot_current:
            self._startup_snapshot_current = False
            return
        if self._shared_syncing:
            return
        dfm_path = str(load_dfm_cache().get("source_path", ""))
        prod_path = str(load_prod_cache().get("source_path", ""))
        needs_dfm = bool(dfm_path and os.path.isfile(dfm_path) and self._shared_dfm_path != dfm_path)
        needs_prod = bool(prod_path and os.path.isfile(prod_path) and self._shared_prod_path != prod_path)
        if not (needs_dfm or needs_prod):
            return

        self._shared_syncing = True

        def worker():
            dfm_result = None
            prod_result = None
            if dfm_path and os.path.isfile(dfm_path) and self._shared_dfm_path != dfm_path:
                try:
                    dfm_result = data_loaders.load_dfm(dfm_path)
                except Exception as exc:  # noqa: BLE001
                    logger.warning("Could not restore shared DFM file: %s", exc)
            if prod_path and os.path.isfile(prod_path) and self._shared_prod_path != prod_path:
                try:
                    prod_result = data_loaders.load_data_prod(prod_path)
                except Exception as exc:  # noqa: BLE001
                    logger.warning("Could not restore shared Produzione file: %s", exc)

            def apply_result():
                self._shared_syncing = False
                if dfm_result and not dfm_result[1] and dfm_result[0] is not None and not dfm_result[0].empty:
                    df = dfm_result[0]
                    self.loaded_frames["dfm"] = df
                    self._shared_dfm_path = dfm_path
                    self.source_rows["dfm"].set_status(True, f"✅ {len(df)} rows")
                    db.save_upload("dfm", Path(dfm_path).name, len(df), "ok", f"✅ {len(df)} rows - {Path(dfm_path).name}", file_path=dfm_path)
                if prod_result and not prod_result[1] and prod_result[0] is not None and not prod_result[0].empty:
                    df = prod_result[0]
                    self.loaded_frames["data_prod"] = df
                    self._shared_prod_path = prod_path
                    self.source_rows["data_prod"].set_status(True, f"✅ {len(df)} rows")
                    db.save_upload("data_prod", Path(prod_path).name, len(df), "ok", f"✅ {len(df)} rows - {Path(prod_path).name}", file_path=prod_path)
                self.after_idle(self.sync_shared_async)

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def _save_shared_dfm(self, path):
        """Update the shared DFM cache when Situazione is the upload source."""
        try:
            entries = build_dfm_lookup(Path(path))
            if entries:
                save_dfm_cache(entries, Path(path).name, Path(path))
                self._shared_dfm_path = str(Path(path))
                if self._on_shared_cache_changed:
                    self._on_shared_cache_changed()
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not update shared DFM reference: %s", exc)

    def sync_shared_prod(self):
        """Load the Produzione file selected in either page from the shared cache."""
        cache = load_prod_cache()
        source_path = Path(str(cache.get("source_path", "")))
        if not source_path.is_file():
            return
        if getattr(self, "_shared_prod_path", "") == str(source_path):
            return

        try:
            df, errors = data_loaders.load_data_prod(str(source_path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not restore shared Produzione file: %s", exc)
            return
        if errors or df is None or df.empty:
            return

        self.loaded_frames["data_prod"] = df
        self._shared_prod_path = str(source_path)
        msg = f"✅ {len(df)} rows - {source_path.name}"
        self.source_rows["data_prod"].set_status(True, f"✅ {len(df)} rows")
        db.save_upload("data_prod", source_path.name, len(df), "ok", msg, file_path=str(source_path))

    def _save_shared_prod(self, path):
        """Update the shared Produzione cache when Situazione is the upload source."""
        try:
            save_prod_cache(Path(path))
            self._shared_prod_path = str(Path(path))
            if self._on_shared_cache_changed:
                self._on_shared_cache_changed()
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not update shared Produzione reference: %s", exc)

    def _handle_upload(self, key, path):
        label, loader_fn = data_loaders.LOADERS[key]
        try:
            df, errors = loader_fn(path)
        except Exception as exc:  # noqa: BLE001
            errors = [f"An error occurred while reading the file: {exc}"]
            df = None

        if errors or df is None or df.empty:
            msg = "; ".join(errors) if errors else "The file is empty after filtering"
            self.source_rows[key].set_status(False, f"❌ {msg}")
            db.save_upload(key, os.path.basename(path), 0, "error", msg, file_path=str(path))
            self.loaded_frames.pop(key, None)
            return

        self.loaded_frames[key] = df
        if key == "copertura":
            self._copertura_revision += 1
        msg = f"✅ {len(df)} rows - {os.path.basename(path)}"
        self.source_rows[key].set_status(True, f"✅ {len(df)} rows")
        db.save_upload(key, os.path.basename(path), len(df), "ok", msg, file_path=str(path))
        if key == "dfm":
            self._save_shared_dfm(path)
        if key == "data_prod":
            self._save_shared_prod(path)
        logger.info("Situazione: %s uploaded — %d rows (%s)", key, len(df), os.path.basename(path))

    def _handle_codes_upload(self, _key, path):
        try:
            df, errors = data_loaders.load_codes(path)
        except Exception as exc:  # noqa: BLE001
            errors = [f"An error occurred while reading the file: {exc}"]
            df = None

        if errors or df is None or df.empty:
            msg = "; ".join(errors) if errors else "The file is empty"
            self.codes_row.set_status(False, f"❌ {msg}")
            return

        db.save_codes(df)
        msg = f"✅ Saved ({len(df)} codes) - {os.path.basename(path)}"
        self.codes_row.set_status(True, msg)
        db.save_upload("codes", os.path.basename(path), len(df), "ok", msg, file_path=str(path))
        logger.info("Situazione: yarn codes reference updated — %d codes (%s)", len(df), os.path.basename(path))

    def _refresh_source_labels_from_db(self):
        uploads = db.get_all_uploads()
        for key, row_widget in self.source_rows.items():
            info = uploads.get(key)
            if info:
                ok = info["status"] == "ok"
                if ok and not info.get("file_path"):
                    row_widget.set_status(
                        False,
                        f"⚠ {info['message']} (select once to save path)",
                    )
                else:
                    count = info.get("row_count", "")
                    status_text = f"✅ {count} rows" if ok else f"❌ {info['message']}"
                    row_widget.set_status(ok, status_text)
        codes_info = uploads.get("codes")
        if codes_info:
            if codes_info["status"] == "ok" and not codes_info.get("file_path"):
                self.codes_row.set_status(False, "⚠ Saved status only (select once to save path)")
            else:
                self.codes_row.set_status(codes_info["status"] == "ok",
                                           f"✅ Saved ({codes_info.get('row_count', '')} codes)")

    # -------------------------------------------------------------- refresh
    def _on_refresh(self, preserve_comment_history=False):
        if "wincoint" not in self.loaded_frames:
            messagebox.showwarning("Missing data", "Upload the WINCOINT orders file before refreshing.")
            return

        missing = [SOURCE_BUTTON_NAMES[k] for k in SOURCE_ORDER if k not in self.loaded_frames]
        if missing:
            messagebox.showwarning(
                "Missing files",
                "These files have not been uploaded in this session:\n- " + "\n- ".join(missing) +
                "\n\nUpload them once, or use Upload Data after their paths have been saved. "
                "Refresh was cancelled so existing derived data is not cleared."
            )
            return

        # previous New Comment per Partita -- this becomes this round's Old Comment,
        # and the cascade in situazione_logic actively uses it, not just for history
        existing = db.get_all_states()
        old_comments = {p: s["new_comment"] for p, s in existing.items()}

        result_df = business_logic.compute_situation(
            orders_df=self.loaded_frames.get("wincoint"),
            dfm_df=self.loaded_frames.get("dfm"),
            data_prod_df=self.loaded_frames.get("data_prod"),
            copertura_df=self.loaded_frames.get("copertura"),
            uscita_df=self.loaded_frames.get("uscita"),
            qualita_df=self.loaded_frames.get("qualita"),
            codes_map=db.load_codes(),
            old_comments=old_comments,
        )

        # --- safety check: has ANYTHING changed vs what's already stored? ---
        any_change = False
        for _, r in result_df.iterrows():
            prev = existing.get(r["partita"])
            if prev is None or prev["new_comment"] != r["new_comment"]:
                any_change = True
                break

        current_partite = {
            str(value).strip()
            for value in result_df.get("partita", pd.Series(dtype=str)).tolist()
            if str(value).strip()
        }
        saved_partite = {str(value).strip() for value in existing}
        partite_changed = current_partite != saved_partite

        same_data = not any_change and not partite_changed and bool(existing)
        if preserve_comment_history or same_data:
            # The user may intentionally upload the same files again to
            # rebuild the Treeview.  Do not show a warning, and do not treat
            # the refresh as a new comment-history step: keep each row's
            # already stored Old Comm. exactly as it is.
            result_df["old_comment"] = result_df["partita"].map(
                lambda partita: existing.get(partita, {}).get("old_comment", "")
            )
        else:
            proceed = messagebox.askyesno(
                "Confirm refresh",
                f"{len(result_df)} batches will be updated. The current New Comment will move to Old Comment "
                "for batches whose status changed. Continue?"
            )
            if not proceed:
                return

        removed = db.remove_states_not_in(current_partite)
        added, updated, unchanged = db.upsert_states(result_df.to_dict(orient="records"))
        self.summary_lbl.config(text=f"Added: {added}  |  Updated: {updated}  |  Unchanged: {unchanged}")
        logger.info(
            "Situazione: refreshed — added=%d updated=%d unchanged=%d removed=%d",
            added, updated, unchanged, removed,
        )
        self._load_table_from_db()

    def _recompute_raw_yarn_match(self) -> None:
        """Fill "Filato Disponibile" for PG-X rows from Magazino Filato's
        current stock -- best-effort, never blocks: if Magazino hasn't been
        loaded yet this just leaves the column blank."""
        if self.current_df.empty:
            return
        if "comment" not in self.current_df.columns:
            self.current_df["raw_yarn_match"] = ""
            return
        magazino_tab = self.magazino_tab
        magazino_summary = getattr(magazino_tab, "magazino_summary", None) if magazino_tab else None
        lotti_summary = getattr(magazino_tab, "lotti_summary", None) if magazino_tab else None
        try:
            self.current_df["raw_yarn_match"] = business_logic.compute_raw_yarn_matches(
                self.current_df, magazino_summary, lotti_summary
            )
        except Exception as exc:  # noqa: BLE001
            logger.error("Situazione: raw yarn auto-match failed: %s", exc)
            self.current_df["raw_yarn_match"] = ""

    def refresh_raw_yarn_match(self) -> None:
        """Public hook: re-run the Filato Disponibile match against whatever
        Magazino Filato/LOTTI data is loaded right now, and re-render."""
        if self.current_df.empty:
            return
        self._recompute_raw_yarn_match()
        self._data_revision += 1
        self._render_tree(self.current_df)

    def _load_table_from_db(self):
        states = db.get_all_states()
        self.current_df = pd.DataFrame(states.values())
        self._data_revision += 1
        if not self.current_df.empty and "bagno" in self.current_df.columns:
            self.current_df = self.current_df.sort_values(
                by="bagno", ascending=True, key=lambda s: s.astype(str)
            )
            self.sort_state["bagno"] = False  # next click on Bagno heading reverses to Z-A
        self._recompute_raw_yarn_match()
        self._render_tree(self.current_df)
        for callback in tuple(self._table_loaded_callbacks):
            try:
                callback()
            except Exception as exc:  # noqa: BLE001
                logger.warning("Situazione: could not update dependent view: %s", exc)

    def add_table_loaded_callback(self, callback):
        """Register a callback invoked after the Situazione Treeview reloads."""
        if callback not in self._table_loaded_callbacks:
            self._table_loaded_callbacks.append(callback)

    # -------------------------------------------------------------- display
    def _render_tree(self, df, autosize=True):
        """Render rows in small UI batches so large snapshots do not freeze Tk."""
        self._tree_render_generation += 1
        generation = self._tree_render_generation
        if self._tree_render_after_id is not None:
            try:
                self.after_cancel(self._tree_render_after_id)
            except tk.TclError:
                pass
            self._tree_render_after_id = None

        self.tree.delete(*self.tree.get_children())
        if autosize:
            self._autosize_columns(df)
        if df.empty:
            return
        display_df = df.reindex(columns=self.columns, fill_value="")
        rows = list(display_df.itertuples(index=False, name=None))

        def insert_chunk(start=0):
            if generation != self._tree_render_generation:
                return
            end = min(start + 150, len(rows))
            for index in range(start, end):
                values = rows[index]
                status = values[self.columns.index("new_comment")]
                tag = "Ritinta" if str(status).startswith("Ritinta") else status
                if not tag:
                    tag = "stripe" if index % 2 else ""
                self.tree.insert("", "end", values=values, tags=((tag,) if tag else ()))
            if end < len(rows):
                self._tree_render_after_id = self.after(1, insert_chunk, end)
            else:
                self._tree_render_after_id = None

        insert_chunk()

    def _on_search_changed(self, *_args):
        """Debounce typing so the whole table is not redrawn per keystroke."""
        if self._filter_after_id is not None:
            self.after_cancel(self._filter_after_id)
        self._filter_after_id = self.after(180, self._apply_filter)

    def _apply_filter(self):
        self._filter_after_id = None
        q = self.search_var.get().strip().lower()
        if not q:
            self._render_tree(self.current_df)
            return
        if self.current_df.empty:
            return
        searchable = self.current_df.reindex(columns=self.columns, fill_value="").astype(str)
        mask = searchable.apply(lambda col: col.str.contains(q, case=False, regex=False)).any(axis=1)
        self._render_tree(self.current_df[mask], autosize=False)

    def _sort_by(self, col):
        ascending = self.sort_state.get(col, True)
        if self.current_df.empty:
            return
        self.current_df = self.current_df.sort_values(by=col, ascending=ascending, key=lambda s: s.astype(str))
        self.sort_state[col] = not ascending
        self._apply_filter()

    def _open_abbina(self):
        if self._focus_child_window("abbina"):
            return
        suggestions = build_suggestions(self.current_df, max_extra_percent=0.20)
        window = tk.Toplevel(self)
        self._child_windows["abbina"] = window
        window.title("Da abbinare")
        window.geometry("1250x600")
        window.minsize(850, 350)
        window.resizable(True, True)

        top = ttk.Frame(window)
        top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text=(f"{len(suggestions)} righe — stesso Codice/Colore, Titolo compatibile; limite extra 20% (oltre = ⚠)"),
                  foreground="#344054").pack(side="left")
        ttk.Label(top, text="Auto search:").pack(side="left", padx=(18, 4))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(top, textvariable=search_var, width=24)
        search_entry.pack(side="left")
        ttk.Button(top, text="Export Abbina", command=lambda: self._export_abbina(suggestions)).pack(side="right")

        cols = ["titolo", "codice", "colore", "rocche", "partita", "bagno", "abbina",
                "tot_rocche", "mc_target", "polmoni", "extra_percent", "motivo", "new_comment"]
        labels = {"titolo": "Titolo", "codice": "Codice", "colore": "Colore", "rocche": "Rocche",
                  "partita": "Partita", "bagno": "Bagno", "abbina": "Abbina", "tot_rocche": "Tot. Rocche",
                  "mc_target": "Capacità", "polmoni": "Polmoni", "extra_percent": "Extra %",
                  "motivo": "Motivo", "new_comment": "New Comment"}
        frame = ttk.Frame(window)
        frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        style = ttk.Style(window)
        style.configure("Abbina.Treeview", rowheight=28, font=("Segoe UI", 9))
        tree = ttk.Treeview(frame, columns=cols, show="headings", style="Abbina.Treeview")
        for col in cols:
            tree.heading(col, text=labels[col])
            tree.column(col, width=105 if col not in ("motivo", "abbina", "new_comment") else 230, anchor="center")
        yscroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        tree.tag_configure("group_a", background="#eaf2f8")
        tree.tag_configure("group_b", background="#fff7e6")

        def render():
            tree.delete(*tree.get_children())
            query = search_var.get().strip().casefold()
            visible = suggestions
            if query:
                mask = suggestions[cols].fillna("").astype(str).apply(
                    lambda column: column.str.casefold().str.contains(query, regex=False)
                ).any(axis=1)
                visible = suggestions[mask]
            group_tags = {
                group: "group_a" if index % 2 == 0 else "group_b"
                for index, group in enumerate(
                    suggestions.apply(
                        lambda row: f"{row.get('codice', '')}|{row.get('colore', '')}|{row.get('motivo', '')}",
                        axis=1,
                    ).drop_duplicates()
                )
            }
            for _, row in visible.iterrows():
                group = f"{row.get('codice', '')}|{row.get('colore', '')}|{row.get('motivo', '')}"
                values = [f"{row[c]:.1%}" if c == "extra_percent" else row.get(c, "") for c in cols]
                tree.insert("", "end", values=values, tags=(group_tags[group],))

        search_var.trace_add("write", lambda *_: render())
        render()

    def _export_abbina(self, suggestions):
        if suggestions.empty:
            messagebox.showinfo("Da abbinare", "Non ci sono combinazioni entro il limite del 20%.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                                             initialfile="Da_abbinare.xlsx")
        if not path:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        headers = ["Titolo", "Codice", "Colore", "Rocche", "Partita", "Bagno", "Abbina",
                   "Tot. Rocche", "Capacità M/C", "Polmoni", "Extra %", "Motivo", "New Comment"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Da abbinare"
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="16324F")
        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        thin = Side(style="thin", color="B8C6D6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        group_colors = {"group_a": "EAF2F8", "group_b": "FFF7E6"}
        group_tags = {
            group: "group_a" if index % 2 == 0 else "group_b"
            for index, group in enumerate(
                suggestions.apply(
                    lambda row: f"{row.get('codice', '')}|{row.get('colore', '')}|{row.get('motivo', '')}",
                    axis=1,
                ).drop_duplicates()
            )
        }
        for _, row in suggestions.iterrows():
            group = f"{row.get('codice', '')}|{row.get('colore', '')}|{row.get('motivo', '')}"
            values = [row.get("titolo", ""), row.get("codice", ""), row.get("colore", ""),
                      row.get("rocche", ""), row.get("partita", ""), row.get("bagno", ""),
                      row.get("abbina", ""), row.get("tot_rocche", ""), row.get("mc_target", ""),
                      row.get("polmoni", ""), f"{row.get('extra_percent', 0):.1%}",
                      row.get("motivo", ""), row.get("new_comment", "")]
            ws.append(values)
            fill = PatternFill("solid", fgColor=group_colors[group_tags[group]])
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.font = Font(name="Arial")
                cell.alignment = center
                cell.border = border

        last_row = ws.max_row
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.freeze_panes = "A2"
        for column_index, header in enumerate(headers, start=1):
            letter = get_column_letter(column_index)
            values = [str(ws.cell(row=row, column=column_index).value or "") for row in range(1, last_row + 1)]
            width = min(max(max(len(value) for value in values) + 3, len(header) + 2, 10), 45)
            ws.column_dimensions[letter].width = width
        ws.row_dimensions[1].height = 28
        wb.save(path)
        messagebox.showinfo("Completed", f"Export completed successfully:\n{path}")

    def _open_yarn_shortage(self):
        """Opens Mancanza Filato as a popup window (same pattern as Da abbinare)."""
        if self._focus_child_window("yarn_shortage"):
            return
        window = tk.Toplevel(self)
        self._child_windows["yarn_shortage"] = window
        window.title("Mancanza Filato")
        window.geometry("1100x650")
        window.minsize(750, 350)
        window.resizable(True, True)

        shortage_view = YarnShortageTab(window, self)
        shortage_view.pack(fill="both", expand=True)
        self.add_table_loaded_callback(shortage_view.refresh)

        def on_close():
            try:
                self._table_loaded_callbacks.remove(shortage_view.refresh)
            except ValueError:
                pass
            self._child_windows.pop("yarn_shortage", None)
            window.destroy()

        window.protocol("WM_DELETE_WINDOW", on_close)
        shortage_view.refresh()

    # -------------------------------------------------------------- export
    @staticmethod
    def _to_number(text):
        """Parse a stored value into a real number. Handles European comma
        decimals (e.g. '128,00') as well as plain '128'. Returns None for
        blank/unparseable values so the cell stays genuinely empty."""
        if text is None:
            return None
        s = str(text).strip()
        if not s or s.lower() == "nan":
            return None
        s = s.replace(".", "").replace(",", ".") if ("," in s and s.count(",") == 1) else s
        try:
            value = float(s)
        except ValueError:
            return None
        return int(value) if value.is_integer() else value

    @staticmethod
    def _to_date(text):
        """Parse a stored 'YYYY-MM-DD' string into a real date. Returns None
        for blank/unparseable values."""
        if not text:
            return None
        s = str(text).strip()
        if not s or s.lower() == "nan":
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None

    def _on_export(self):
        if self.current_df.empty:
            messagebox.showinfo("No data", "Refresh the data before exporting.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")],
                                             initialfile="Situazione_Generale.xlsx")
        if not path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule, FormulaRule

        # Export exactly the rows currently visible after Search.  This keeps
        # Excel consistent with the user's filtered Treeview.
        export_df = self.current_df.copy()
        q = self.search_var.get().strip().lower()
        if q and not export_df.empty:
            searchable = export_df.reindex(columns=self.columns, fill_value="").astype(str)
            mask = searchable.apply(lambda col: col.str.contains(q, case=False, regex=False)).any(axis=1)
            export_df = export_df[mask]

        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="16324F", end_color="16324F", fill_type="solid")

        wb = Workbook()
        ws = wb.active
        ws.title = "Situazione"
        headers = [header for _, header, _ in COLUMN_SPEC]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        date_format = "dd/mm/yyyy"

        for _, r in export_df.iterrows():
            row_values = []
            for key, _header, ctype in COLUMN_SPEC:
                raw = r.get(key, "")
                if ctype == "number":
                    row_values.append(self._to_number(raw))
                elif ctype == "date":
                    row_values.append(self._to_date(raw))
                else:
                    row_values.append(raw if raw not in (None, "") else None)
            ws.append(row_values)

            status = str(r.get("new_comment", ""))
            color = color_for_status(status if not status.startswith("Ritinta") else "Ritinta")
            fill = PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
            for col_index, (key, _header, ctype) in enumerate(COLUMN_SPEC, start=1):
                cell = ws.cell(row=ws.max_row, column=col_index)
                cell.fill = fill
                cell.font = Font(name="Arial")
                cell.alignment = center
                cell.border = border
                if ctype == "date" and cell.value is not None:
                    cell.number_format = date_format

        last_row = ws.max_row
        last_col_letter = get_column_letter(len(COLUMN_SPEC))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        ws.freeze_panes = "A2"
        col_letter = {key: get_column_letter(i) for i, (key, _h, _t) in enumerate(COLUMN_SPEC, start=1)}

        if last_row > 1:
            # Custom == "Check" -> red
            rng = f"{col_letter['custom']}2:{col_letter['custom']}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Check"'], fill=red_fill))

            # Bagno duplicated anywhere in the column -> red
            rng = f"{col_letter['bagno']}2:{col_letter['bagno']}{last_row}"
            first = f"{col_letter['bagno']}2"
            formula = f'AND({first}<>"",COUNTIF(${col_letter["bagno"]}$2:${col_letter["bagno"]}${last_row},{first})>1)'
            ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=red_fill))

            # Days in Q.C > 5 -> red
            rng = f"{col_letter['days_in_qc']}2:{col_letter['days_in_qc']}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["5"], fill=red_fill))

            # Consegna before today (overdue) -> red
            rng = f"{col_letter['consegna']}2:{col_letter['consegna']}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["TODAY()"], fill=red_fill))

        for key, header, _ctype in COLUMN_SPEC:
            letter = col_letter[key]
            values = [str(v) for v in export_df[key].tolist()] if key in export_df.columns else []
            longest = max([len(header)] + [len(v) for v in values]) if values else len(header)
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 35)

        wb.save(path)
        logger.info("Situazione: exported %d visible rows to %s", len(export_df), path)
        messagebox.showinfo("Completed", f"Export completed successfully:\n{path}")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Situazione (standalone test)")
    root.geometry("1400x760")
    tab = SituazioneTab(root)
    tab.pack(fill="both", expand=True)
    root.mainloop()
