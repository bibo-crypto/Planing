"""
prezzi_tab.py — "Prezzi" tab: browse Delta's Listini (price list) export.

Upload the Listini file once (its path is cached and auto-restored on the
next launch, same as Magazino Filato). Prices are looked up by Articolo
(CLARTICOLO) + colour code (CLCOLORE) together -- the same Articolo can
have several colours, each priced differently, since a different colour
often means a different raw yarn.
"""
import threading
import tkinter as tk
from pathlib import Path
from tkinter import ttk, filedialog, messagebox

import pandas as pd

import prezzi_logic as logic
from prezzi_cache import load_prezzi_cache, save_prezzi_cache
from utils import logger

COLUMNS = logic.DISPLAY_COLUMNS
HEADERS = logic.HEADERS


class PrezziTab(ttk.Frame):
    """Embeddable 'Prezzi' tab."""

    def __init__(self, master):
        super().__init__(master)
        self._base_df = pd.DataFrame()
        self.prezzi_df = pd.DataFrame()
        self.summary_df = pd.DataFrame()
        self._uploading = False
        self._filter_after_id = None
        self._render_after_id = None
        self._render_generation = 0
        self._loaded_source_path = ""
        self._sort_column = ""
        self._sort_reverse = False
        self._articolo_var = tk.StringVar()
        self._codice_var = tk.StringVar()

        self._configure_styles()
        self._build_upload_panel()
        self._build_treeview()
        self.after_idle(self._restore_from_cache)

    # ------------------------------------------------------------------
    def _configure_styles(self) -> None:
        # Matches the navy heading style already used elsewhere in the app
        # (ui/gui.py sets the "clam" theme globally at startup, needed for
        # a Treeview heading's background colour to actually show).
        style = ttk.Style(self)
        style.configure("Prezzi.Treeview", background="#ffffff", fieldbackground="#ffffff",
                         foreground="#1d2939", rowheight=27, font=("Segoe UI", 9))
        style.configure("Prezzi.Treeview.Heading", background="#16324f", foreground="#ffffff",
                         relief="raised", borderwidth=1, padding=(8, 6),
                         font=("Segoe UI", 9, "bold"))
        style.map("Prezzi.Treeview.Heading",
                  foreground=[("pressed", "#ffffff"), ("active", "#ffffff"), ("!active", "#ffffff")],
                  background=[("pressed", "#0b2239"), ("active", "#244b70"), ("!active", "#16324f")])
        style.map("Prezzi.Treeview", background=[("selected", "#2563eb")],
                  foreground=[("selected", "#ffffff")])

    def _build_upload_panel(self) -> None:
        panel = ttk.LabelFrame(self, text="1) Upload Listini")
        panel.pack(side="top", fill="x", padx=8, pady=6)

        row = ttk.Frame(panel)
        row.pack(fill="x", padx=4, pady=4)
        self.status_var = tk.StringVar(value="No Listini file uploaded")
        self._btn_upload = ttk.Button(row, text="📤 Upload Listini", command=self._on_upload)
        self._btn_upload.pack(side="left")
        ttk.Label(row, textvariable=self.status_var, foreground="#666666").pack(side="left", padx=8)

        row_search = ttk.Frame(panel)
        row_search.pack(fill="x", padx=4, pady=4)
        ttk.Label(row_search, text="Articolo:").pack(side="left")
        ttk.Entry(row_search, textvariable=self._articolo_var, width=20).pack(side="left", padx=(4, 16))
        ttk.Label(row_search, text="Codice:").pack(side="left")
        ttk.Entry(row_search, textvariable=self._codice_var, width=14).pack(side="left", padx=(4, 16))
        self._articolo_var.trace_add("write", lambda *_: self._on_search_changed())
        self._codice_var.trace_add("write", lambda *_: self._on_search_changed())

        ttk.Button(row_search, text="Clear", command=self._clear_filters).pack(side="left", padx=4)

        self._btn_export = ttk.Button(row_search, text="📤 Extract Excel", command=self._on_export)
        self._btn_export.pack(side="right")

    def _build_treeview(self) -> None:
        frame = ttk.Frame(self)
        frame.pack(side="top", fill="both", expand=True, padx=8, pady=6)

        self.tree = ttk.Treeview(frame, columns=COLUMNS, show="headings", selectmode="browse",
                                  style="Prezzi.Treeview")
        for c in COLUMNS:
            self.tree.heading(c, text=HEADERS[c], command=lambda c=c: self._sort_by(c))
            self.tree.column(c, width=150 if c in ("DESCRIZARTICOLOLI", "CLDESCR") else 120,
                              anchor="w" if c in ("CLARTICOLO", "DESCRIZARTICOLOLI", "CLDESCR") else "center")
        self.tree.tag_configure("oddrow", background="#ffffff")
        self.tree.tag_configure("evenrow", background="#eaf1fb")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

    # ------------------------------------------------------------------
    # Upload + startup restore
    # ------------------------------------------------------------------
    def _on_upload(self) -> None:
        if self._uploading:
            return
        path = filedialog.askopenfilename(
            title="Select the Listini file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        self._load_path(path, save_cache=True)

    def _restore_from_cache(self) -> None:
        cache = load_prezzi_cache()
        source_path = cache.get("source_path")
        if source_path and Path(source_path).is_file():
            self._load_path(source_path, save_cache=False)

    def _load_path(self, path: str, save_cache: bool) -> None:
        normalized_path = str(Path(path).resolve())
        if self._base_df is not None and not self._base_df.empty and normalized_path == self._loaded_source_path:
            self.status_var.set(f"{Path(path).name} — {len(self._base_df)} price rows")
            return
        self._uploading = True
        self._btn_upload.config(state="disabled")
        self.status_var.set(f"Loading {Path(path).name}…")

        def worker():
            try:
                df, errors = logic.load_prezzi(path)
            except Exception as exc:  # noqa: BLE001
                errors = [f"An error occurred while reading the file: {exc}"]
                df = None

            def apply_result():
                self._uploading = False
                self._btn_upload.config(state="normal")
                if df is None:
                    self.status_var.set("Failed to load Listini file")
                    if save_cache:
                        messagebox.showerror("Error", "\n".join(errors))
                    else:
                        logger.warning("Prezzi: could not restore cached file: %s", "; ".join(errors))
                    return
                self._base_df = df
                # Public, unfiltered-by-search accessor for other tabs
                # (Ordine Elvy's Livello/Prezzo lookup) -- self.summary_df
                # changes with the search boxes, this doesn't.
                self.prezzi_df = df
                self._loaded_source_path = normalized_path
                self.status_var.set(f"{Path(path).name} — {len(df)} price rows")
                if save_cache:
                    save_prezzi_cache(path)
                self._apply_search_and_sort()

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    # ------------------------------------------------------------------
    # Search + sort + render
    # ------------------------------------------------------------------
    def _on_search_changed(self) -> None:
        if self._filter_after_id is not None:
            try:
                self.after_cancel(self._filter_after_id)
            except tk.TclError:
                pass
        self._filter_after_id = self.after(180, self._apply_search_and_sort)

    def _clear_filters(self) -> None:
        self._articolo_var.set("")
        self._codice_var.set("")
        self._sort_column = ""
        self._sort_reverse = False
        self._apply_search_and_sort()

    def _apply_search_and_sort(self) -> None:
        self._filter_after_id = None
        df = self._base_df
        if df.empty:
            self.summary_df = pd.DataFrame(columns=COLUMNS)
            self._render()
            return

        articolo_q = self._articolo_var.get().strip().lower()
        if articolo_q:
            df = df[df["CLARTICOLO"].str.lower().str.contains(articolo_q, regex=False)]
        codice_q = self._codice_var.get().strip().lower()
        if codice_q:
            df = df[df["CLCOLORE"].str.lower().str.contains(codice_q, regex=False)]

        if self._sort_column:
            ascending = not self._sort_reverse
            df = df.sort_values(self._sort_column, ascending=ascending, kind="mergesort")

        self.summary_df = df.reset_index(drop=True)
        self._render()

    def _sort_by(self, column: str) -> None:
        if self._sort_column == column:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_column = column
            self._sort_reverse = False
        self._apply_search_and_sort()

    def _render(self) -> None:
        self._render_generation += 1
        generation = self._render_generation
        if self._render_after_id is not None:
            try:
                self.after_cancel(self._render_after_id)
            except tk.TclError:
                pass
            self._render_after_id = None
        self.tree.delete(*self.tree.get_children())
        if self.summary_df.empty:
            return
        rows = list(self.summary_df[COLUMNS].itertuples(index=False, name=None))

        def insert_chunk(start=0):
            if generation != self._render_generation:
                return
            end = min(start + 150, len(rows))
            for idx in range(start, end):
                row = rows[idx]
                livello = "" if pd.isna(row[4]) else f"{row[4]:g}"
                prezzo = "" if pd.isna(row[5]) else f"{row[5]:.2f}"
                self.tree.insert("", "end", values=(
                    row[0], row[1], row[2], row[3], livello, prezzo,
                ), tags=("evenrow" if idx % 2 == 0 else "oddrow",))
            if end < len(rows):
                self._render_after_id = self.after(1, insert_chunk, end)
            else:
                self._render_after_id = None

        insert_chunk()

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def _on_export(self) -> None:
        if self.summary_df.empty:
            messagebox.showinfo("No data", "Upload the Listini file first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Prezzi.xlsx"
        )
        if not path:
            return

        # Snapshot the visible result and move all workbook work off the Tk
        # thread.  openpyxl can be slow on large Listini exports.
        export_df = self.summary_df[COLUMNS].copy()
        self._btn_export.config(state="disabled")

        def worker():
            error = None
            try:
                self._write_export_file(path, export_df)
            except Exception as exc:  # noqa: BLE001
                error = exc

            def finish():
                self._btn_export.config(state="normal")
                if error is not None:
                    messagebox.showerror("Export failed", str(error), parent=self)
                    return
                logger.info("Prezzi: exported to %s", path)
                messagebox.showinfo("Completed", f"Export completed successfully:\n{path}", parent=self)

            self.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    @staticmethod
    def _write_export_file(path: str, export_df: pd.DataFrame) -> None:
        """Write a Prezzi workbook without blocking the Tk event loop."""

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="16324F", end_color="16324F", fill_type="solid")

        wb = Workbook()
        ws = wb.active
        ws.title = "Prezzi"
        ws.append([HEADERS[c] for c in COLUMNS])
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for _, r in export_df.iterrows():
            livello = None if pd.isna(r["LIVELLOLPZ"]) else r["LIVELLOLPZ"]
            prezzo = None if pd.isna(r["PREZZOLPZ"]) else r["PREZZOLPZ"]
            ws.append([r["CLARTICOLO"], r["DESCRIZARTICOLOLI"], r["CLCOLORE"], r["CLDESCR"], livello, prezzo])
            for cell in ws[ws.max_row]:
                cell.font = Font(name="Arial")
                cell.alignment = center
                cell.border = border

        last_row = ws.max_row
        last_col_letter = get_column_letter(len(COLUMNS))
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        ws.freeze_panes = "A2"
        widths = {"CLARTICOLO": 16, "DESCRIZARTICOLOLI": 24, "CLCOLORE": 14,
                  "CLDESCR": 20, "LIVELLOLPZ": 12, "PREZZOLPZ": 12}
        for i, c in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 16)

        wb.save(path)
