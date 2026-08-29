"""
overview_tab.py — Overview dashboard.

Read-only cross-tab summary, built entirely from data the app already has
in memory — nothing new to upload. Current sources:

  • Situazione Generale (situazione_tab.current_df) — columns include
    cliente, colore, titolo, partita, rocche, comment, new_comment.
      - "comment" starting with "PG-X"        -> yarn shortage (Filato)
      - "new_comment" containing "pronto da
         spedire"                              -> ready to ship, not yet
                                                   marked as shipped
                                                   (no Data Uscita yet, i.e.
                                                   no exit document made)
  • Magazino Filato (magazino_tab.magazino_summary) — columns articolo,
    partita, mag_rocche, mag_peso. articolo prefix identifies the client
    family (G130 = Elvy, G170 = Kamal) since Magazino has no Cliente column.

There's no historical/time-series table in situazione_db.py (partita_state
only tracks current state, not day-by-day snapshots), so the charts here
are current-snapshot breakdowns (by client / status) rather than true
trend-over-time lines. If a history table gets added later this can grow
a real trend chart on top of it.

Auto-refreshes when the tab is shown (call `on_shown()` from the notebook's
<<NotebookTabChanged>> handler) and has a manual Refresh button too.

Every Treeview on this page can export exactly what it's showing (after
search/filter) to an .xlsx file — right-click a row, or use the "Export to
Excel" button above each table. `export_treeview_to_excel()` is a small,
tree-agnostic helper: it only reads `tree["columns"]`, `tree.heading()`
and `tree.get_children()`, so it can be reused as-is on any other
ttk.Treeview in the app, not just the ones built here.
"""
from __future__ import annotations

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils import load_settings, parse_number
import situazione_logic as business_logic

ARTICOLO_CLIENT_LABELS = {"G130": "Elvy", "G170": "Kamal"}
READY_MARK = "pronto da spedire"
SHORTAGE_MARK = "PG-X"
DELIVERY_ALERT_DAYS = 3
QUALITY_READY_CODES = {"AA", "AC", "AU", "AT"}
HEADERS_IT = {
    "cliente": "Cliente", "colore": "Colore", "titolo": "Titolo",
    "articolo": "Articolo", "codice": "Codice", "ordine": "Ordine",
    "riga": "Riga", "data": "Data", "consegna": "Consegna",
    "partita": "Partita", "rocche": "Rocche", "mc": "M/C",
    "comment": "Commento", "raw_yarn_match": "Filato Disponibile",
    "cq": "C.Q", "tinto": "Tinto", "bagno": "Bagno",
    "old_comment": "Vecchio commento", "new_comment": "Nuovo commento",
    "planedate": "PlaneDate", "data_qualita": "Data Qualità",
    "data_uscita": "Data Uscita", "custom": "Controllo",
    "days_in_qc": "Giorni in C.Q",
    "days_to_delivery": "Giorni alla consegna",
    "prezzo": "Prezzo", "issue": "Problema",
}
CHECK_COLUMNS = [
    "cliente", "articolo", "titolo", "codice", "colore", "ordine", "riga",
    "data", "consegna", "partita", "rocche", "mc", "comment", "raw_yarn_match",
    "cq", "tinto", "bagno", "old_comment", "new_comment", "planedate",
    "data_qualita", "data_uscita", "custom", "days_in_qc",
]


# ----------------------------------------------------------------------
# Generic "export any Treeview to Excel" helper — reusable elsewhere.
# ----------------------------------------------------------------------
def treeview_to_dataframe(tree: ttk.Treeview) -> pd.DataFrame:
    """Read a Treeview's current columns/rows (as displayed) into a DataFrame."""
    columns = tree["columns"]
    headers = [str(tree.heading(c)["text"] or c) for c in columns]
    rows = [tree.item(item)["values"] for item in tree.get_children()]
    return pd.DataFrame(rows, columns=headers)


def export_treeview_to_excel(tree: ttk.Treeview, default_filename: str, parent: tk.Widget | None = None) -> None:
    """Export a Treeview's currently displayed rows/columns to a new .xlsx file."""
    df = treeview_to_dataframe(tree)
    if df.empty:
        messagebox.showinfo("Nessun dato", "Non ci sono dati da esportare.", parent=parent)
        return
    path = filedialog.asksaveasfilename(
        title="Esporta in Excel",
        defaultextension=".xlsx",
        filetypes=[("File Excel", "*.xlsx")],
        initialfile=default_filename,
    )
    if not path:
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overview"
        _write_typed_excel_table(ws, df)
        ws.freeze_panes = "A2"
        wb.save(path)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        messagebox.showerror("Esportazione non riuscita", str(exc), parent=parent)
        return
    messagebox.showinfo("Completato", f"Esportato in:\n{path}", parent=parent)


def _write_typed_excel_table(ws, df: pd.DataFrame) -> None:
    """Write an Overview export with typed cells, styling and Excel filters."""
    ws.append(list(df.columns))

    date_headers = {"data", "consegna", "data qualità", "data uscita", "planedate"}
    number_headers = {
        "ordine", "riga", "rocche", "m/c",
        "giorni in c.q", "giorni alla consegna", "extra %",
    }
    general_headers = {"cliente", "codice", "partita", "titolo", "colore"}
    header_fills = {
        "date": "70AD47",    # green
        "number": "5B9BD5",  # blue
        "general": "A5A5A5",  # gray
        "text": "ED7D31",    # orange
    }

    column_types = []
    for col in df.columns:
        header = str(col).strip().casefold()
        if header in date_headers or "data" in header or "date" in header:
            column_types.append("date")
        elif header in general_headers:
            column_types.append("general")
        elif header in number_headers or "giorni" in header:
            column_types.append("number")
        else:
            column_types.append("text")

    for row in df.itertuples(index=False, name=None):
        values = []
        for value, kind in zip(row, column_types):
            if value is None or (isinstance(value, float) and pd.isna(value)):
                values.append(None)
            elif kind == "date":
                parsed = pd.to_datetime(value, errors="coerce")
                values.append(None if pd.isna(parsed) else parsed.to_pydatetime())
            elif kind == "number":
                parsed = parse_number(value)
                values.append(parsed if parsed is not None else str(value))
            else:
                values.append(str(value))
        ws.append(values)

    for index, (col, kind) in enumerate(zip(df.columns, column_types), start=1):
        letter = get_column_letter(index)
        header_cell = ws.cell(row=1, column=index)
        header_cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        header_cell.fill = openpyxl.styles.PatternFill(
            "solid", fgColor=header_fills[kind]
        )
        header_cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        if kind == "date":
            for cell in ws[letter][1:]:
                cell.number_format = "yyyy-mm-dd"
        elif kind == "number":
            for cell in ws[letter][1:]:
                # Overview quantities and identifiers are whole numbers:
                # show 6, not 6.00, while keeping the cells numeric.
                cell.number_format = "#,##0"
        max_len = max(
            [len(str(col))]
            + [len(str(ws.cell(row=row, column=index).value or "")) for row in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[letter].width = max(10, min(45, max_len + 2))

    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    table = Table(displayName="OverviewExport", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def attach_excel_export(tree: ttk.Treeview, default_filename: str) -> None:
    """Right-click on any row of *tree* -> "Export to Excel"."""
    menu = tk.Menu(tree, tearoff=0)
    menu.add_command(
        label="📤 Esporta in Excel",
        command=lambda: export_treeview_to_excel(tree, default_filename, parent=tree),
    )

    def _popup(event):
        row_id = tree.identify_row(event.y)
        if row_id:
            tree.selection_set(row_id)
        menu.tk_popup(event.x_root, event.y_root)

    tree.bind("<Button-3>", _popup)


class OverviewTab(ttk.Frame):
    """Dashboard: KPI cards + charts + detail lists, all read-only."""

    AUTO_REFRESH_MS = 5000

    def __init__(self, master, situazione_tab, magazino_tab, biglietti_tab=None, prezzi_tab=None, save_prefs=None, prefs=None):
        super().__init__(master)
        self.situazione_tab = situazione_tab
        self.magazino_tab = magazino_tab
        self.biglietti_tab = biglietti_tab
        self.prezzi_tab = prezzi_tab
        self._save_prefs = save_prefs
        self._prefs = prefs or {}
        p_dir = self._prefs.get("master_data_dir")
        self._data_folder = Path(p_dir) if p_dir and Path(p_dir).is_dir() else None

        self._data_signature = None
        self._auto_refresh_id = None
        self._build_ui()
        self.refresh(force=True)
        self._schedule_auto_refresh()

    # ------------------------------------------------------------------
    # UI scaffolding
    # ------------------------------------------------------------------
    def _build_ui(self) -> None:
        style = ttk.Style(self)
        style.configure("Overview.Card.TFrame", background="#ffffff", relief="solid", borderwidth=1)
        style.configure("Overview.CardTitle.TLabel", background="#ffffff", foreground="#526173",
                        font=("Segoe UI", 9, "bold"))
        style.configure("Overview.CardValue.TLabel", background="#ffffff", foreground="#16324f",
                        font=("Segoe UI", 21, "bold"))
        style.configure("Overview.Table.TLabelframe", background="#ffffff", borderwidth=1, relief="solid")
        style.configure("Overview.Table.TLabelframe.Label", background="#ffffff", foreground="#16324f",
                        font=("Segoe UI", 10, "bold"))
        style.configure("Overview.Treeview", background="#ffffff", fieldbackground="#ffffff",
                        foreground="#25364d", rowheight=28, font=("Segoe UI", 9))
        style.configure("Overview.Treeview.Heading", background="#16324f", foreground="#ffffff",
                        font=("Segoe UI", 9, "bold"), padding=(7, 6))
        style.map("Overview.Treeview", background=[("selected", "#2f80ed")],
                  foreground=[("selected", "#ffffff")])
        style.configure("Overview.AlertCard.TFrame", background="#fff7ed", relief="solid", borderwidth=1)
        style.configure("Overview.AlertCardTitle.TLabel", background="#fff7ed", foreground="#c2410c",
                        font=("Segoe UI", 9, "bold"))
        style.configure("Overview.AlertCardValue.TLabel", background="#fff7ed", foreground="#9a3412",
                        font=("Segoe UI", 19, "bold"))

        toolbar = ttk.Frame(self)
        toolbar.pack(side="top", fill="x", padx=8, pady=(8, 4))
        ttk.Label(toolbar, text="📊  Operations Overview", foreground="#16324f",
                  font=("Segoe UI", 14, "bold")).pack(side="left")
        self._lbl_updated = ttk.Label(toolbar, text="", foreground="#667085", font=("Segoe UI", 9))
        self._lbl_updated.pack(side="right")

        outer = ttk.Frame(self)
        outer.pack(side="top", fill="both", expand=True, padx=4, pady=4)

        canvas = tk.Canvas(outer, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self._body = ttk.Frame(canvas, padding=(4, 2))
        self._body.columnconfigure(0, weight=1)
        body_window = canvas.create_window((0, 0), window=self._body, anchor="nw")

        def _on_body_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(event):
            canvas.itemconfig(body_window, width=event.width)

        self._body.bind("<Configure>", _on_body_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")

        # ── Master Data Synchronization Card ──
        sync_frame = ttk.LabelFrame(self._body, text=" 📂 Master Data Synchronization (Update All Data) ", padding=10)
        sync_frame.pack(side="top", fill="x", padx=8, pady=(4, 10))
        sync_frame.columnconfigure(1, weight=1)

        ttk.Button(sync_frame, text="📂  Upload All Data from Folder...", command=self._on_choose_data_folder, width=28).grid(row=0, column=0, sticky="w", pady=2)

        folder_text = str(self._data_folder) if self._data_folder else "No data folder selected"
        folder_color = "#111827" if self._data_folder else "grey"
        self._lbl_folder = ttk.Label(sync_frame, text=folder_text, foreground=folder_color, anchor="w")
        self._lbl_folder.grid(row=0, column=1, sticky="ew", padx=(10, 10), pady=2)

        self._btn_sync = ttk.Button(sync_frame, text="🔄  Update All Data", command=self._on_sync_all_data, width=20)
        self._btn_sync.grid(row=0, column=2, sticky="e", pady=2)

        sec_row = ttk.Frame(sync_frame)
        sec_row.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        sec_row.columnconfigure(0, weight=1)

        self._lbl_sync_status = ttk.Label(sec_row, text="● Ready to sync", foreground="#2E7D32", font=("Segoe UI", 9))
        self._lbl_sync_status.grid(row=0, column=1, sticky="e", padx=(10, 0))

        self._cards_frame = ttk.Frame(self._body, padding=(0, 2))
        self._cards_frame.pack(side="top", fill="x", padx=10, pady=(3, 12))
        for column in range(3):
            self._cards_frame.columnconfigure(column, weight=1, uniform="overview_card")

        self._tables_frame = ttk.Frame(self._body, padding=(0, 2))
        self._tables_frame.pack(side="top", fill="both", expand=True, padx=10, pady=(0, 14))
        self._tables_frame.columnconfigure(0, weight=1, uniform="overview_table")
        self._tables_frame.columnconfigure(1, weight=1, uniform="overview_table")
        self._table_count = 0

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def on_shown(self) -> None:
        """Call this when the tab becomes visible — keeps it auto-fresh."""
        self._schedule_auto_refresh()
        if self._data_signature != self._current_data_signature():
            self.refresh()

    def _schedule_auto_refresh(self) -> None:
        if self._auto_refresh_id is None and self.winfo_exists():
            self._auto_refresh_id = self.after(self.AUTO_REFRESH_MS, self._auto_refresh_tick)

    def _auto_refresh_tick(self) -> None:
        self._auto_refresh_id = None
        if not self.winfo_exists():
            return
        # Do not touch the Treeviews while the user is on another page.
        if self.winfo_ismapped() and self._data_signature != self._current_data_signature():
            self.refresh()
        self._schedule_auto_refresh()

    def _current_data_signature(self):
        """Return cheap revision counters maintained by the source tabs."""
        return (
            getattr(self.situazione_tab, "_data_revision", 0),
            getattr(self.situazione_tab, "_copertura_revision", 0),
            getattr(self.magazino_tab, "_data_revision", 0),
        )

    def _on_choose_data_folder(self) -> None:
        from tkinter import filedialog, messagebox
        initial = str(self._data_folder) if self._data_folder else None
        path = filedialog.askdirectory(
            title="Select Data Folder containing factory files (Articoli, DFM, WINCOINT, etc.)",
            initialdir=initial,
        )
        if not path:
            return
        self._data_folder = Path(path).expanduser().resolve()
        self._lbl_folder.config(text=str(self._data_folder), foreground="#111827")
        # Persist immediately in the same settings store used at startup.
        # The absolute path also avoids a changed working directory making a
        # valid selected folder look missing on the next launch.
        self._prefs["master_data_dir"] = str(self._data_folder)
        if self._save_prefs:
            self._save_prefs(master_data_dir=str(self._data_folder))
        # Selecting the folder only stores the master directory. The update
        # button is the only action that reads and distributes its files.
        self._lbl_sync_status.config(text="● Folder saved — press Update All Data", foreground="#1565C0")

    def _on_sync_all_data(self) -> None:
        from tkinter import messagebox
        # Reload the persisted value on every update click. This covers the
        # case where Overview was rebuilt or another tab replaced its prefs
        # mapping after the folder was selected.
        saved_folder = load_settings().get("master_data_dir")
        if saved_folder:
            candidate = Path(str(saved_folder)).expanduser()
            if candidate.is_dir():
                self._data_folder = candidate.resolve()
                self._lbl_folder.config(text=str(self._data_folder), foreground="#111827")
        if not self._data_folder or not self._data_folder.is_dir():
            messagebox.showwarning(
                "Data folder not selected",
                "Select the master data folder first with Upload All Data from Folder.",
            )
            return

        self._btn_sync.config(state="disabled")
        self._lbl_sync_status.config(text="⏳ Synchronizing all factory files in background...", foreground="#1565C0")

        import threading
        import master_import

        def worker():
            try:
                loaded, skipped = master_import.import_master_directory(
                    self._data_folder,
                    self.situazione_tab,
                    self.magazino_tab,
                    self.biglietti_tab,
                    self.prezzi_tab,
                )
                err = None
            except Exception as exc:  # noqa: BLE001
                loaded, skipped, err = [], [], str(exc)

            def apply_result():
                self._btn_sync.config(state="normal")
                if err:
                    self._lbl_sync_status.config(text=f"✖ Sync error: {err}", foreground="#C62828")
                    messagebox.showerror("Sync Failed", f"Failed to sync data folder:\n{err}")
                    return

                now_str = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                self._lbl_sync_status.config(text=f"✔ Synced ({len(loaded)} sources) at {now_str}", foreground="#2E7D32")
                self._lbl_updated.config(text=f"Last sync: {now_str}")

                # The individual _handle_upload/_on_upload_magazino calls
                # inside import_master_directory populate each tab's
                # in-memory frames and caches, but Situazione's own table
                # (compute_situation + its Treeview) only recomputes when
                # its own Refresh runs -- do that now too, silently, if
                # every source it needs actually came through this sync.
                try:
                    st = self.situazione_tab
                    required = {"wincoint", "dfm", "data_prod", "copertura", "uscita", "qualita"}
                    if st is not None and required.issubset(getattr(st, "loaded_frames", {}).keys()):
                        st._on_refresh()
                except Exception:
                    pass

                msg_parts = []
                if loaded:
                    msg_parts.append("✅ Successfully Loaded & Distributed:\n" + "\n".join(f"  • {item}" for item in loaded))
                if skipped:
                    msg_parts.append("\n⚠️ Missing / Not Found in Folder:\n" + "\n".join(f"  • {item}" for item in skipped))

                summary = "\n".join(msg_parts) if msg_parts else "No recognized files found in folder."
                messagebox.showinfo("Data Synchronization Complete", summary)
                self.refresh(force=True)

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def refresh(self, force: bool = False) -> None:
        signature = self._current_data_signature()
        if not force and self._data_signature == signature:
            return
        df = getattr(self.situazione_tab, "current_df", None)
        magazino = getattr(self.magazino_tab, "magazino_summary", None)
        df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
        magazino = magazino.copy() if isinstance(magazino, pd.DataFrame) else pd.DataFrame()
        self._render(df, magazino)
        self._data_signature = signature
        self._lbl_updated.config(text=f"Ultimo aggiornamento: {pd.Timestamp.now():%Y-%m-%d %H:%M:%S}")

    # ------------------------------------------------------------------
    # Rendering
    # ------------------------------------------------------------------
    def _render(self, df: pd.DataFrame, magazino: pd.DataFrame) -> None:
        for frame in (self._cards_frame, self._tables_frame):
            for widget in frame.winfo_children():
                widget.destroy()
        self._has_rendered = True
        self._table_count = 0

        if not df.empty:
            for col in ("cliente", "colore", "titolo", "comment", "new_comment", "custom"):
                values = df[col] if col in df.columns else pd.Series("", index=df.index)
                df[col] = values.fillna("").astype(str).str.strip()

        # Keep every shortage row whose Commento contains PG-X.  The raw-yarn
        # match is only an informational column; it must never filter out a
        # shortage that has no available yarn match.
        shortage_df = df.iloc[0:0].copy()
        if not df.empty and "comment" in df.columns:
            shortage_mask = df["comment"].astype(str).str.contains(
                SHORTAGE_MARK, case=False, regex=False, na=False
            )
            shortage_df = df.loc[shortage_mask].copy()
            if "raw_yarn_match" not in shortage_df.columns:
                shortage_df["raw_yarn_match"] = ""
        # "Colori pronti" means: quality is completed with a ready code,
        # more than three days have passed since the quality date, and no
        # warehouse exit/invoice has been recorded yet.  Do not rely only on
        # New Comment, because that status can be carried from old history.
        ready_df = df.iloc[0:0].copy()
        if not df.empty:
            cq_ready = df.get(
                "cq", pd.Series("", index=df.index)
            ).astype(str).str.strip().str.upper().isin(QUALITY_READY_CODES)
            quality_dates = pd.to_datetime(
                df.get("data_qualita", pd.Series(index=df.index)), errors="coerce"
            )
            quality_older_than_3_days = quality_dates.notna() & (
                (pd.Timestamp.now().normalize() - quality_dates.dt.normalize()).dt.days > 3
            )
            no_invoice = pd.to_datetime(
                df.get("data_uscita", pd.Series(index=df.index)), errors="coerce"
            ).isna()
            ready_df = df.loc[cq_ready & quality_older_than_3_days & no_invoice].copy()
        # Quality-delay alert is intentionally based on the three visible
        # report fields, so stale Custom values cannot leak old rows into it.
        check_df = df.iloc[0:0].copy()
        if not df.empty:
            cq_oo = df.get("cq", pd.Series("", index=df.index)).astype(str).str.strip().str.upper().eq("OO")
            new_comment_cq = df.get("new_comment", pd.Series("", index=df.index)).astype(str).str.strip().str.casefold().eq("c.q")
            days_in_qc = pd.to_numeric(df.get("days_in_qc", pd.Series(index=df.index)), errors="coerce")
            check_df = df.loc[cq_oo & new_comment_cq & days_in_qc.gt(4)].copy()

        # Delivery alert: include overdue orders and orders due within the
        # next DELIVERY_ALERT_DAYS days, while excluding already shipped rows.
        # The current situation data stores dates as ISO strings, so parse them
        # explicitly instead of comparing display text.
        delivery_df = df.iloc[0:0].copy()
        if not df.empty and "consegna" in df.columns:
            due_dates = pd.to_datetime(df["consegna"], errors="coerce")
            today = pd.Timestamp.now().normalize()
            deadline = today + pd.Timedelta(days=DELIVERY_ALERT_DAYS)
            # Delivery risk is limited to batches still queued in quality
            # (C.Q = OO), as requested. Include overdue and due-within-window
            # rows, but never rows already shipped.
            not_shipped = ~df.get("new_comment", pd.Series("", index=df.index)).str.casefold().eq("spedita")
            still_in_quality = df.get("cq", pd.Series("", index=df.index)).astype(str).str.strip().eq("OO")
            due_mask = due_dates.notna() & due_dates.le(deadline) & not_shipped & still_in_quality
            delivery_df = df.loc[due_mask].copy()
            delivery_df["days_to_delivery"] = (due_dates.loc[due_mask] - today).dt.days
            delivery_df = delivery_df.sort_values("days_to_delivery")

        n_shortage_colors = shortage_df["colore"].nunique() if not shortage_df.empty else 0
        n_ready_colors = ready_df["colore"].nunique() if not ready_df.empty else 0
        total_yarn_kg = (
            magazino["mag_peso"].sum() if not magazino.empty and "mag_peso" in magazino else 0.0
        )

        self._add_card(0, "📋 Totale partite", f"{len(df):,}")
        self._add_card(1, "🧵 Filato disponibile", f"{total_yarn_kg:,.0f} kg")
        self._add_card(2, "📦 Pronte da spedire", str(n_ready_colors))
        self._add_card(3, "⚠️ In attesa di filato", str(n_shortage_colors))
        self._add_card(4, "⚠️ Ritardo in Q.C", f"{len(check_df):,}", alert=True)
        self._add_card(5, "📅 Ritardo Consegna", f"{len(delivery_df):,}", alert=True)

        price_anomalies = business_logic.find_price_anomalies(df)
        price_problem_colors = len({
            str(item.get("colore", "")).strip()
            for item in price_anomalies
            if str(item.get("colore", "")).strip()
        })
        self._add_card(
            6, "💲 Errori Prezzo — colori", str(price_problem_colors),
            alert=price_problem_colors > 0,
        )

        self._add_machine_summary(df)

        self._add_table(
            "Colori pronti da spedire (senza uscita)",
            ready_df, ["cliente", "codice", "colore", "titolo", "partita", "rocche", "mc", "bagno"],
            "colori_pronti.xlsx",
        )
        self._add_table(
            "Colori in attesa di filato",
            shortage_df, ["cliente", "codice", "colore", "titolo", "ordine", "riga", "partita", "rocche",
                          "comment", "raw_yarn_match"],
            "colori_filato_mancante.xlsx",
        )
        self._add_table(
            "Ritardo in Q.C (C.Q = OO, oltre 4 giorni)",
            check_df,
            ["cliente", "codice", "colore", "titolo", "partita", "rocche",
             "days_in_qc", "new_comment"],
            "ritardo_in_qc.xlsx",
        )
        self._add_table(
            f"📅 Ritardo consegna / entro {DELIVERY_ALERT_DAYS} giorni (C.Q = OO)",
            delivery_df, ["cliente", "codice", "colore", "titolo", "partita", "rocche",
                          "consegna", "days_to_delivery", "new_comment"],
            "ordini_urgenti_consegna.xlsx",
        )
        self._add_table(
            "💲 Errori Prezzo (Prezzo mancante o sospetto)",
            pd.DataFrame(price_anomalies), ["cliente", "articolo", "codice", "ordine", "riga", "bagno", "mc", "prezzo", "issue"],
            "errori_prezzo.xlsx", count_column="colore",
        )

    def _add_machine_summary(self, situation_df: pd.DataFrame) -> None:
        """Render Copertura machine totals as compact Overview cards."""
        copertura = getattr(self.situazione_tab, "loaded_frames", {}).get("copertura")
        counts = {machine: 0 for machine in range(3, 13)}
        if isinstance(copertura, pd.DataFrame) and not copertura.empty and not situation_df.empty:
            if {"bagno", "machine"}.issubset(copertura.columns) and "bagno" in situation_df.columns:
                def key(value):
                    digits = "".join(ch for ch in str(value or "") if ch.isdigit()).lstrip("0")
                    return digits or str(value or "").strip().casefold()

                left = situation_df.copy()
                right = copertura.copy()
                left["_bagno_key"] = left["bagno"].map(key)
                right["_bagno_key"] = right["bagno"].map(key)
                joined = left.merge(right[["_bagno_key", "machine"]].drop_duplicates("_bagno_key"), on="_bagno_key", how="inner")

                # Same parser the Copertura popup uses (business_logic.
                # machine_number_from_label) -- kept in one place so a
                # "5" vs "3305" style label is read identically in both
                # views instead of silently under-counting here.
                joined["_machine_number"] = joined["machine"].map(business_logic.machine_number_from_label)
                valid = joined[joined["_machine_number"].between(3, 12, inclusive="both")]
                counts.update(valid["_machine_number"].value_counts().to_dict())

        frame = tk.LabelFrame(
            self._cards_frame, text="Copertura macchine", bg="#f8fafc", fg="#16324f",
            font=("Segoe UI", 10, "bold"), padx=8, pady=6,
        )
        # Row 2 is reserved for the Errori Prezzi color counter card.
        frame.grid(row=3, column=0, columnspan=3, padx=6, pady=(0, 10), sticky="ew")
        self._cards_frame.columnconfigure(0, weight=1)
        self._cards_frame.columnconfigure(1, weight=1)
        self._cards_frame.columnconfigure(2, weight=1)

        covered_until = business_logic.machine_coverage_until

        for index, machine in enumerate(range(3, 13)):
            count = int(counts[machine])
            empty = count == 0
            card = tk.Frame(
                frame, bg="#fee2e2" if empty else "#eaf2f8",
                highlightbackground="#ef4444" if empty else "#cbd5e1",
                highlightthickness=1, width=106, height=76,
            )
            card.grid(row=0, column=index, padx=3, pady=3, sticky="ew")
            card.grid_propagate(False)
            frame.columnconfigure(index, weight=1, minsize=106)
            tk.Label(card, text=f"M{machine}", bg=card["bg"], fg="#991b1b" if empty else "#16324f",
                     font=("Segoe UI", 9, "bold"), anchor="center").pack(fill="x")
            tk.Label(card, text=f"{count} colori", bg=card["bg"], fg="#991b1b" if empty else "#344054",
                     font=("Segoe UI", 11, "bold"), anchor="center").pack(fill="x")
            tk.Label(card, text=f"Fino al: {covered_until(count)}", bg=card["bg"],
                     fg="#991b1b" if empty else "#667085",
                     font=("Segoe UI", 8, "bold"), anchor="center").pack(fill="x", pady=(2, 0))

    def _add_card(self, col: int, title: str, value: str, alert: bool = False) -> None:
        self._cards_frame.columnconfigure(col, weight=1)
        card_style = "Overview.AlertCard.TFrame" if alert else "Overview.Card.TFrame"
        title_style = "Overview.AlertCardTitle.TLabel" if alert else "Overview.CardTitle.TLabel"
        value_style = "Overview.AlertCardValue.TLabel" if alert else "Overview.CardValue.TLabel"
        card = ttk.Frame(self._cards_frame, style=card_style, padding=(14, 12))
        card.grid(row=col // 3, column=col % 3, padx=5, pady=5, sticky="nsew")
        self._cards_frame.rowconfigure(col // 3, weight=1, minsize=86)
        ttk.Label(card, text=title, style=title_style, wraplength=180).pack(anchor="w")
        ttk.Label(card, text=value, style=value_style).pack(anchor="w", pady=(4, 0))

    def _add_table(self, title: str, df: pd.DataFrame, cols: list[str], export_filename: str, count_column: str | None = None) -> None:
        frame = ttk.LabelFrame(self._tables_frame, text=title, padding=(9, 8), style="Overview.Table.TLabelframe")
        col = self._table_count % 2
        row = self._table_count // 2
        self._tables_frame.columnconfigure(col, weight=1)
        frame.grid(row=row, column=col, padx=6, pady=6, sticky="nsew")
        self._table_count += 1

        header = ttk.Frame(frame)
        header.pack(side="top", fill="x", pady=(0, 4))

        available_cols = [c for c in cols if df.empty or c in df.columns] or cols
        search_var = tk.StringVar()
        ttk.Label(header, text="Cerca:").pack(side="left", padx=(2, 5))
        search_entry = ttk.Entry(header, textvariable=search_var, width=28)
        search_entry.pack(side="left", padx=(0, 8))
        count_label = None
        if count_column:
            count_label = ttk.Label(header, text="", foreground="#667085")
            count_label.pack(side="left", padx=(4, 0))
        table_area = ttk.Frame(frame)
        table_area.pack(fill="both", expand=True)
        tree = ttk.Treeview(table_area, columns=available_cols, show="headings", height=11, style="Overview.Treeview")
        for c in available_cols:
            tree.heading(c, text=HEADERS_IT.get(c, c.capitalize()))
            tree.column(c, width=132, minwidth=92, anchor="center", stretch=True)
        vsb = ttk.Scrollbar(table_area, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_area, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_area.rowconfigure(0, weight=1)
        table_area.columnconfigure(0, weight=1)

        def render_filtered_rows(*_args):
            tree.delete(*tree.get_children())
            visible = df
            query = search_var.get().strip()
            if query and not df.empty:
                mask = df[available_cols].astype(str).apply(
                    lambda column: column.str.contains(query, case=False, regex=False, na=False)
                ).any(axis=1)
                visible = df.loc[mask]
            for index, (_, row) in enumerate(visible.iterrows()):
                tree.insert("", "end", values=[row.get(c, "") for c in available_cols],
                            tags=("evenrow" if index % 2 == 0 else "oddrow",))
            if count_label is not None:
                count = int(visible[count_column].nunique()) if count_column in visible.columns else len(visible)
                count_label.config(
                    text=f"Colori visualizzati: {count}",
                    foreground="#C62828" if count > 0 else "#667085",
                    font=("Segoe UI", 9, "bold") if count > 0 else ("Segoe UI", 9),
                )
        tree.tag_configure("evenrow", background="#f8fafc")
        tree.tag_configure("oddrow", background="#ffffff")
        search_var.trace_add("write", render_filtered_rows)
        render_filtered_rows()

        ttk.Button(
            header, text="📤 Esporta in Excel",
            command=lambda: export_treeview_to_excel(tree, export_filename, parent=self),
        ).pack(side="right")
        attach_excel_export(tree, export_filename)
