"""
elvy_invoice_exporter.py — Builds an Excel workbook from parsed Elvy
raw-yarn invoice data (see elvy_invoice_parser.py).
"""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from elvy_invoice_parser import ElvyInvoiceRow
from utils import logger

COLUMNS: list[tuple[str, str, str]] = [
    ("Inv No",         "inv_no",         "text"),
    ("Date",           "inv_date",       "text"),
    ("Pos",            "pos",            "integer"),
    ("Yarn Code",      "yarn_code",      "text"),
    ("Articolo Delta", "articolo_delta", "text"),
    ("NM",             "nm",             "text"),
    ("Ne",             "ne",             "text"),
    ("Yarn Type",      "yarn_type",      "text"),
    ("LOT",            "lot",            "text"),
    ("Price USD",      "price_usd",      "number"),
    ("Net Qty (kg)",   "net_qty_kg",     "number"),
    ("Gross Qty (kg)", "gross_qty_kg",   "number"),
    ("Value USD",      "value_usd",      "number"),
    ("Rocche",         "rocche",         "integer"),
]

HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
HEADER_FONT = Font(bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER_SIDE = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE,
)

NUMBER_FORMAT = "#,##0.00"
INTEGER_FORMAT = "0"
TEXT_FORMAT = "@"
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 8
YARN_TYPE_COL_WIDTH = 40


class ElvyInvoiceExporter:
    """Writes parsed Elvy invoice rows to a single-sheet Excel workbook."""

    def __init__(self, rows: Sequence[ElvyInvoiceRow], output_path: Path) -> None:
        self.rows = rows
        self.output_path = output_path

    def export(self) -> None:
        logger.info(
            "Exporting %d Elvy invoice row(s) -> %s",
            len(self.rows), self.output_path.name,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Elvy Invoice"
        ws.sheet_view.rightToLeft = False

        for col_idx, (label, _attr, _dtype) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = CELL_BORDER

        for row_idx, obj in enumerate(self.rows, start=2):
            for col_idx, (_, attr, dtype) in enumerate(COLUMNS, start=1):
                raw_value = getattr(obj, attr, "")
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = BODY_FONT
                cell.border = CELL_BORDER

                if dtype == "number":
                    cell.value = float(raw_value) if raw_value is not None else None
                    cell.number_format = NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif dtype == "integer":
                    cell.value = int(raw_value) if raw_value is not None else None
                    cell.number_format = INTEGER_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:  # "text"
                    cell.value = str(raw_value) if raw_value not in (None, "") else ""
                    cell.number_format = TEXT_FORMAT
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col_idx, (label, attr, _) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            if attr == "yarn_type":
                ws.column_dimensions[col_letter].width = YARN_TYPE_COL_WIDTH
                continue
            max_width = len(label)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_width = max(max_width, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(MIN_COL_WIDTH, min(max_width + 4, MAX_COL_WIDTH))

        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            last_col = get_column_letter(len(COLUMNS))
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        logger.info("Export completed: %s", self.output_path)
