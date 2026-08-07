"""
magazino_filato_tab.py
Raw yarn warehouse summary tab ("Magazino Filato"): upload the Magazino
export, apply the filter rules in magazino_logic.py, and show one row per
raw-yarn Partita with its total cones (Mag.rocche) and weight (Mag.peso).
Optionally upload the LOTTI reference too, to add each Partita's Lotto
number (the same reference Kamal calls their raw yarn "message number").
"""
import os
import threading
import tkinter as tk
from pathlib import Path
from tkinter import ttk, filedialog, messagebox
from typing import Callable
import pandas as pd

import lotti_logic
import magazino_logic as logic
from lotti_cache import load_lotti_cache, save_lotti_cache
from magazino_cache import load_magazino_cache, save_magazino_cache
from utils import logger

COLUMNS = ["articolo", "partita", "mag_rocche", "mag_peso", "lotto"]
HEADERS = {
    "articolo": "Articolo", "partita": "Partita", "mag_rocche": "Mag.rocche",
    "mag_peso": "Mag.peso", "lotto": "Lotto",
}


class MagazinoFilatoTab(ttk.Frame):
    """Embeddable 'Magazino Filato' tab."""

    def __init__(self, master, on_shared_cache_changed: Callable[[], None] | None = None):
        super().__init__(master)
        self.magazino_summary = pd.DataFrame()
        self.lotti_summary = pd.DataFrame()
        self._base_df = pd.DataFrame()
        self.summary_df = pd.DataFrame()
        self._shared_path = ""
        self._shared_lotti_path = ""
        self._lotti_syncing = False
        self._syncing = False
        self._uploading = False
        self._on_shared_cache_changed = on_shared_cache_changed
        self._search_var = tk.StringVar()
        self._sort_column = ""
        self._sort_reverse = False

        self._build_upload_panel()
        self._build_treeview()
        self.after_idle(self.sync_shared_lotti_async)
        self.after_idle(self.sync_shared_async)

    def _restore_shared_lotti(self):
        cache = load_lotti_cache()
        source_path = cache.get("source_path", "")
        if not source_path or not Path(source_path).is_file() or not self.lotti_summary.empty:
            return

        try:
            df, errors = lotti_logic.load_lotti(source_path)
        except Exception as exc:  # noqa: BLE001
            df, errors = None, [str(exc)]

        if errors or df is None or df.empty:
            return

        self.lotti_summary = lotti_logic.summarize_by_partita(df)
        self._shared_lotti_path = str(source_path)
        self.lotti_status_var.set(
            f"✅ {len(self.lotti_summary)} Partita/Lotto pairs - {Path(source_path)}"
        )
        self._recompute()

    def sync_shared_lotti(self):
        """Restore a LOTTI file selected elsewhere in the app."""
        cache = load_lotti_cache()
        source_path = cache.get("source_path", "")
        if not source_path or not Path(source_path).is_file():
            return
        if self._shared_lotti_path == str(source_path):
            return

        try:
            df, errors = lotti_logic.load_lotti(source_path)
        except Exception as exc:  # noqa: BLE001
            df, errors = None, [str(exc)]

        if errors or df is None or df.empty:
            return

        self.lotti_summary = lotti_logic.summarize_by_partita(df)
        self._shared_lotti_path = str(source_path)
        self.lotti_status_var.set(
            f"✅ {len(self.lotti_summary)} Partita/Lotto pairs - {Path(source_path)}"
        )
        self._recompute()

    def sync_shared_lotti_async(self):
        """Restore LOTTI in a worker so opening the app stays responsive."""
        if self._lotti_syncing:
            return
        source_path = str(load_lotti_cache().get("source_path", ""))
        if not source_path or not Path(source_path).is_file() or self._shared_lotti_path == source_path:
            return

        self._lotti_syncing = True

        def worker():
            try:
                df, errors = lotti_logic.load_lotti(source_path)
            except Exception as exc:  # noqa: BLE001
                df, errors = None, [str(exc)]

            def apply_result():
                self._lotti_syncing = False
                if errors or df is None or df.empty:
                    return
                self.lotti_summary = lotti_logic.summarize_by_partita(df)
                self._shared_lotti_path = source_path
                self.lotti_status_var.set(
                    f"✅ {len(self.lotti_summary)} Partita/Lotto pairs - {Path(source_path)}"
                )
                self._recompute()

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def _build_upload_panel(self):
        panel = ttk.LabelFrame(self, text="1) Upload files")
        panel.pack(side="top", fill="x", padx=8, pady=6)

        row = ttk.Frame(panel)
        row.pack(fill="x", padx=4, pady=4)
        self.status_var = tk.StringVar(value="No Magazino file uploaded")
        self._btn_magazino = ttk.Button(row, text="Select Magazino file...", command=self._on_upload_magazino)
        self._btn_magazino.pack(side="left")
        ttk.Label(row, textvariable=self.status_var, foreground="#666666").pack(side="left", padx=8)

        row2 = ttk.Frame(panel)
        row2.pack(fill="x", padx=4, pady=4)
        self.lotti_status_var = tk.StringVar(value="No LOTTI file uploaded (optional)")
        ttk.Button(row2, text="Select LOTTI file...", command=self._on_upload_lotti).pack(side="left")
        ttk.Label(row2, textvariable=self.lotti_status_var, foreground="#666666").pack(side="left", padx=8)

        row_search = ttk.Frame(panel)
        row_search.pack(fill="x", padx=4, pady=4)
        ttk.Label(row_search, text="Search:").pack(side="left", padx=(0, 8))
        search_entry = ttk.Entry(row_search, textvariable=self._search_var, width=36)
        search_entry.pack(side="left", padx=(0, 8))
        search_entry.bind("<KeyRelease>", lambda _e: self._on_search_changed())
        self._search_var.trace_add("write", lambda *_: self._on_search_changed())

        row3 = ttk.Frame(panel)
        row3.pack(fill="x", padx=4, pady=4)
        ttk.Button(row3, text="Export Excel", command=self._on_export).pack(side="right")

    def _build_treeview(self):
        frame = ttk.Frame(self)
        frame.pack(side="top", fill="both", expand=True, padx=8, pady=6)

        self.tree = ttk.Treeview(frame, columns=COLUMNS, show="headings", selectmode="browse")
        for c in COLUMNS:
            self.tree.heading(c, text=HEADERS[c], command=lambda c=c: self._sort_by(c))
            self.tree.column(c, width=140, anchor="w" if c == "articolo" else "center")
        self.tree.tag_configure("oddrow", background="#FFFFFF")
        self.tree.tag_configure("evenrow", background="#EAF1FB")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

    def _on_upload_magazino(self):
        if self._uploading:
            return
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return
        self._uploading = True
        self._btn_magazino.config(state="disabled")
        self.status_var.set("Loading Magazino file…")

        def worker():
            try:
                df, errors = logic.load_magazino(path)
            except Exception as exc:  # noqa: BLE001
                errors = [f"An error occurred while reading the file: {exc}"]
                df = None

            def apply_result():
                self._uploading = False
                self._btn_magazino.config(state="normal")
                if errors or df is None or df.empty:
                    msg = "; ".join(errors) if errors else "The file is empty after filtering"
                    self.status_var.set(f"❌ {msg}")
                    messagebox.showerror("Error", msg)
                    return

                # Raw files need aggregation; app-generated summary files are
                # already normalized and can be used as-is.
                if {"mag_rocche", "mag_peso"}.issubset(df.columns):
                    self.magazino_summary = df
                else:
                    self.magazino_summary = logic.summarize_by_partita(df)
                self._base_df = self.magazino_summary.copy()
                self._shared_path = str(path)
                save_magazino_cache(path, self.magazino_summary)
                if self._on_shared_cache_changed:
                    self._on_shared_cache_changed()
                self.status_var.set(f"✅ {len(self.magazino_summary)} batches - {os.path.basename(path)}")
                logger.info("Magazino Filato: loaded %d batches from %s",
                            len(self.magazino_summary), os.path.basename(path))
                self._recompute()

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def _on_upload_lotti(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return
        try:
            df, errors = lotti_logic.load_lotti(path)
        except Exception as exc:  # noqa: BLE001
            errors = [f"An error occurred while reading the file: {exc}"]
            df = None

        if errors or df is None or df.empty:
            msg = "; ".join(errors) if errors else "The file is empty after filtering"
            self.lotti_status_var.set(f"❌ {msg}")
            messagebox.showerror("Error", msg)
            return

        self.lotti_summary = lotti_logic.summarize_by_partita(df)
        self._shared_lotti_path = str(path)
        save_lotti_cache(path)
        if self._on_shared_cache_changed:
            self._on_shared_cache_changed()
        self.lotti_status_var.set(f"✅ {len(self.lotti_summary)} Partita/Lotto pairs - {path}")
        logger.info("Magazino Filato: loaded %d Lotto entries from %s",
                    len(self.lotti_summary), os.path.basename(path))
        self._recompute()

    def sync_shared_async(self):
        """Restore a Magazino export selected elsewhere in the app without blocking the UI."""
        if self._syncing:
            return
        cache = load_magazino_cache()
        source_path = str(cache.get("source_path", ""))
        if not source_path or self._shared_path == source_path or not os.path.isfile(source_path):
            return

        self._syncing = True

        def worker():
            try:
                df, errors = logic.load_magazino(source_path)
            except Exception as exc:  # noqa: BLE001
                df, errors = None, [str(exc)]

            def apply_result():
                self._syncing = False
                if errors or df is None or df.empty:
                    return
                self.magazino_summary = logic.summarize_by_partita(df)
                self._base_df = self.magazino_summary.copy()
                self._shared_path = source_path
                save_magazino_cache(source_path, self.magazino_summary)
                self.status_var.set(f"✅ {len(self.magazino_summary)} batches - {source_path}")
                self._recompute()

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def _get_unfiltered_df(self):
        if self._base_df.empty:
            return pd.DataFrame()

        if not self.lotti_summary.empty:
            df = self._base_df.merge(self.lotti_summary, on="partita", how="left")
        else:
            df = self._base_df.copy()
            df["lotto"] = ""

        return df

    def _recompute(self):
        self._apply_search_and_sort()

    def _render(self):
        self.tree.delete(*self.tree.get_children())
        for idx, r in self.summary_df.iterrows():
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert("", "end", values=(
                r["articolo"], r["partita"], f"{r['mag_rocche']:g}", f"{r['mag_peso']:g}",
                r.get("lotto", "") if pd.notna(r.get("lotto", "")) else "",
            ), tags=(tag,))

    def _apply_search_and_sort(self):
        df = self._get_unfiltered_df()
        if df.empty:
            self.summary_df = pd.DataFrame()
            self._render()
            return

        q = self._search_var.get().strip().lower()
        if q:
            mask = df.apply(lambda row: row.astype(str).str.lower().str.contains(q).any(), axis=1)
            df = df[mask]

        if self._sort_column:
            ascending = not self._sort_reverse
            df = df.sort_values(self._sort_column, ascending=ascending, kind="mergesort")

        self.summary_df = df.reset_index(drop=True)
        self._render()

    def _sort_by(self, column: str) -> None:
        if self._sort_column == column:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_column = column
            self._sort_reverse = False
        self._apply_search_and_sort()

    def _on_search_changed(self):
        self._apply_search_and_sort()

    def _on_export(self):
        if self.summary_df.empty:
            messagebox.showinfo("No data", "Upload the Magazino file first.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")],
                                             initialfile="Magazino_Filato.xlsx")
        if not path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="16324F", end_color="16324F", fill_type="solid")

        wb = Workbook()
        ws = wb.active
        ws.title = "Magazino Filato"
        ws.append([HEADERS[c] for c in COLUMNS])
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for _, r in self.summary_df.iterrows():
            lotto_val = r.get("lotto", "")
            lotto_val = "" if pd.isna(lotto_val) else lotto_val
            ws.append([r["articolo"], r["partita"], r["mag_rocche"], r["mag_peso"], lotto_val])
            for cell in ws[ws.max_row]:
                cell.font = Font(name="Arial")
                cell.alignment = center
                cell.border = border

        last_row = ws.max_row
        last_col_letter = get_column_letter(len(COLUMNS))
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        ws.freeze_panes = "A2"
        for i in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        wb.save(path)
        logger.info("Magazino Filato: exported to %s", path)
        messagebox.showinfo("Completed", f"Export completed successfully:\n{path}")
