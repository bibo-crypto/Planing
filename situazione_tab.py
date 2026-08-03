"""
situazione_tab.py
Dyeing Situation Tracker — embedded as the "Situazione" tab of the Delta
Dyeing PDF/Excel converter. Replaces the old Old Situazione / WOORKSHEET /
New Situazione Excel + Power Query chain: each source is uploaded and
validated separately, and per-Partita Old/New Comment history is kept in
a local SQLite database (see situazione_db.py) instead of copy-pasted
sheets.
"""
import os
import json
import tkinter as tk
from pathlib import Path
from tkinter import ttk, filedialog, messagebox
from tkinter import font as tkfont
from datetime import datetime
import pandas as pd

import situazione_db as db
import situazione_loaders as data_loaders
import situazione_logic as business_logic
from dfm_lookup import build_dfm_lookup, load_dfm_cache, save_dfm_cache
from utils import logger

STATUS_COLORS = {
    "Filato": "#e0e0e0",
    "pronto da spedire": "#c6efce",
    "Spedita": "#bdd7ee",
    "spedita": "#bdd7ee",
    "Check": "#ffc7ce",
    "C.Q": "#fff2cc",
}


def color_for_status(status):
    if not status:
        return "#ffffff"
    if status.startswith("Ritinta"):
        return "#d9c6f0"
    return STATUS_COLORS.get(status, "#ffffff")


SOURCE_ORDER = ["copertura", "data_prod", "dfm", "wincoint", "uscita", "qualita"]

SOURCE_BUTTON_NAMES = {
    "copertura": "Copertura",
    "data_prod": "Produzione",
    "dfm": "DFM",
    "wincoint": "Wincoint",
    "uscita": "Uscita",
    "qualita": "Qualita",
    "codes": "Articoli",
}

# (internal_key, header shown in the app / exported file, cell type)
# Order and headers match the real "New Situazione" sheet exactly.
COLUMN_SPEC = [
    ("cliente", "CLIENTE", "text"),
    ("articolo", "Articolo", "text"),
    ("titolo", "Titolo", "text"),
    ("codice", "Codice", "number"),
    ("colore", "Colore", "text"),
    ("ordine", "Ordine", "text"),
    ("riga", "Riga", "number"),
    ("data", "Data", "date"),
    ("consegna", "Consegna", "date"),
    ("partita", "Partita", "number"),
    ("rocche", "Rocche", "number"),
    ("mc", "M/C", "number"),
    ("comment", "Comment", "text"),
    ("cq", "C.Q", "text"),
    ("tinto", "Tinto", "date"),
    ("bagno", "Bagno", "text"),
    ("old_comment", "Old Comm.", "text"),
    ("new_comment", "New Comm.", "text"),
    ("planedate", "PlaneDate", "text"),
    ("data_qualita", "Data Qualita", "date"),
    ("data_uscita", "Data Uscita", "date"),
    ("custom", "Custom", "text"),
    ("days_in_qc", "Days in Q.C", "number"),
]
COLUMN_LABELS = {key: header for key, header, _ in COLUMN_SPEC}
COLUMN_TYPES = {key: ctype for key, _, ctype in COLUMN_SPEC}


class SourceRow(ttk.Frame):
    """Compact upload card: file name button + current upload status."""

    def __init__(self, master, key, label, on_upload):
        super().__init__(master)
        self.key = key
        self.on_upload = on_upload

        self.status_var = tk.StringVar(value="Not uploaded")
        self.dot = tk.Label(self, text="●", fg="#9e9e9e", bg="#f4f6f8", font=("Segoe UI", 11))
        self.dot.grid(row=1, column=0, padx=(4, 2), sticky="w")

        button_name = SOURCE_BUTTON_NAMES.get(key, label)
        self.button = ttk.Button(
            self,
            text=f"Upload {button_name}",
            command=self._browse,
            width=21,
            font=("Segoe UI", 9, "bold"),
            height=30,
        )
        self.button.grid(row=0, column=0, columnspan=2, padx=4, pady=(3, 2), sticky="ew")

        self.status_lbl = ttk.Label(self, textvariable=self.status_var, anchor="w",
                                    foreground="#667085", width=22)
        self.status_lbl.grid(row=1, column=1, padx=(2, 4), sticky="w")

        self.columnconfigure(1, weight=1)

    def _browse(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return
        self.on_upload(self.key, path)

    def set_status(self, ok, message):
        self.dot.config(fg="#2e7d32" if ok else "#c62828")
        self.status_var.set(message)


class SituazioneTab(ttk.Frame):
    """Embeddable 'Situazione' tab — hosted inside gui.py's main Notebook."""

    def __init__(self, master):
        super().__init__(master)

        self._configure_styles()

        db.init_db()

        self.loaded_frames = {}   # key -> DataFrame (validated, ready to merge)
        self.sort_state = {}      # column -> ascending bool
        self.current_df = pd.DataFrame()
        self._filter_after_id = None

        self._build_upload_panel()
        self._build_toolbar()
        self._build_treeview()
        self._refresh_source_labels_from_db()
        # Load the saved table after all widgets have completed initialization.
        self.after_idle(self._load_table_from_db)

    # ------------------------------------------------------------------ UI
    def _configure_styles(self):
        # Note: intentionally does NOT call style.theme_use(...) — this tab
        # shares a ttk.Style with the rest of the app (Data Elvy, Ordini
        # ELVY, Med Bolla, Elvy Invoice tabs), so switching the global theme
        # here would also change how every other tab looks.  The native
        # Windows theme ignores Treeview heading background colors, so use
        # the built-in clam renderer when available; it honors the explicit
        # header colors below.
        style = ttk.Style(self)
        if "clam" in style.theme_names():
            style.theme_use("clam")
        style.configure("Upload.TLabelframe", background="#f4f6f8", borderwidth=1, relief="solid")
        style.configure("Upload.TLabelframe.Label", background="#f4f6f8", foreground="#344054",
                        font=("Segoe UI", 10, "bold"))
        style.configure("Situazione.Treeview", background="#ffffff", fieldbackground="#ffffff",
                        foreground="#1d2939", rowheight=29, borderwidth=1, relief="solid",
                        bordercolor="#b8c6d6", lightcolor="#b8c6d6", darkcolor="#b8c6d6",
                        font=("Segoe UI", 9))
        style.configure("Situazione.Treeview.Heading", background="#16324f", foreground="#ffffff",
                        relief="raised", borderwidth=1, padding=(8, 8),
                        font=("Segoe UI", 9, "bold"))
        style.map("Situazione.Treeview.Heading",
                  foreground=[("pressed", "#ffffff"), ("active", "#ffffff"), ("!active", "#ffffff")],
                  background=[("pressed", "#0b2239"), ("active", "#244b70"), ("!active", "#16324f")])
        style.map("Situazione.Treeview", background=[("selected", "#2563eb")],
                  foreground=[("selected", "#ffffff")])

    def _build_upload_panel(self):
        upload_area = ttk.Frame(self)
        upload_area.pack(side="top", fill="x", anchor="w", padx=8, pady=(6, 2))

        panel = ttk.LabelFrame(upload_area, text="1) Required files", style="Upload.TLabelframe")
        panel.pack(side="left", anchor="nw")
        panel.configure(padding=(5, 3))

        self.source_rows = {}
        for key in SOURCE_ORDER:
            label, _ = data_loaders.LOADERS[key]
            row = SourceRow(panel, key, label, self._handle_upload)
            row.grid(row=0, column=SOURCE_ORDER.index(key), padx=3, pady=1, sticky="nw")
            self.source_rows[key] = row

        codes_panel = ttk.LabelFrame(upload_area, text="Optional file", style="Upload.TLabelframe")
        codes_panel.pack(side="left", anchor="nw", padx=(8, 0), fill="y")
        codes_panel.configure(padding=(5, 3))

        codes_row = SourceRow(codes_panel, "codes", "Yarn codes (Articoli) - optional",
                               self._handle_codes_upload)
        codes_row.grid(row=0, column=0, padx=3, pady=1, sticky="nw")
        self.codes_row = codes_row

    def _build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(side="top", fill="x", padx=8, pady=(3, 4))

        self.refresh_btn = ttk.Button(bar, text="2) Refresh", command=self._on_refresh)
        self.refresh_btn.pack(side="left", padx=4)

        self.upload_data_btn = ttk.Button(bar, text="Upload Data", command=self._on_upload_data)
        self.upload_data_btn.pack(side="left", padx=4)

        self.export_btn = ttk.Button(bar, text="Export to Excel", command=self._on_export)
        self.export_btn.pack(side="left", padx=4)

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search_changed)
        ttk.Label(bar, text="Search:").pack(side="left", padx=(18, 4))
        ttk.Entry(bar, textvariable=self.search_var, width=30).pack(side="left")

        self.summary_lbl = ttk.Label(bar, text="")
        self.summary_lbl.pack(side="right", padx=8)

    def _build_treeview(self):
        cols = [key for key, _, _ in COLUMN_SPEC]
        self.columns = cols

        frame = ttk.Frame(self, borderwidth=1, relief="solid")
        frame.pack(side="top", fill="both", expand=True, padx=8, pady=(2, 8))

        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse",
                                 style="Situazione.Treeview")
        for c in cols:
            ctype = COLUMN_TYPES.get(c, "text")
            anchor = "center" if ctype in ("number", "date") else "w"
            self.tree.heading(c, text=COLUMN_LABELS.get(c, c), anchor="center",
                              command=lambda c=c: self._sort_by(c))
            self.tree.column(c, width=100, minwidth=60, anchor=anchor, stretch=False)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        for status, color in STATUS_COLORS.items():
            self.tree.tag_configure(status, background=color)
        self.tree.tag_configure("Ritinta", background="#d9c6f0")
        self.tree.tag_configure("stripe", background="#f3f6fa")
        self.tree.tag_configure("normal", background="#ffffff")

    def _autosize_columns(self, df):
        """Fit columns to visible content while keeping the table usable."""
        body_font = tkfont.Font(family="Segoe UI", size=9)
        heading_font = tkfont.Font(family="Segoe UI", size=9, weight="bold")
        for col in self.columns:
            values = []
            if not df.empty and col in df.columns:
                values = [str(value) for value in df[col].fillna("").tolist()]
            # Measuring only the longest string is much faster than measuring
            # every cell, especially for large Excel exports.
            longest_value = max(values, key=len, default="")
            longest = max(heading_font.measure(COLUMN_LABELS.get(col, col)), body_font.measure(longest_value))
            width = max(65, min(longest + 22, 300))
            self.tree.column(col, width=width, minwidth=min(width, 65), stretch=False)

    # ------------------------------------------------------------- shared DFM / uploads
    def _on_upload_data(self):
        """Reload every previously uploaded source from its saved file path."""
        uploads = db.get_all_uploads()
        reloaded = []
        missing = []

        for key in SOURCE_ORDER:
            info = uploads.get(key, {})
            path = info.get("file_path", "")
            if not path:
                missing.append(SOURCE_BUTTON_NAMES[key] + " (not uploaded yet)")
            elif not os.path.isfile(path):
                missing.append(f"{SOURCE_BUTTON_NAMES[key]} ({path})")
            else:
                self._handle_upload(key, path)
                reloaded.append(SOURCE_BUTTON_NAMES[key])

        codes_info = uploads.get("codes", {})
        codes_path = codes_info.get("file_path", "")
        if not codes_path:
            missing.append("Articoli (not uploaded yet)")
        elif not os.path.isfile(codes_path):
            missing.append(f"Articoli ({codes_path})")
        else:
            self._handle_codes_upload("codes", codes_path)
            reloaded.append("Articoli")

        summary = []
        if reloaded:
            summary.append("Reloaded: " + ", ".join(reloaded))
        if missing:
            summary.append("Not available:\n- " + "\n- ".join(missing))
        messagebox.showinfo("Upload Data", "\n\n".join(summary) or "No saved file paths found.")

    def sync_shared_dfm(self):
        """Load the DFM selected in either page from the shared persistent cache."""
        cache = load_dfm_cache()
        source_path = Path(str(cache.get("source_path", "")))
        if not source_path.is_file():
            return
        if getattr(self, "_shared_dfm_path", "") == str(source_path):
            return

        try:
            df, errors = data_loaders.load_dfm(str(source_path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not restore shared DFM file: %s", exc)
            return
        # Wincoint is the base file.  An empty but structurally valid Wincoint
        # is meaningful: it means all orders were removed and the saved
        # situation should be cleared on the next Refresh.
        if errors or df is None or (df.empty and key != "wincoint"):
            return

        self.loaded_frames["dfm"] = df
        self._shared_dfm_path = str(source_path)
        msg = f"✅ {len(df)} rows - {source_path.name}"
        self.source_rows["dfm"].set_status(True, f"✅ {len(df)} rows")
        db.save_upload("dfm", source_path.name, len(df), "ok", msg, file_path=str(source_path))

    def _save_shared_dfm(self, path):
        """Update the shared DFM cache when Situazione is the upload source."""
        try:
            entries = build_dfm_lookup(Path(path))
            if entries:
                save_dfm_cache(entries, Path(path).name, Path(path))
                self._shared_dfm_path = str(Path(path))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not update shared DFM reference: %s", exc)

    def _handle_upload(self, key, path):
        label, loader_fn = data_loaders.LOADERS[key]
        try:
            df, errors = loader_fn(path)
        except Exception as exc:  # noqa: BLE001
            errors = [f"An error occurred while reading the file: {exc}"]
            df = None

        if errors or df is None or df.empty:
            msg = "; ".join(errors) if errors else "The file is empty after filtering"
            self.source_rows[key].set_status(False, f"❌ {msg}")
            db.save_upload(key, os.path.basename(path), 0, "error", msg, file_path=str(path))
            self.loaded_frames.pop(key, None)
            return

        self.loaded_frames[key] = df
        msg = f"✅ {len(df)} rows - {os.path.basename(path)}"
        self.source_rows[key].set_status(True, f"✅ {len(df)} rows")
        db.save_upload(key, os.path.basename(path), len(df), "ok", msg, file_path=str(path))
        if key == "dfm":
            self._save_shared_dfm(path)
        logger.info("Situazione: %s uploaded — %d rows (%s)", key, len(df), os.path.basename(path))

    def _handle_codes_upload(self, _key, path):
        try:
            df, errors = data_loaders.load_codes(path)
        except Exception as exc:  # noqa: BLE001
            errors = [f"An error occurred while reading the file: {exc}"]
            df = None

        if errors or df is None or df.empty:
            msg = "; ".join(errors) if errors else "The file is empty"
            self.codes_row.set_status(False, f"❌ {msg}")
            return

        db.save_codes(df)
        msg = f"✅ Saved ({len(df)} codes) - {os.path.basename(path)}"
        self.codes_row.set_status(True, msg)
        db.save_upload("codes", os.path.basename(path), len(df), "ok", msg, file_path=str(path))
        logger.info("Situazione: yarn codes reference updated — %d codes (%s)", len(df), os.path.basename(path))

    def _refresh_source_labels_from_db(self):
        uploads = db.get_all_uploads()
        for key, row_widget in self.source_rows.items():
            info = uploads.get(key)
            if info:
                ok = info["status"] == "ok"
                if ok and not info.get("file_path"):
                    row_widget.set_status(
                        False,
                        f"⚠ {info['message']} (select once to save path)",
                    )
                else:
                    count = info.get("row_count", "")
                    status_text = f"✅ {count} rows" if ok else f"❌ {info['message']}"
                    row_widget.set_status(ok, status_text)
        codes_info = uploads.get("codes")
        if codes_info:
            if codes_info["status"] == "ok" and not codes_info.get("file_path"):
                self.codes_row.set_status(False, "⚠ Saved status only (select once to save path)")
            else:
                self.codes_row.set_status(codes_info["status"] == "ok",
                                           f"✅ Saved ({codes_info.get('row_count', '')} codes)")

    # -------------------------------------------------------------- refresh
    def _on_refresh(self):
        if "wincoint" not in self.loaded_frames:
            messagebox.showwarning("Missing data", "Upload the WINCOINT orders file before refreshing.")
            return

        missing = [SOURCE_BUTTON_NAMES[k] for k in SOURCE_ORDER if k not in self.loaded_frames]
        if missing:
            messagebox.showwarning(
                "Missing files",
                "These files have not been uploaded in this session:\n- " + "\n- ".join(missing) +
                "\n\nUpload them once, or use Upload Data after their paths have been saved. "
                "Refresh was cancelled so existing derived data is not cleared."
            )
            return

        # previous New Comment per Partita -- this becomes this round's Old Comment,
        # and the cascade in situazione_logic actively uses it, not just for history
        existing = db.get_all_states()
        old_comments = {p: s["new_comment"] for p, s in existing.items()}

        result_df = business_logic.compute_situation(
            orders_df=self.loaded_frames.get("wincoint"),
            dfm_df=self.loaded_frames.get("dfm"),
            data_prod_df=self.loaded_frames.get("data_prod"),
            copertura_df=self.loaded_frames.get("copertura"),
            uscita_df=self.loaded_frames.get("uscita"),
            qualita_df=self.loaded_frames.get("qualita"),
            codes_map=db.load_codes(),
            old_comments=old_comments,
        )

        # --- safety check: has ANYTHING changed vs what's already stored? ---
        any_change = False
        for _, r in result_df.iterrows():
            prev = existing.get(r["partita"])
            if prev is None or prev["new_comment"] != r["new_comment"]:
                any_change = True
                break

        current_partite = {
            str(value).strip()
            for value in result_df.get("partita", pd.Series(dtype=str)).tolist()
            if str(value).strip()
        }
        saved_partite = {str(value).strip() for value in existing}
        partite_changed = current_partite != saved_partite

        if not any_change and not partite_changed and existing:
            proceed = messagebox.askyesno(
                "No data changes",
                "The new data was compared with the saved data, and no new comment changed.\n"
                "Refreshing now will not move anything. Make sure you did not upload "
                "the same files by mistake.\n\nDo you want to continue?"
            )
            if not proceed:
                return
        else:
            proceed = messagebox.askyesno(
                "Confirm refresh",
                f"{len(result_df)} batches will be updated. The current New Comment will move to Old Comment "
                "for batches whose status changed. Continue?"
            )
            if not proceed:
                return

        removed = db.remove_states_not_in(current_partite)
        added, updated, unchanged = db.upsert_states(result_df.to_dict(orient="records"))
        self.summary_lbl.config(text=f"Added: {added}  |  Updated: {updated}  |  Unchanged: {unchanged}")
        logger.info(
            "Situazione: refreshed — added=%d updated=%d unchanged=%d removed=%d",
            added, updated, unchanged, removed,
        )
        self._load_table_from_db()

    def _load_table_from_db(self):
        states = db.get_all_states()
        self.current_df = pd.DataFrame(states.values())
        if not self.current_df.empty and "bagno" in self.current_df.columns:
            self.current_df = self.current_df.sort_values(
                by="bagno", ascending=True, key=lambda s: s.astype(str)
            )
            self.sort_state["bagno"] = False  # next click on Bagno heading reverses to Z-A
        self._render_tree(self.current_df)

    # -------------------------------------------------------------- display
    def _render_tree(self, df, autosize=True):
        self.tree.delete(*self.tree.get_children())
        if autosize:
            self._autosize_columns(df)
        if df.empty:
            return
        display_df = df.reindex(columns=self.columns, fill_value="")
        status_values = display_df["new_comment"].tolist()
        for index, values in enumerate(display_df.itertuples(index=False, name=None)):
            status = status_values[index]
            tag = "Ritinta" if str(status).startswith("Ritinta") else status
            if not tag:
                tag = "stripe" if index % 2 else ""
            self.tree.insert("", "end", values=values, tags=((tag,) if tag else ()))

    def _on_search_changed(self, *_args):
        """Debounce typing so the whole table is not redrawn per keystroke."""
        if self._filter_after_id is not None:
            self.after_cancel(self._filter_after_id)
        self._filter_after_id = self.after(180, self._apply_filter)

    def _apply_filter(self):
        self._filter_after_id = None
        q = self.search_var.get().strip().lower()
        if not q:
            self._render_tree(self.current_df)
            return
        if self.current_df.empty:
            return
        searchable = self.current_df.reindex(columns=self.columns, fill_value="").astype(str)
        mask = searchable.apply(lambda col: col.str.contains(q, case=False, regex=False)).any(axis=1)
        self._render_tree(self.current_df[mask], autosize=False)

    def _sort_by(self, col):
        ascending = self.sort_state.get(col, True)
        if self.current_df.empty:
            return
        self.current_df = self.current_df.sort_values(by=col, ascending=ascending, key=lambda s: s.astype(str))
        self.sort_state[col] = not ascending
        self._apply_filter()

    # -------------------------------------------------------------- export
    @staticmethod
    def _to_number(text):
        """Parse a stored value into a real number. Handles European comma
        decimals (e.g. '128,00') as well as plain '128'. Returns None for
        blank/unparseable values so the cell stays genuinely empty."""
        if text is None:
            return None
        s = str(text).strip()
        if not s or s.lower() == "nan":
            return None
        s = s.replace(".", "").replace(",", ".") if ("," in s and s.count(",") == 1) else s
        try:
            value = float(s)
        except ValueError:
            return None
        return int(value) if value.is_integer() else value

    @staticmethod
    def _to_date(text):
        """Parse a stored 'YYYY-MM-DD' string into a real date. Returns None
        for blank/unparseable values."""
        if not text:
            return None
        s = str(text).strip()
        if not s or s.lower() == "nan":
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None

    def _on_export(self):
        if self.current_df.empty:
            messagebox.showinfo("No data", "Refresh the data before exporting.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")],
                                             initialfile="Situazione.xlsx")
        if not path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule, FormulaRule

        # always export sorted by Bagno A -> Z, regardless of what's on screen
        export_df = self.current_df.sort_values(by="bagno", ascending=True,
                                                  key=lambda s: s.astype(str)).copy()

        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="16324F", end_color="16324F", fill_type="solid")

        wb = Workbook()
        ws = wb.active
        ws.title = "Situazione"
        headers = [header for _, header, _ in COLUMN_SPEC]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        date_format = "dd/mm/yyyy"

        for _, r in export_df.iterrows():
            row_values = []
            for key, _header, ctype in COLUMN_SPEC:
                raw = r.get(key, "")
                if ctype == "number":
                    row_values.append(self._to_number(raw))
                elif ctype == "date":
                    row_values.append(self._to_date(raw))
                else:
                    row_values.append(raw if raw not in (None, "") else None)
            ws.append(row_values)

            status = str(r.get("new_comment", ""))
            color = color_for_status(status if not status.startswith("Ritinta") else "Ritinta")
            fill = PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
            for col_index, (key, _header, ctype) in enumerate(COLUMN_SPEC, start=1):
                cell = ws.cell(row=ws.max_row, column=col_index)
                cell.fill = fill
                cell.font = Font(name="Arial")
                cell.alignment = center
                cell.border = border
                if ctype == "date" and cell.value is not None:
                    cell.number_format = date_format

        last_row = ws.max_row
        last_col_letter = get_column_letter(len(COLUMN_SPEC))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        ws.freeze_panes = "A2"
        col_letter = {key: get_column_letter(i) for i, (key, _h, _t) in enumerate(COLUMN_SPEC, start=1)}

        if last_row > 1:
            # Custom == "Check" -> red
            rng = f"{col_letter['custom']}2:{col_letter['custom']}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Check"'], fill=red_fill))

            # Bagno duplicated anywhere in the column -> red
            rng = f"{col_letter['bagno']}2:{col_letter['bagno']}{last_row}"
            first = f"{col_letter['bagno']}2"
            formula = f'AND({first}<>"",COUNTIF(${col_letter["bagno"]}$2:${col_letter["bagno"]}${last_row},{first})>1)'
            ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=red_fill))

            # Days in Q.C > 5 -> red
            rng = f"{col_letter['days_in_qc']}2:{col_letter['days_in_qc']}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["5"], fill=red_fill))

            # Consegna before today (overdue) -> red
            rng = f"{col_letter['consegna']}2:{col_letter['consegna']}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["TODAY()"], fill=red_fill))

        for key, header, _ctype in COLUMN_SPEC:
            letter = col_letter[key]
            values = [str(v) for v in export_df[key].tolist()] if key in export_df.columns else []
            longest = max([len(header)] + [len(v) for v in values]) if values else len(header)
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 35)

        wb.save(path)
        logger.info("Situazione: exported %d rows to %s", len(self.current_df), path)
        messagebox.showinfo("Completed", f"Export completed successfully:\n{path}")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Situazione (standalone test)")
    root.geometry("1400x760")
    tab = SituazioneTab(root)
    tab.pack(fill="both", expand=True)
    root.mainloop()
