"""
ordini_elvy.py — Derives the "Ordini ELVY" sheet from parsed Purchase Order
rows (see pdf_parser.OrderRow), for the Elvy raw-yarn dyeing workflow.

Column mapping (fixed values confirmed with the user; header labels and
types match EXCEL_PER_ORDINE_VENDITA_EGITTO.xlsx, the real ERP import
template, so this sheet can be pasted straight into it)
--------------------------------------------------------------------------
CLIENTE              — always 3009 (number).
ARTICOLO             — the row's raw code (already G-prefixed via the Elvy
                        mapping) converted to "C" — it's entering dyeing.
COLORE                — copied from the row's DFM colour lookup (number).
Q.TA                 — copied as-is (number).
CONSEGNA             — today + 14 days (real date).
COMMENTO             — "PG-X-PO-<POD No>-<year>", year from the row's Po Date.
LAVORANTE            — always 900901 (number).
LAV. SUCC            — always 900161 (number).
DATA RICONSEGNA      — CONSEGNA - 1 day (real date).
MAG. GREGGIO         — always 900923 (number).
SIGLA DISPOSIZIONE   — always "D".
BAGNO PREPOSTO       — always blank.
GRUPPO MACCHINA      — the physical machine number for the row's Abbina
                        cone count (see abbina_calculator.MACHINE_CODES),
                        as a number.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

from abbina_calculator import MACHINE_CODES
from pdf_parser import OrderRow
from utils import clean_text

CLIENTE_CODE = 3009
LAVORANTE_CODE = 900901
LAVORANTE_SUCCESSIVO_CODE = 900161
STOCK_MAG_CODE = 900923
SIGLA_DISPO_CODE = "D"
CONSEGNA_LEAD_DAYS = 14

RE_ABBINA_CONES = re.compile(r"Machine\s+(\d+)\s*Cones?", re.IGNORECASE)
RE_YEAR = re.compile(r"(\d{4})")


@dataclass
class OrdiniElvyRow:
    """One derived row for the "Ordini ELVY" sheet."""
    cliente: int = CLIENTE_CODE
    articolo_delta: str = ""
    coloredfm: int | None = None
    quantity_cones: float | None = None
    data_consegna: datetime | None = None
    commento: str = ""
    lavorante: int = LAVORANTE_CODE
    lavorante_successivo: int = LAVORANTE_SUCCESSIVO_CODE
    data_riconsegna: datetime | None = None
    stock_mag: int = STOCK_MAG_CODE
    sigla_dispo: str = SIGLA_DISPO_CODE
    bagno_proposto: str = ""
    macchina: int | None = None


def _to_c_code(articolo_delta: str) -> str:
    """Convert a G-prefixed raw code to its C-prefixed dyeing-entry form."""
    articolo_delta = clean_text(articolo_delta)
    if articolo_delta.upper().startswith("G"):
        return "C" + articolo_delta[1:]
    return ""


def _extract_year(po_date: str) -> str:
    """Pull the 4-digit year out of a Po Date string like '19/4/2026'."""
    m = RE_YEAR.search(clean_text(po_date))
    return m.group(1) if m else ""


def _to_int(value: str) -> int | None:
    value = clean_text(value)
    try:
        return int(value)
    except ValueError:
        return None


def _machine_code_for_abbina(abbina: str) -> int | None:
    """Look up the MACCHINA code from an Abbina value like 'Machine 72 Cones'."""
    m = RE_ABBINA_CONES.search(clean_text(abbina))
    if not m:
        return None
    return MACHINE_CODES.get(int(m.group(1)))


def build_ordini_elvy_rows(rows: list[OrderRow]) -> list[OrdiniElvyRow]:
    """Build one OrdiniElvyRow per OrderRow, using today's date as the basis
    for CONSEGNA / DATA RICONSEGNA."""
    today = datetime.now()
    data_consegna_dt = today + timedelta(days=CONSEGNA_LEAD_DAYS)
    data_riconsegna_dt = data_consegna_dt - timedelta(days=1)

    out: list[OrdiniElvyRow] = []
    for r in rows:
        year = _extract_year(r.po_date)
        commento = f"PG-X-PO-{clean_text(r.pod_no)}-{year}" if year else f"PG-X-PO-{clean_text(r.pod_no)}"

        out.append(OrdiniElvyRow(
            articolo_delta=_to_c_code(r.articolo_delta),
            coloredfm=_to_int(r.coloredfm),
            quantity_cones=r.quantity_cones,
            data_consegna=data_consegna_dt,
            commento=commento,
            data_riconsegna=data_riconsegna_dt,
            macchina=_machine_code_for_abbina(r.abbina),
        ))
    return out


# ---------------------------------------------------------------------------
# Updating an existing external ERP file in place
# ---------------------------------------------------------------------------

# Header text (normalized: stripped, case-insensitive) -> OrdiniElvyRow
# attribute. Matches EXCEL_PER_ORDINE_VENDITA_EGITTO.xlsx's own headers, so
# the target file's column ORDER doesn't have to match ours exactly — only
# the header text does.
_HEADER_TO_ATTR = {
    "cliente": "cliente",
    "articolo": "articolo_delta",
    "colore": "coloredfm",
    "q.ta": "quantity_cones",
    "consegna": "data_consegna",
    "commento": "commento",
    "lavorante": "lavorante",
    "lav. succ": "lavorante_successivo",
    "data riconsegna": "data_riconsegna",
    "mag. greggio": "stock_mag",
    "sigla disposizione": "sigla_dispo",
    "bagno preposto": "bagno_proposto",
    "gruppo macchina": "macchina",
}


def update_existing_ordini_file(target_path: Path, ordini_rows: list[OrdiniElvyRow]) -> int:
    """
    Open the existing ERP import file at *target_path*, clear every data
    row (everything from row 2 down — the header row is left untouched),
    write *ordini_rows* starting at row 2 matched to that file's own
    header text, save, and close.

    Returns the number of rows written. Raises if the file can't be
    opened (e.g. it's currently open in Excel) or has no recognizable
    header row.
    """
    import openpyxl  # local import: this module doesn't need openpyxl otherwise

    wb = openpyxl.load_workbook(target_path)
    ws = None
    col_attr: dict[int, str] = {}
    # ERP templates may contain several sheets and the active sheet is not
    # guaranteed to be the import sheet. Find the sheet by its headers.
    for candidate in wb.worksheets:
        candidate_attrs: dict[int, str] = {}
        header_cells = next(candidate.iter_rows(min_row=1, max_row=1))
        for cell in header_cells:
            if cell.value is None:
                continue
            key = " ".join(str(cell.value).strip().lower().split())
            attr = _HEADER_TO_ATTR.get(key)
            if attr:
                candidate_attrs[cell.column] = attr
        if candidate_attrs:
            ws = candidate
            col_attr = candidate_attrs
            break

    if ws is None or not col_attr:
        wb.close()
        raise ValueError(
            "None of this file's column headers match the expected Ordini "
            "ELVY headers (CLIENTE, ARTICOLO, COLORE, ...) — is this the "
            "right file?"
        )

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_offset, row_data in enumerate(ordini_rows):
        row_idx = 2 + row_offset
        for col_idx, attr in col_attr.items():
            value = getattr(row_data, attr, None)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if isinstance(value, datetime):
                cell.number_format = "DD/MM/YYYY"

    wb.save(target_path)
    wb.close()
    return len(ordini_rows)


_FILATO_HEADER_TO_ATTR = {
    "articolo": "articolo",
    "titolo": "titolo",
    "partita": "partita",
    "rocce": "rocce",
    "peso": "peso",
    "تحضير خام": "label",
}


def read_filato_tinturia_sheet(source_path: Path) -> list[RawYarnMatch]:
    """Read an existing 'Filato x Tinturia' sheet back into RawYarnMatch rows."""
    import openpyxl  # local import: this module doesn't need openpyxl otherwise

    wb = openpyxl.load_workbook(source_path, data_only=True)
    if "Filato x Tinturia" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{source_path.name} has no \"Filato x Tinturia\" sheet.")
    ws = wb["Filato x Tinturia"]

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    col_attr: dict[int, str] = {}
    for cell in header_cells:
        if cell.value is None:
            continue
        attr = _FILATO_HEADER_TO_ATTR.get(str(cell.value).strip().lower())
        if attr:
            col_attr[cell.column] = attr

    rows: list[RawYarnMatch] = []
    for row in ws.iter_rows(min_row=2):
        values = {attr: row[col_idx - 1].value for col_idx, attr in col_attr.items()}
        if not any(v not in (None, "") for v in values.values()):
            continue
        rows.append(RawYarnMatch(
            articolo=str(values.get("articolo") or ""),
            titolo=str(values.get("titolo") or ""),
            partita=str(values.get("partita") or ""),
            rocce=float(values.get("rocce") or 0),
            peso=float(values.get("peso") or 0),
            label=str(values.get("label") or "تحضير خام"),
        ))
    wb.close()
    return rows


def update_existing_filato_file(target_path: Path, matches: list[RawYarnMatch]) -> int:
    """
    Same transfer logic as update_existing_ordini_file(): open the existing
    file at *target_path*, match its header row (any sheet/order, matched
    by header TEXT) to Articolo/Titolo/Partita/Rocce/Peso/تحضير خام, clear
    every data row from row 2 down, write *matches* starting at row 2, save.
    Returns the number of rows written.
    """
    import openpyxl  # local import: this module doesn't need openpyxl otherwise

    wb = openpyxl.load_workbook(target_path)
    ws = wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    col_attr: dict[int, str] = {}
    for cell in header_cells:
        if cell.value is None:
            continue
        attr = _FILATO_HEADER_TO_ATTR.get(str(cell.value).strip().lower())
        if attr:
            col_attr[cell.column] = attr

    if not col_attr:
        wb.close()
        raise ValueError(
            "None of this file's column headers match the expected Filato x "
            "Tinturia headers (Articolo, Titolo, Partita, Rocce, Peso, تحضير خام) "
            "— is this the right file?"
        )

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_offset, match in enumerate(matches):
        row_idx = 2 + row_offset
        for col_idx, attr in col_attr.items():
            ws.cell(row=row_idx, column=col_idx).value = getattr(match, attr, None)

    wb.save(target_path)
    wb.close()
    return len(matches)


def export_ordini_full(target_path: Path, ordini_rows: list) -> int:
    """
    Create a brand-new "EXCEL PER ORDINE VENDITA EGITTO" workbook at
    *target_path*, fully formatted (header styling, column widths, freeze
    panes/filter) with ordini_rows -- unlike update_existing_ordini_file(),
    this does not open/edit an existing file, it (re)creates the file from
    scratch every time it's called, overwriting whatever was there before.
    Returns the number of rows written.
    """
    import openpyxl  # local import: this module doesn't need openpyxl otherwise
    from excel_exporter import ORDINI_ELVY_COLUMNS, apply_column_widths, freeze_and_filter, write_data, write_header

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ORDINE VENDITA EGITTO"
    ws.sheet_view.rightToLeft = False
    write_header(ws, ORDINI_ELVY_COLUMNS)
    write_data(ws, ORDINI_ELVY_COLUMNS, ordini_rows)
    apply_column_widths(ws, ORDINI_ELVY_COLUMNS)
    freeze_and_filter(ws, ORDINI_ELVY_COLUMNS)

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    wb.close()
    return len(ordini_rows)


def export_filato_full(target_path: Path, matches: list["RawYarnMatch"]) -> int:
    """
    Create a brand-new "Filato x Tinturia" workbook at *target_path*, fully
    formatted the same way as the sheet embedded in the PO/Kamal export --
    unlike update_existing_filato_file(), this does not open/edit an
    existing file, it (re)creates the file from scratch every time it's
    called, overwriting whatever was there before.
    Returns the number of rows written.
    """
    import openpyxl  # local import: this module doesn't need openpyxl otherwise
    from openpyxl.styles import Alignment, Font
    from excel_exporter import FILATO_TINTURIA_COLUMNS, apply_column_widths, freeze_and_filter, write_data, write_header

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filato x Tinturia"
    ws.sheet_view.rightToLeft = False
    write_header(ws, FILATO_TINTURIA_COLUMNS)
    write_data(ws, FILATO_TINTURIA_COLUMNS, matches)
    apply_column_widths(ws, FILATO_TINTURIA_COLUMNS)
    freeze_and_filter(ws, FILATO_TINTURIA_COLUMNS)

    label_col_idx = next(
        (i + 1 for i, (_, attr, _) in enumerate(FILATO_TINTURIA_COLUMNS) if attr == "label"), None
    )
    if label_col_idx:
        bold_label_font = Font(bold=True, name="Calibri", size=11)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=label_col_idx, max_col=label_col_idx):
            for cell in row:
                cell.font = bold_label_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    wb.close()
    return len(matches)


@dataclass
class RawYarnMatch:
    """One row assigned for the "Filato x Tinturia" summary sheet."""
    articolo: str = ""
    titolo: str = ""
    partita: str = ""
    rocce: float = 0.0
    peso: float = 0.0
    label: str = "تحضير خام"


def match_raw_yarn(
    ordini_rows: list["OrdiniElvyRow"],
    magazino_summary: "pd.DataFrame",
    codes_map: dict | None = None,
    quantity_attr: str = "quantity_cones",
) -> list[RawYarnMatch]:
    """
    Fills in the Partita number in place of "X" in "PG-X-..." for rows
    that can be covered by available raw yarn, and returns the list of
    batches actually assigned (for the "Filato x Tinturia" export sheet).

    magazino_summary: output of magazino_logic.summarize_by_partita()
      (columns: articolo, partita, mag_rocche, mag_peso) -- articolo here
      is the raw (G-prefixed) code.
    codes_map: optional {articolo_filato: titolo} lookup (Articoli.xlsx),
      used only to fill in the Titolo column of the returned matches.
    quantity_attr: the OrdiniElvyRow attribute to use for requested raw-yarn
      quantity (default: quantity_cones). Kamal rows use peso_kg.

    Matching rule (as specified): group the order rows still needing raw
    yarn ("PG-X" in commento) by their G-code Articolo. For each group, if
    a single available batch covers the group's total quantity, assign that
    batch's Partita to every row in the group. Otherwise, try to cover
    each row individually with the smallest batch that fits it, leaving
    any row that can't be covered as "PG-X".
    """
    if magazino_summary is None or magazino_summary.empty:
        return []

    batch_capacity_attr = "rocche" if quantity_attr == "quantity_cones" else "peso"

    pool: dict[str, list[dict]] = {}
    for _, r in magazino_summary.iterrows():
        pool.setdefault(str(r["articolo"]), []).append(
            {"partita": str(r["partita"]), "rocche": float(r["mag_rocche"]), "peso": float(r["mag_peso"])}
        )

    groups: dict[str, list[int]] = {}
    for i, row in enumerate(ordini_rows):
        if "PG-X" not in (row.commento or ""):
            continue
        articolo_c = clean_text(row.articolo_delta).upper()
        if not articolo_c.startswith("C"):
            continue
        articolo_g = "G" + articolo_c[1:]
        groups.setdefault(articolo_g, []).append(i)

    matches: list[RawYarnMatch] = []

    def _assign(idxs, batch, quantity_used):
        for i in idxs:
            ordini_rows[i].commento = ordini_rows[i].commento.replace("PG-X", f"PG-{batch['partita']}")
        articolo_c_for_lookup = "C" + articolo_g[1:] if articolo_g.upper().startswith("G") else articolo_g
        titolo = (codes_map or {}).get(articolo_c_for_lookup, "")
        if batch_capacity_attr == "rocche":
            peso_used = batch["peso"] * (quantity_used / batch["rocche"]) if batch["rocche"] else 0.0
            rocce_used = quantity_used
        else:
            peso_used = quantity_used
            rocce_used = quantity_used
        matches.append(RawYarnMatch(
            articolo=articolo_g,
            titolo=titolo,
            partita=batch["partita"],
            rocce=rocce_used,
            peso=round(peso_used, 2),
        ))

    for articolo_g, idxs in groups.items():
        batches = pool.get(articolo_g, [])
        if not batches:
            continue
        total_needed = sum(getattr(ordini_rows[i], quantity_attr, 0) or 0 for i in idxs)

        candidates = sorted(
            (b for b in batches if b.get(batch_capacity_attr, 0) >= total_needed),
            key=lambda b: b.get(batch_capacity_attr, 0),
        )
        if candidates:
            batch = candidates[0]
            _assign(idxs, batch, total_needed)
            batches.remove(batch)
            continue

        for i in idxs:
            row = ordini_rows[i]
            need = getattr(row, quantity_attr, 0) or 0
            fit = sorted(
                (b for b in batches if b.get(batch_capacity_attr, 0) >= need),
                key=lambda b: b.get(batch_capacity_attr, 0),
            )
            if fit:
                batch = fit[0]
                _assign([i], batch, need)
                batches.remove(batch)
            # else: leave as PG-X -- nothing in stock covers this row yet

    return matches
