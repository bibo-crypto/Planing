"""
kamal_excel_exporter.py
Builds the Kamal order Excel workbook: a raw "Kamal Order" sheet (parsed
PDF rows, including Dye Type), an "Ordine Kamal" sheet (matched rows,
mirroring Ordini ELVY's structure), and -- if raw yarn stock was supplied
-- a "Filato x Tinturia" sheet listing what was assigned.
"""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

import openpyxl
from utils import clean_text

from excel_exporter import (
    FILATO_TINTURIA_COLUMNS,
    apply_column_widths,
    freeze_and_filter,
    write_data,
    write_header,
)
from kamal_parser import KamalOrderRow
from ordine_kamal import assign_ordine_kamal_machines, build_ordine_kamal_rows, match_by_lotto
from ordini_elvy import RawYarnMatch, match_raw_yarn
from utils import logger
from abbina_calculator import _smallest_fitting_machine, MACHINE_CODES

RAW_KAMAL_COLUMNS: list[tuple[str, str, str]] = [
    ("Reply No",       "reply_no",       "text"),
    ("Order Date",     "order_date",     "text"),
    ("Dye Type",       "dye_type",       "text"),
    ("Notes",          "notes",          "text"),
    ("Sender Name",    "sender_name",    "text"),
    ("Sender Msg No",  "sender_msg_no",  "text"),
    ("COLOREDFM",      "coloredfm",      "text"),
    ("Titolo",         "titolo",         "text"),
    ("Colore",         "colore_name",    "text"),
    ("Peso (kg)",      "peso_kg",        "number"),
    ("Abbina",         "abbina",         "text"),
]

ORDINE_KAMAL_COLUMNS: list[tuple[str, str, str]] = [
    ("CLIENTE",              "cliente",              "integer"),
    ("ARTICOLO",             "articolo_delta",       "text"),
    ("COLORE",               "coloredfm",            "integer"),
    ("Q.TA",                 "peso_kg",              "number"),
    ("CONSEGNA",             "data_consegna",        "date"),
    ("COMMENTO",             "commento",             "text"),
    ("LAVORANTE",            "lavorante",            "integer"),
    ("LAV. SUCC",            "lavorante_successivo", "integer"),
    ("DATA RICONSEGNA",      "data_riconsegna",      "date"),
    ("MAG. GREGGIO",         "stock_mag",            "integer"),
    ("SIGLA DISPOSIZIONE",   "sigla_dispo",          "text"),
    ("BAGNO PREPOSTO",       "bagno_proposto",       "text"),
    ("GRUPPO MACCHINA",      "macchina",             "integer"),
]


class KamalExcelExporter:
    def __init__(self, rows: Sequence[KamalOrderRow], output_path: Path,
                 dfm_c170_entries: list[dict] | None = None,
                 magazino_summary=None, codes_map: dict | None = None,
                 lotti_summary=None) -> None:
        self.rows = rows
        self.output_path = output_path
        self.dfm_c170_entries = dfm_c170_entries or []
        self.magazino_summary = magazino_summary
        self.codes_map = codes_map
        self.lotti_summary = lotti_summary

    def export(self) -> None:
        logger.info("Exporting %d Kamal row(s) → %s", len(self.rows), self.output_path.name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kamal Order"
        ws.sheet_view.rightToLeft = False
        write_header(ws, RAW_KAMAL_COLUMNS)
        write_data(ws, RAW_KAMAL_COLUMNS, self.rows)
        apply_column_widths(ws, RAW_KAMAL_COLUMNS)
        freeze_and_filter(ws, RAW_KAMAL_COLUMNS)

        kamal_rows = build_ordine_kamal_rows(list(self.rows), self.dfm_c170_entries)

        # Primary: match by Kamal's own raw-yarn message number (Lotto),
        # which is a deterministic reference from their supplier.
        if self.lotti_summary is not None and not self.lotti_summary.empty:
            lotto_matched = match_by_lotto(kamal_rows, self.lotti_summary)
            logger.info("Raw yarn matched via Lotto: %d row(s)", lotto_matched)

        raw_yarn_matches = self._build_lotto_matches(kamal_rows)

        if self.magazino_summary is not None and not self.magazino_summary.empty:
            fallback_matches = match_raw_yarn(
                kamal_rows, self.magazino_summary, self.codes_map, quantity_attr="peso_kg"
            )
            if fallback_matches:
                raw_yarn_matches.extend(fallback_matches)
            logger.info("Raw yarn matched via quantity fallback: %d batch(es) assigned", len(fallback_matches))

        ws2 = wb.create_sheet("Ordine Kamal")
        ws2.sheet_view.rightToLeft = False
        # Use the same machine calculation for the workbook and ERP update.
        assign_ordine_kamal_machines(kamal_rows)

        # Also set Abbina on the original parsed rows (first sheet)
        for r in self.rows:
            c = clean_text(getattr(r, "coloredfm", ""))
            if not c:
                continue
            matching = next((row for row in kamal_rows if clean_text(row.coloredfm) == c), None)
            r.abbina = matching.abbina if matching else ""

        # The raw sheet was written before the derived rows were calculated.
        # Update its Abbina cells after the machine assignment so the parsed
        # Kamal Order sheet contains the same values used by Ordine Kamal.
        abbina_col = next(
            (index + 1 for index, (_header, attr, _dtype) in enumerate(RAW_KAMAL_COLUMNS)
             if attr == "abbina"),
            None,
        )
        if abbina_col is not None:
            for row_index, parsed_row in enumerate(self.rows, start=2):
                ws.cell(row=row_index, column=abbina_col).value = getattr(parsed_row, "abbina", "") or ""

        write_header(ws2, ORDINE_KAMAL_COLUMNS)
        write_data(ws2, ORDINE_KAMAL_COLUMNS, kamal_rows)
        apply_column_widths(ws2, ORDINE_KAMAL_COLUMNS)
        freeze_and_filter(ws2, ORDINE_KAMAL_COLUMNS)

        if raw_yarn_matches:
            ws3 = wb.create_sheet("Filato x Tinturia")
            ws3.sheet_view.rightToLeft = False
            write_header(ws3, FILATO_TINTURIA_COLUMNS)
            write_data(ws3, FILATO_TINTURIA_COLUMNS, raw_yarn_matches)
            apply_column_widths(ws3, FILATO_TINTURIA_COLUMNS)
            freeze_and_filter(ws3, FILATO_TINTURIA_COLUMNS)

            from openpyxl.styles import Alignment, Font
            label_col_idx = next(
                (i + 1 for i, (_, attr, _) in enumerate(FILATO_TINTURIA_COLUMNS) if attr == "label"), None
            )
            if label_col_idx:
                bold_label_font = Font(bold=True, name="Calibri", size=11)
                for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row,
                                          min_col=label_col_idx, max_col=label_col_idx):
                    for cell in row:
                        cell.font = bold_label_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        logger.info("Export completed: %s", self.output_path)

    def _build_lotto_matches(self, kamal_rows: list[OrdineKamalRow]) -> list[RawYarnMatch]:
        matches: list[RawYarnMatch] = []
        if self.lotti_summary is None or self.lotti_summary.empty:
            return matches

        assigned: dict[tuple[str, str, str], float] = {}
        for row in kamal_rows:
            commento = clean_text(row.commento)
            if "PG-X" in commento.upper():
                continue
            partita = self._extract_partita(commento)
            if not partita or partita.upper() == "X":
                continue
            articolo = (
                "G" + row.articolo_delta[1:]
                if row.articolo_delta.upper().startswith("C")
                else row.articolo_delta
            )
            titolo = (self.codes_map or {}).get(row.articolo_delta, "")
            key = (articolo, titolo, partita)
            assigned[key] = assigned.get(key, 0.0) + (row.peso_kg or 0.0)

        for (articolo, titolo, partita), peso in assigned.items():
            matches.append(RawYarnMatch(
                articolo=articolo,
                titolo=titolo,
                partita=partita,
                rocce=round(peso, 2),
                peso=round(peso, 2),
            ))
        return matches

    @staticmethod
    def _extract_partita(commento: str) -> str:
        if not commento:
            return ""
        normalized = commento.upper()
        idx = normalized.find("PG-")
        if idx == -1:
            return ""
        start = idx + 3
        end = start
        while end < len(commento) and commento[end] not in (" ", "\t", "\r", "\n", "-", "("):
            end += 1
        return commento[start:end].strip()
