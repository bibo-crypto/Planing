"""
dfm_lookup.py — Loads the DFM.xlsx historical dye-job export (Elvy-specific)
and looks up a Purchase Order row's colour against it to retrieve Delta's
own colour code (COLOREDFM) and colour name (CLDESCR).

Background
----------
DFM.xlsx is a historical export of dye jobs. Article codes (ARTICOLODFM)
for the Elvy client always start with "C130...". Their raw (griege)
equivalent — "Articolo Delta" — is the same digits with "G" instead of
"C" (e.g. "C130027S" <-> "G130027S"). "Articolo Delta" itself is resolved
separately from a PDF row's Article No via the user-maintained mapping in
elvy_mapping.py; this module only handles the colour side.

Matching algorithm (verified against real DFM.xlsx + confirmed examples)
--------------------------------------------------------------------------
1. The PDF's Colore field (OrderRow.colour, e.g. "440") is NOT the same
   number as COLOREDFM or any fixed slice of it — matching against
   COLOREDFM directly is unreliable. What actually works is CLDESCR: it's
   almost always "EL-<colour><1-2 extra client digits><optional suffix>",
   e.g. "EL-44011-DOUBLEYARN" for colour "440". So colour is matched as a
   prefix of the digits after "EL-" in CLDESCR.

2. DESCRIZARTICOLOLI (e.g. "0070  100.00 2") ends with a twist digit that
   exactly matches Ne's denominator: Ne "100/2" -> twist "2" (double),
   "80/1" -> "1" (single), ".../3" -> "3" (triple). This is a reliable
   structural fact, unlike the colour digits.

3. Dye type / yarn keywords: CLDESCR sometimes carries extra qualifiers
   that must also match:
     - dye_type containing "react" (Reattivo/Reactive) -> CLDESCR must
       also contain "REATT" or "REACT".
     - yarn containing "tencel" -> CLDESCR must also contain "TENCEL".
     - yarn containing "kashmir"/"cashmere" -> CLDESCR must also contain
       "KASHMIR" (no confirmed example in the data yet, kept best-effort).
     - Vat is the default/unmarked case — no extra keyword required.

4. Search order:
     a. PRIMARY: rows whose ARTICOLODFM matches this row's Articolo Delta
        (G->C swapped) exactly, plus the colour/twist/keyword filters
        above. Most recent DATAINS wins among matches.
     b. FALLBACK: if (a) finds nothing — e.g. this exact article was
        never dyed in this colour before — broaden to ANY Elvy (C130*)
        article with the same colour/twist/keyword filters. This mirrors
        what happens at the factory: the same colour recipe is reused
        across a family of related articles, so the most recent match
        from a *different* article is still the right answer far more
        often than leaving the cell blank.
     c. If neither finds anything, leave both columns blank — never guess.

This is explicitly a best-effort heuristic on messy, free-text factory
data (confirmed ~85-95% consistent per criterion, not 100%) — it is
expected to be refined further as more real examples turn up.

Caching
-------
Parsing the full DFM export is a one-off, explicit action (via the Elvy
tab's "Load DFM Color Reference" button). The *filtered* Elvy-only rows
are then cached as JSON in the same writable per-user AppData directory
used for settings/logs, so the app doesn't need the original (large)
DFM.xlsx again until the user reloads a fresh export.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from utils import APP_DATA_DIR, clean_text, logger

CACHE_FILE = APP_DATA_DIR / "settings" / "dfm_color_cache.json"

# Only rows whose ARTICOLODFM starts with this prefix are relevant to Elvy.
ELVY_ARTICLE_PREFIX = "C130"

REQUIRED_COLUMNS = ("ARTICOLODFM", "COLOREDFM", "CLDESCR", "DESCRIZARTICOLOLI", "DATAINS")

# Keyword groups: if the PDF-side condition (checked against dye_type or
# yarn) is met, CLDESCR must contain at least one of the listed variants.
_REACTIVE_CLDESCR_HINTS = ("REATT", "REACT")
_TENCEL_CLDESCR_HINTS = ("TENCEL",)
_KASHMIR_CLDESCR_HINTS = ("KASHMIR", "KASCHMIR", "CASHMER")


def _parse_date(text: str) -> datetime | None:
    """Parse a DD/MM/YYYY date string. Returns None if unparseable."""
    text = clean_text(text)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%d/%m/%Y")
    except ValueError:
        return None


def _twist_digit(descrizarticololi: str) -> str:
    """Extract the trailing twist digit ('1'/'2'/'3'/...) from a
    DESCRIZARTICOLOLI string like '0070  100.00 2'. Returns "" if absent."""
    parts = clean_text(descrizarticololi).split()
    return parts[-1] if parts and parts[-1].isdigit() else ""


def _ne_twist_digit(ne: str) -> str:
    """Extract the twist digit from an Ne value like '100/2'. Returns "" if absent."""
    ne = clean_text(ne)
    if "/" in ne:
        denom = ne.split("/")[-1].strip()
        if denom.isdigit():
            return denom
    return ""


# ---------------------------------------------------------------------------
# Building / caching the filtered reference
# ---------------------------------------------------------------------------

def build_dfm_lookup(xlsx_path: Path) -> list[dict[str, str]]:
    """
    Read *xlsx_path* (a DFM.xlsx-style export) and return the list of Elvy
    (ARTICOLODFM starting with "C130") rows, reduced to just the fields
    needed for matching: articolo, coloredfm, cldescr, twist, date (ISO
    string, for sorting — "" if unparseable).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header_row]

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(
            f"This file doesn't look like a DFM export — missing column(s): "
            f"{', '.join(missing)}"
        )
    col_idx = {name: headers.index(name) for name in REQUIRED_COLUMNS}

    entries: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        articolo = clean_text(row[col_idx["ARTICOLODFM"]])
        if not articolo.startswith(ELVY_ARTICLE_PREFIX):
            continue

        coloredfm = clean_text(row[col_idx["COLOREDFM"]])
        cldescr = clean_text(row[col_idx["CLDESCR"]])
        if not coloredfm or not cldescr:
            continue

        date = _parse_date(row[col_idx["DATAINS"]])
        entries.append({
            "articolo": articolo,
            "coloredfm": coloredfm,
            "cldescr": cldescr,
            "twist": _twist_digit(row[col_idx["DESCRIZARTICOLOLI"]]),
            "date": date.isoformat() if date else "",
        })

    wb.close()
    logger.info(
        "Built DFM colour reference: %d Elvy row(s) from %s",
        len(entries), xlsx_path.name,
    )
    return entries


def save_dfm_cache(entries: list[dict[str, str]], source_name: str) -> None:
    """Persist *entries* (plus metadata) to the cache file."""
    try:
        CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload: dict[str, Any] = {
            "source_file": source_name,
            "loaded_at": datetime.now().isoformat(timespec="seconds"),
            "entries": entries,
        }
        CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    except Exception as exc:  # noqa: BLE001
        logger.warning("Could not save DFM colour cache: %s", exc)


def load_dfm_cache() -> dict[str, Any]:
    """
    Load the cached DFM reference from disk.
    Returns {"source_file": "", "loaded_at": "", "entries": []} if no
    cache exists yet or it can't be read.
    """
    empty: dict[str, Any] = {"source_file": "", "loaded_at": "", "entries": []}
    try:
        if CACHE_FILE.is_file():
            data = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict) and isinstance(data.get("entries"), list):
                return data
    except Exception as exc:  # noqa: BLE001
        logger.warning("Could not load DFM colour cache: %s", exc)
    return empty


# ---------------------------------------------------------------------------
# Lookup
# ---------------------------------------------------------------------------

# Real POs express Colour in several different raw formats:
#   "205"                  -> plain numeric code
#   "009.Red Water Color"  -> numeric code + descriptive name after a dot
#   "R.1372"               -> "R." (Reactive marker) + numeric code — this
#                             one does NOT require a literal "R" in CLDESCR;
#                             it just echoes dye_type, already checked
#                             separately via _required_cldescr_hints.
#   "G.203"                -> "G" client marker + numeric code
#   "MO_0030.CELESTE"      -> "MO" client marker + numeric code + name
#
# "G" and "MO" are NOT decorative — they mark colours belonging to a
# different client/family that happen to share DFM's colour numbering.
# CLDESCR encodes the same marker as "EL-G-<num>" / "EL-MO-<num>" (with a
# dash) vs plain "EL-<num>" — and these can be genuinely different colours
# for the very same numeric code (confirmed: article C130078S has both
# "EL-45311" (plain) and "EL-G-453" (G-marked) as two distinct COLOREDFM
# entries). So the marker must match exactly, not just be tolerated.
_COLOUR_PREFIX_RE = re.compile(r"^(?P<marker>R|G|MO)[._\-\s]+", re.IGNORECASE)
_COLOUR_CODE_RE = re.compile(r"(\d+)[.\s_\-]*(.*)$")

_CLDESCR_CODE_RE = re.compile(r"EL-(?:(?P<marker>MO|G)-?)?(?P<code>\d+)", re.IGNORECASE)


def _extract_colour_code(colour: str) -> tuple[str, str, str]:
    """
    Split a raw Colour field into (numeric_code, descriptive_name, marker).
    marker is "G" or "MO" when that client-prefix is present, "" otherwise
    (including for the "R." Reactive prefix, which carries no CLDESCR
    marker of its own). Returns ("", "", "") if no digits are found.
    """
    text = clean_text(colour)
    prefix_match = _COLOUR_PREFIX_RE.match(text)
    marker = ""
    if prefix_match:
        found = prefix_match.group("marker").upper()
        if found in ("G", "MO"):
            marker = found
        text = text[prefix_match.end():]
    m = _COLOUR_CODE_RE.match(text)
    if not m:
        return "", "", ""
    return m.group(1), m.group(2).strip(" .-_"), marker


def _required_cldescr_hints(dye_type: str, yarn: str) -> list[str]:
    """
    Return groups of CLDESCR keyword alternatives that must each be
    satisfied (at least one variant per group) given this row's dye_type
    and yarn composition.
    """
    dye_type = clean_text(dye_type).lower()
    yarn = clean_text(yarn).lower()
    groups: list[tuple[str, ...]] = []

    if "react" in dye_type:
        groups.append(_REACTIVE_CLDESCR_HINTS)
    if "tencel" in yarn:
        groups.append(_TENCEL_CLDESCR_HINTS)
    if "kashmir" in yarn or "cashmere" in yarn or "kaschmir" in yarn:
        groups.append(_KASHMIR_CLDESCR_HINTS)
    return groups


def _matches_filters(
    entry: dict[str, str],
    colour_code: str,
    marker: str,
    twist: str,
    hint_groups: list[tuple[str, ...]],
    name_hint: str = "",
) -> bool:
    cldescr_upper = entry["cldescr"].upper()

    m = _CLDESCR_CODE_RE.match(cldescr_upper)
    if not m:
        return False
    code = m.group("code")
    # CLDESCR's number is either the colour code exactly, or the code plus
    # a standard 2-digit client sub-code suffix (e.g. "440" -> "44011").
    # Any OTHER extra length (e.g. "151" -> "151711") is a coincidental
    # prefix match against a genuinely different, longer colour number
    # (here, "1517") — not the same colour — so it must be rejected.
    if len(code) not in (len(colour_code), len(colour_code) + 2) or not code.startswith(colour_code):
        return False
    if (m.group("marker") or "") != marker:
        return False

    if twist and entry["twist"] and entry["twist"] != twist:
        return False

    for group in hint_groups:
        if not any(hint in cldescr_upper for hint in group):
            return False

    if name_hint and name_hint.upper() not in cldescr_upper:
        return False

    return True


def _best_match(candidates: list[dict[str, str]]) -> dict[str, str] | None:
    if not candidates:
        return None
    return max(candidates, key=lambda e: e["date"])


def _guess_new_colour(
    colour_code: str, colour_name: str, yarn: str
) -> tuple[str, str]:
    """
    Best-effort guess for a colour with NO dyeing history at all yet (not
    even via the broadened fallback search) — confirmed with the user that
    such codes are sometimes invented by hand for a first-time colour, and
    that a genuinely new code isn't necessarily in the DFM export we have.

    Only one case has a reliable enough mechanical pattern to predict:
    Kaschmir/Cashmere yarns consistently get a COLOREDFM starting with
    "37" followed by the colour digits exactly (confirmed ~99% consistent
    across the full Elvy history — e.g. colour "1100" -> "371100"). For
    anything else, there's no reliable pattern to invent a code from, so
    we only echo the colour's own descriptive name (if it has one) into
    CLDESCR and leave COLOREDFM blank for manual entry, per the user's
    explicit instruction.
    """
    yarn_lower = clean_text(yarn).lower()
    if colour_code and ("kaschmir" in yarn_lower or "kashmir" in yarn_lower or "cashmere" in yarn_lower):
        return f"37{colour_code}", f"EL-{colour_code}-KASCHMIR"

    if colour_name:
        return "", colour_name.upper()

    return "", ""


def lookup_dfm_color(
    articolo_delta: str,
    colour: str,
    ne: str,
    dye_type: str,
    yarn: str,
    entries: list[dict[str, str]],
) -> tuple[str, str]:
    """
    Look up (COLOREDFM, CLDESCR) for a Purchase Order row.

    articolo_delta : this row's resolved Delta raw code (e.g. "G130027S"),
                      from the Elvy Article No mapping — "" if unresolved.
    colour          : the PDF's Colore value. Handles the several raw
                      formats seen in practice — plain numbers ("440"),
                      a number + descriptive name ("009.Red Water Color",
                      "MO_0030.CELESTE"), or a letter-prefixed number
                      ("R.1372", "G.203") — see _extract_colour_code.
    ne              : the PDF's Ne value (e.g. "100/2") — its denominator
                      gives the required twist.
    dye_type, yarn  : used to require extra CLDESCR keywords (Reattivo/
                      Reactive, Tencel, Kashmir/Kaschmir) when applicable.

    Search order: same article code first, then any Elvy article as a
    fallback; within each, prefer a match whose CLDESCR also contains the
    colour's descriptive name (when it has one) before dropping that
    requirement. If nothing at all is found — this colour has never been
    dyed for Elvy before — falls back to _guess_new_colour as a last
    resort: a mechanical prediction for Kaschmir yarns, or just the
    colour's own descriptive name with the code left blank otherwise.
    """
    colour_code, colour_name, marker = _extract_colour_code(colour)
    if not colour_code:
        return "", ""
    twist = _ne_twist_digit(ne)
    hint_groups = _required_cldescr_hints(dye_type, yarn)

    articolo_c = ""
    articolo_delta = clean_text(articolo_delta)
    if articolo_delta.upper().startswith("G"):
        articolo_c = "C" + articolo_delta[1:]

    pools = [
        [e for e in entries if e["articolo"] == articolo_c] if articolo_c else [],
        entries,
    ]
    name_preferences = [colour_name, ""] if colour_name else [""]
    # Try requiring the dye/yarn keywords (e.g. "REATTIVO") first, but some
    # CLDESCR entries (fibre-special ones like Kaschmir in particular) don't
    # carry that tag even when it applies, so fall back to not requiring it.
    hint_levels = [hint_groups, []] if hint_groups else [[]]

    for pool in pools:
        if not pool:
            continue
        for hints in hint_levels:
            for name_hint in name_preferences:
                candidates = [
                    e for e in pool
                    if _matches_filters(e, colour_code, marker, twist, hints, name_hint)
                ]
                best = _best_match(candidates)
                if best:
                    return best["coloredfm"], best["cldescr"]

    return _guess_new_colour(colour_code, colour_name, yarn)
