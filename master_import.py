"""master_import.py — one directory or one file, every data source.

Lets the user select either:
1. A data folder (directory) containing individual factory exports
   (Articoli.xlsx, Copertura.xlsx, Data Ordine.xlsx, Densita' Query.xlsx,
   DFM.xlsx, Dispo Bagno.xlsx, LISTINI.XLS, LOTTI.xlsx, Magazino.xlsx,
   Produzione.xlsx, Qualita.xlsx, Uscita.xlsx, WINCOINT.xlsx).
2. A single master workbook that bundles all sheets in one file.

Routes each recognized file/sheet through the exact loader/cache pipeline
its own tab uses, distributing data to Situazione, Magazino, Biglietti,
Prezzi, and Overview tabs.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

import openpyxl

from path_manager import save_source

SHEET_CANDIDATES: dict[str, list[str]] = {
    "dfm": ["dfm"],
    "copertura": ["copertura"],
    "data_prod": ["produzione", "prod", "data prod", "dataprod"],
    "wincoint": ["wincoint"],
    "uscita": ["uscita"],
    "qualita": ["qualita", "qualità"],
    "codes": ["articoli"],
    "magazino": ["magazino", "magazzino"],
    "lotti": ["lotti"],
    "listini": ["listini", "prezzi"],
}

DIRECTORY_FILE_CANDIDATES: dict[str, list[str]] = {
    "dfm": ["dfm"],
    "copertura": ["copertura"],
    "data_prod": ["produzione", "prod", "data prod", "data_prod", "dataprod"],
    "wincoint": ["wincoint"],
    "uscita": ["uscita"],
    "qualita": ["qualita", "qualità"],
    "codes": ["articoli", "codes"],
    "magazino": ["magazino", "magazzino"],
    "lotti": ["lotti"],
    "listini": ["listini", "prezzi"],
    "densita": ["densita' query", "densita query", "densita", "density query"],
    "data_ordine": ["data ordine", "data_ordine", "dataordine", "ordine data"],
    "dispo_bagno": ["dispo bagno", "dispo-bagno", "dispo_bagno", "doispo bagno", "doispo-bagno"],
}

LABELS: dict[str, str] = {
    "dfm": "DFM",
    "copertura": "Copertura",
    "data_prod": "Produzione",
    "wincoint": "WINCOINT",
    "uscita": "Uscita",
    "qualita": "Qualita",
    "codes": "Articoli",
    "magazino": "Magazino",
    "lotti": "LOTTI",
    "listini": "LISTINI",
    "densita": "Densita' Query",
    "data_ordine": "Data Ordine",
    "dispo_bagno": "Dispo Bagno",
}


def _norm(s: str) -> str:
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


def _match_sheets(sheetnames: list[str]) -> dict[str, str]:
    """{key: actual_sheet_name} for whichever keys have a matching sheet."""
    found = {}
    normalized = {_norm(s): s for s in sheetnames}
    for key, candidates in SHEET_CANDIDATES.items():
        for cand in candidates:
            n = _norm(cand)
            if n in normalized:
                found[key] = normalized[n]
                break
    return found


def find_files_in_directory(dir_path: str | Path) -> dict[str, Path]:
    """Scan *dir_path* and map each known source key to its matching file
    Path. Real export filenames are rarely the bare candidate name (e.g.
    'Densita__Query.xlsx', 'DFM Agosto 2026.xlsx') so this matches by
    substring, not exact equality -- exact matches are preferred first
    (across every file) before falling back to substring matches, and a
    file already claimed by one key isn't offered to another."""
    dir_path = Path(dir_path)
    if not dir_path.is_dir():
        return {}

    valid_extensions = {".xlsx", ".xls", ".xlsm", ".csv"}
    files = [p for p in dir_path.iterdir() if p.is_file() and p.suffix.lower() in valid_extensions]
    file_norms = [(f, _norm(f.stem)) for f in files]

    found: dict[str, Path] = {}
    claimed: set[Path] = set()

    def try_match(exact: bool) -> None:
        for key, candidates in DIRECTORY_FILE_CANDIDATES.items():
            if key in found:
                continue
            for candidate in candidates:
                cand_norm = _norm(candidate)
                for f, f_norm in file_norms:
                    if f in claimed:
                        continue
                    hit = (f_norm == cand_norm) if exact else (cand_norm in f_norm)
                    if hit:
                        found[key] = f
                        claimed.add(f)
                        break
                if key in found:
                    break

    try_match(exact=True)
    try_match(exact=False)

    # data_ordine / dispo_bagno filenames are inherently order-specific
    # (e.g. 'MED-D-505449-2026.xlsx') so keyword matching rarely finds
    # them -- fall back to peeking at content shape for whatever files
    # are still unclaimed.
    if "data_ordine" not in found or "dispo_bagno" not in found:
        _content_match_ordine_files(files, claimed, found)

    return found


def _content_match_ordine_files(files: list[Path], claimed: set[Path], found: dict[str, Path]) -> None:
    import csv as _csv

    for f in files:
        if f in claimed:
            continue
        try:
            if f.suffix.lower() == ".csv":
                if "dispo_bagno" in found:
                    continue
                with open(f, encoding="utf-8-sig", newline="") as fh:
                    sample = fh.readline()
                if _norm("Dispo") in _norm(sample) and _norm("Articolo") in _norm(sample):
                    found["dispo_bagno"] = f
                    claimed.add(f)
                continue

            wb = openpyxl.load_workbook(f, read_only=True)
            try:
                sheetnames = set(wb.sheetnames)
                if "data_ordine" not in found and "Sheet1" in sheetnames:
                    header = {_norm(v) for v in next(wb["Sheet1"].iter_rows(values_only=True), []) if v}
                    signature = {_norm("Descrizione aggiuntiva ordine"), _norm("Cliente")}
                    el_kamal_signature = {_norm("CODICE"), _norm("Clienti"), _norm("M/C")}
                    if signature.issubset(header) or el_kamal_signature.issubset(header):
                        found["data_ordine"] = f
                        claimed.add(f)
                        continue
                if "dispo_bagno" not in found and "Doispo-Bagno" in sheetnames:
                    # already embedded in a Data Ordine file -- nothing
                    # separate to assign, so leave dispo_bagno unmatched
                    # rather than pointing it at the same file twice.
                    pass
            finally:
                wb.close()
        except Exception:
            continue


def _extract_sheet_to_temp(master_path: str, sheet_name: str) -> str:
    """Copy one sheet into its own standalone xlsx."""
    src_wb = openpyxl.load_workbook(master_path, data_only=True, read_only=True)
    try:
        src_ws = src_wb[sheet_name]
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = (sheet_name or "Sheet1")[:31]
        for row in src_ws.iter_rows(values_only=True):
            out_ws.append(row)
    finally:
        src_wb.close()
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="master_import_")
    os.close(fd)
    out_wb.save(tmp_path)
    out_wb.close()
    return tmp_path


def import_master_file(master_path: str, situazione_tab, magazino_tab, biglietti_tab=None, prezzi_tab=None) -> tuple[list[str], list[str]]:
    """Import all data from a single multi-sheet Excel file."""
    master_path = str(master_path)
    wb = openpyxl.load_workbook(master_path, read_only=True)
    sheetnames = list(wb.sheetnames)
    wb.close()

    matches = _match_sheets(sheetnames)
    loaded: list[str] = []
    skipped: list[str] = [LABELS[k] for k in SHEET_CANDIDATES if k not in matches]

    for key, sheet_name in matches.items():
        tmp_path = None
        try:
            tmp_path = _extract_sheet_to_temp(master_path, sheet_name)
            _route(key, tmp_path, master_path, situazione_tab, magazino_tab, biglietti_tab, prezzi_tab)
            loaded.append(LABELS[key])
        except Exception:
            skipped.append(LABELS[key])
        finally:
            if tmp_path:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

    normalized = {_norm(s) for s in sheetnames}
    if _norm("Entry") in normalized and _norm("Densita") in normalized:
        try:
            import biglietti_exporter
            from densita_cache import save_densita_cache
            densita_map, _errors = biglietti_exporter.load_densita_query(Path(master_path))
            if densita_map:
                save_densita_cache(master_path)
                if situazione_tab is not None and hasattr(situazione_tab, "refresh_prezzo_densita"):
                    situazione_tab.refresh_prezzo_densita()
                if biglietti_tab is not None:
                    biglietti_tab.densita_path = Path(master_path)
                    if hasattr(biglietti_tab, "_restore"):
                        biglietti_tab._restore()
                loaded.append(LABELS["densita"])
            else:
                skipped.append(LABELS["densita"])
        except Exception:
            skipped.append(LABELS["densita"])
    else:
        skipped.append(LABELS["densita"])

    return loaded, skipped


def import_master_directory(dir_path: str | Path, situazione_tab, magazino_tab, biglietti_tab=None, prezzi_tab=None) -> tuple[list[str], list[str]]:
    """Scan *dir_path* for all recognized factory files and route each to its
    respective loaders, caches, and GUI tabs.
    Returns (loaded_items, skipped_items)."""
    dir_path = Path(dir_path)
    matched = find_files_in_directory(dir_path)

    loaded: list[str] = []
    skipped: list[str] = []

    for key in DIRECTORY_FILE_CANDIDATES:
        if key not in matched:
            skipped.append(LABELS.get(key, key))
            continue

        file_path = matched[key]
        try:
            _route(key, str(file_path), str(file_path), situazione_tab, magazino_tab, biglietti_tab, prezzi_tab)
            loaded.append(f"{LABELS.get(key, key)} ({file_path.name})")
        except Exception as exc:
            skipped.append(f"{LABELS.get(key, key)} (Error: {exc})")

    return loaded, skipped


def _route(key: str, tmp_path: str, master_path: str, situazione_tab, magazino_tab, biglietti_tab=None, prezzi_tab=None) -> None:
    """Route one file/source to its proper tab, database, and cache handlers."""
    if key in ("dfm", "copertura", "data_prod", "wincoint", "uscita", "qualita"):
        if situazione_tab is not None:
            situazione_tab._handle_upload(key, tmp_path, cache_path=master_path)

    elif key == "codes":
        if situazione_tab is not None:
            situazione_tab._handle_codes_upload("codes", tmp_path, cache_path=master_path)
        else:
            import situazione_db, situazione_loaders
            from articoli_cache import save_articoli_cache
            df, _ = situazione_loaders.load_codes(tmp_path)
            if df is not None:
                situazione_db.save_codes(df)
            save_articoli_cache(master_path)
        if biglietti_tab is not None:
            biglietti_tab.articoli_path = Path(master_path)
            if hasattr(biglietti_tab, "_restore"):
                try:
                    biglietti_tab._restore()
                except Exception:
                    pass

    elif key == "listini":
        if situazione_tab is not None:
            situazione_tab._handle_listini_upload("listini", tmp_path, cache_path=master_path)
        else:
            from prezzi_cache import save_prezzi_cache
            save_prezzi_cache(master_path)
        if prezzi_tab is not None:
            try:
                if hasattr(prezzi_tab, "_load_path"):
                    prezzi_tab.after(0, lambda: prezzi_tab._load_path(master_path, save_cache=False))
                elif hasattr(prezzi_tab, "load_file"):
                    prezzi_tab.after(0, lambda: prezzi_tab.load_file(master_path))
            except Exception:
                pass
        if biglietti_tab is not None and hasattr(biglietti_tab, "_restore"):
            try:
                biglietti_tab._restore()
            except Exception:
                pass

    elif key == "magazino":
        if magazino_tab is not None:
            magazino_tab._on_upload_magazino(tmp_path, cache_path=master_path)
        if biglietti_tab is not None and hasattr(biglietti_tab, "_restore"):
            try:
                biglietti_tab._restore()
            except Exception:
                pass

    elif key == "lotti":
        if magazino_tab is not None:
            magazino_tab._on_upload_lotti(tmp_path, cache_path=master_path)

    elif key == "densita":
        from densita_cache import save_densita_cache
        save_densita_cache(master_path)
        if situazione_tab is not None and hasattr(situazione_tab, "refresh_prezzo_densita"):
            try:
                situazione_tab.refresh_prezzo_densita()
            except Exception:
                pass
        # Densita is shared by Biglietti and Ordine MED. Persist the same
        # source path and update Ordine MED's visible selection directly.
        save_source("densita", master_path)
        if biglietti_tab is not None:
            biglietti_tab.densita_path = Path(master_path)
            if hasattr(biglietti_tab, "densita_label"):
                biglietti_tab.after(0, lambda: biglietti_tab.densita_label.config(text=str(master_path), foreground="#111827"))
            if hasattr(biglietti_tab, "_restore"):
                try:
                    biglietti_tab._restore()
                except Exception:
                    pass
        ordine_med_tab = getattr(situazione_tab, "ordine_med_tab", None) if situazione_tab is not None else None
        if ordine_med_tab is not None:
            ordine_med_tab.densita_path = Path(master_path)
            if hasattr(ordine_med_tab, "_lbl_densita"):
                ordine_med_tab.after(0, lambda: ordine_med_tab._lbl_densita.config(text=str(master_path), foreground="black"))

    elif key == "data_ordine":
        save_source("data_ordine", master_path)
        if biglietti_tab is not None:
            biglietti_tab.data_path = Path(master_path)
            if hasattr(biglietti_tab, "biglietti_data_path_label"):
                biglietti_tab.biglietti_data_path_label.config(text=str(master_path), foreground="#111827")
            if hasattr(biglietti_tab, "_save_prefs"):
                biglietti_tab._save_prefs(biglietti_data_path=str(master_path))

    elif key == "dispo_bagno":
        save_source("dispo_bagno", master_path)
        if biglietti_tab is not None:
            biglietti_tab.dispo_path = Path(master_path)
            if hasattr(biglietti_tab, "biglietti_dispo_path_label"):
                biglietti_tab.biglietti_dispo_path_label.config(text=str(master_path), foreground="#111827")
            if hasattr(biglietti_tab, "_save_prefs"):
                biglietti_tab._save_prefs(biglietti_dispo_path=str(master_path))

